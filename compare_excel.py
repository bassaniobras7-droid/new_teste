import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border
from openpyxl.utils import get_column_letter, column_index_from_string
import json

ANTIGO = r"C:/Users/Wesllen.Santana/Downloads/bassani/Relatorios_Com_Formulas.xlsx"
NOVO   = r"C:/Users/Wesllen.Santana/Downloads/bassani/Relatorios_Com_Formulas. novo layoutxlsx.xlsx"

wb_a = openpyxl.load_workbook(ANTIGO)
wb_n = openpyxl.load_workbook(NOVO)

print("=== ABAS ===")
print(f"ANTIGO: {wb_a.sheetnames}")
print(f"NOVO:   {wb_n.sheetnames}")
print()

def rgb(color_obj):
    if color_obj is None:
        return None
    try:
        t = color_obj.type
        if t == 'rgb':
            val = color_obj.rgb
            return val if val else None
        if t == 'theme':
            return f"THEME:{color_obj.theme}"
        if t == 'indexed':
            return f"INDEXED:{color_obj.indexed}"
    except:
        pass
    return None

def fill_info(cell):
    f = cell.fill
    if f is None or f.fill_type is None or f.fill_type == 'none':
        return None
    return {
        'type': f.fill_type,
        'fg': rgb(f.fgColor) if f.fgColor else None,
        'bg': rgb(f.bgColor) if f.bgColor else None,
    }

def font_info(cell):
    f = cell.font
    if f is None:
        return None
    return {
        'name': f.name,
        'size': f.size,
        'bold': f.bold,
        'italic': f.italic,
        'color': rgb(f.color) if f.color else None,
        'underline': f.underline,
    }

def align_info(cell):
    a = cell.alignment
    if a is None:
        return None
    return {
        'horizontal': a.horizontal,
        'vertical': a.vertical,
        'wrap': a.wrap_text,
    }

def border_info(cell):
    b = cell.border
    if b is None:
        return None
    def side(s):
        if s is None:
            return None
        return {'style': s.border_style, 'color': rgb(s.color) if s.color else None}
    return {
        'left': side(b.left),
        'right': side(b.right),
        'top': side(b.top),
        'bottom': side(b.bottom),
    }

def cell_value(cell):
    v = cell.value
    if v is None:
        return None
    return str(v)

def sheet_summary(ws):
    info = {}
    info['dims'] = ws.dimensions
    info['max_row'] = ws.max_row
    info['max_col'] = ws.max_column

    # Column widths
    col_widths = {}
    for col_letter, col_dim in ws.column_dimensions.items():
        col_widths[col_letter] = col_dim.width
    info['col_widths'] = col_widths

    # Row heights
    row_heights = {}
    for row_idx, row_dim in ws.row_dimensions.items():
        if row_dim.height is not None:
            row_heights[row_idx] = row_dim.height
    info['row_heights'] = row_heights

    # Merged cells
    info['merged_cells'] = [str(m) for m in ws.merged_cells.ranges]

    # Page setup
    ps = ws.page_setup
    info['page_setup'] = {
        'orientation': ps.orientation,
        'paper_size': ps.paperSize,
        'fit_to_page': ps.fitToPage,
        'fit_to_width': ps.fitToWidth,
        'fit_to_height': ps.fitToHeight,
        'scale': ps.scale,
    }

    # Print area
    info['print_area'] = ws.print_area

    # Freeze panes
    info['freeze_panes'] = str(ws.freeze_panes) if ws.freeze_panes else None

    # Cells
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None or fill_info(cell) or font_info(cell):
                coord = cell.coordinate
                cells[coord] = {
                    'value': cell_value(cell),
                    'data_type': cell.data_type,
                    'number_format': cell.number_format,
                    'fill': fill_info(cell),
                    'font': font_info(cell),
                    'align': align_info(cell),
                    'border': border_info(cell),
                }
    info['cells'] = cells
    return info

# Processar cada aba
sheets_a = {name: sheet_summary(wb_a[name]) for name in wb_a.sheetnames}
sheets_n = {name: sheet_summary(wb_n[name]) for name in wb_n.sheetnames}

all_names = list(dict.fromkeys(list(wb_a.sheetnames) + list(wb_n.sheetnames)))

print("=== COMPARAÇÃO ABA A ABA ===\n")

for sheet_name in all_names:
    print(f"\n{'='*70}")
    print(f"ABA: {sheet_name}")
    print(f"{'='*70}")

    if sheet_name not in sheets_a:
        print(f"  [NOVO] Aba existe apenas no arquivo NOVO")
        continue
    if sheet_name not in sheets_n:
        print(f"  [ANTIGO] Aba existe apenas no arquivo ANTIGO")
        continue

    sa = sheets_a[sheet_name]
    sn = sheets_n[sheet_name]

    # 1. Dimensoes
    if sa['max_row'] != sn['max_row'] or sa['max_col'] != sn['max_col']:
        print(f"\n[DIMENSOES]")
        print(f"  ANTIGO: {sa['max_row']} linhas x {sa['max_col']} colunas")
        print(f"  NOVO:   {sn['max_row']} linhas x {sn['max_col']} colunas")

    # 2. Larguras de colunas
    all_cols = set(list(sa['col_widths'].keys()) + list(sn['col_widths'].keys()))
    col_diffs = {}
    for col in sorted(all_cols):
        wa = sa['col_widths'].get(col)
        wn = sn['col_widths'].get(col)
        if wa != wn:
            col_diffs[col] = (wa, wn)
    if col_diffs:
        print(f"\n[LARGURAS DE COLUNAS]")
        for col, (wa, wn) in sorted(col_diffs.items()):
            print(f"  Coluna {col}: ANTIGO={wa}  -->  NOVO={wn}")

    # 3. Alturas de linhas
    all_rows = set(list(sa['row_heights'].keys()) + list(sn['row_heights'].keys()))
    row_diffs = {}
    for row in sorted(all_rows):
        ha = sa['row_heights'].get(row)
        hn = sn['row_heights'].get(row)
        if ha != hn:
            row_diffs[row] = (ha, hn)
    if row_diffs:
        print(f"\n[ALTURAS DE LINHAS]")
        for row, (ha, hn) in sorted(row_diffs.items()):
            print(f"  Linha {row}: ANTIGO={ha}  -->  NOVO={hn}")

    # 4. Merged cells
    ma = set(sa['merged_cells'])
    mn = set(sn['merged_cells'])
    if ma != mn:
        print(f"\n[MESCLAGENS DE CELULAS]")
        only_a = ma - mn
        only_n = mn - ma
        if only_a:
            print(f"  Removidas do NOVO (existem apenas no ANTIGO): {sorted(only_a)}")
        if only_n:
            print(f"  Adicionadas no NOVO (nao existem no ANTIGO): {sorted(only_n)}")

    # 5. Page setup
    psa = sa['page_setup']
    psn = sn['page_setup']
    ps_diffs = {}
    for k in set(list(psa.keys()) + list(psn.keys())):
        if psa.get(k) != psn.get(k):
            ps_diffs[k] = (psa.get(k), psn.get(k))
    if ps_diffs:
        print(f"\n[CONFIGURACOES DE PAGINA]")
        for k, (va, vn) in ps_diffs.items():
            print(f"  {k}: ANTIGO={va}  -->  NOVO={vn}")

    # Print area
    if sa['print_area'] != sn['print_area']:
        print(f"\n[AREA DE IMPRESSAO]")
        print(f"  ANTIGO: {sa['print_area']}")
        print(f"  NOVO:   {sn['print_area']}")

    # Freeze panes
    if sa['freeze_panes'] != sn['freeze_panes']:
        print(f"\n[CONGELAR PAINEIS]")
        print(f"  ANTIGO: {sa['freeze_panes']}")
        print(f"  NOVO:   {sn['freeze_panes']}")

    # 6. Celulas
    all_coords = set(list(sa['cells'].keys()) + list(sn['cells'].keys()))
    cell_diffs = []
    for coord in sorted(all_coords, key=lambda c: (int(''.join(filter(str.isdigit, c))), c)):
        ca = sa['cells'].get(coord, {})
        cn = sn['cells'].get(coord, {})
        diffs = []

        # Value
        va = ca.get('value')
        vn = cn.get('value')
        if va != vn:
            diffs.append(f"valor: '{va}' -> '{vn}'")

        # Number format
        nfa = ca.get('number_format')
        nfn = cn.get('number_format')
        if nfa != nfn:
            diffs.append(f"formato: '{nfa}' -> '{nfn}'")

        # Fill
        fa = ca.get('fill')
        fn = cn.get('fill')
        if fa != fn:
            diffs.append(f"fill: {fa} -> {fn}")

        # Font
        fta = ca.get('font')
        ftn = cn.get('font')
        if fta != ftn:
            font_d = []
            for k in set(list((fta or {}).keys()) + list((ftn or {}).keys())):
                va2 = (fta or {}).get(k)
                vn2 = (ftn or {}).get(k)
                if va2 != vn2:
                    font_d.append(f"{k}: {va2}->{vn2}")
            if font_d:
                diffs.append(f"font: [{', '.join(font_d)}]")

        # Align
        ala = ca.get('align')
        aln = cn.get('align')
        if ala != aln:
            al_d = []
            for k in set(list((ala or {}).keys()) + list((aln or {}).keys())):
                va2 = (ala or {}).get(k)
                vn2 = (aln or {}).get(k)
                if va2 != vn2:
                    al_d.append(f"{k}: {va2}->{vn2}")
            if al_d:
                diffs.append(f"align: [{', '.join(al_d)}]")

        # Border
        bra = ca.get('border')
        brn = cn.get('border')
        if bra != brn:
            br_d = []
            for side in ['left','right','top','bottom']:
                sa2 = (bra or {}).get(side)
                sn2 = (brn or {}).get(side)
                if sa2 != sn2:
                    br_d.append(f"{side}: {sa2}->{sn2}")
            if br_d:
                diffs.append(f"border: [{', '.join(br_d)}]")

        if diffs:
            cell_diffs.append((coord, diffs))

    if cell_diffs:
        print(f"\n[CELULAS COM DIFERENCAS] ({len(cell_diffs)} celulas)")
        for coord, diffs in cell_diffs:
            print(f"  {coord}:")
            for d in diffs:
                print(f"    - {d}")

print("\n\n=== FIM DA COMPARACAO ===")
