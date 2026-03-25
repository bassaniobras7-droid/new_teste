import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
from src.utils import natural_sort_key
from src.lix_j_logic import _lix_j_should_include, _lix_j_extract_factor_cm, _lix_j_extract_factor_m
from src.aspg_logic import build_aspg_formula_parts
from src.lp_tub_logic import build_lptub_formula

# ==============================================================================
# FUNÇÕES AUXILIARES DE FORMATAÇÃO
# ==============================================================================

def apply_borders_to_range(sheet, min_row, min_col, max_row, max_col):
    thin_border = Border(left=Side(border_style="thin", color="000000"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
    for row in sheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = thin_border

def format_empty_row(sheet, row_idx, regular_font):
    sheet.row_dimensions[row_idx].height = 5
    sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    # Add a fill for visibility
    gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    sheet.cell(row=row_idx, column=1).fill = gray_fill

# ==============================================================================
# GERAÇÃO DO ARQUIVO EXCEL
# ==============================================================================

def write_excel_with_formulas(summary_normal, summary_distratado, client_normal, client_distratado, price_data, filename):
    workbook = openpyxl.Workbook()
    bold_font = Font(name='Verdana', bold=True, size=8)
    regular_font = Font(name='Verdana', size=8)
    header_fill = PatternFill(start_color="99ccff", end_color="99ccff", fill_type="solid")
    currency_format = '#,##0.00'
    accounting_format = '_("R$"* #,##0.00_);_("R$"* -#,##0.00_);_("R$"* "-"??_);_(@_)'

    ws_resumo = workbook.active
    ws_resumo.title = 'Resumo'
    write_summary_sheet(ws_resumo, summary_normal, summary_distratado, price_data, bold_font, header_fill, currency_format, regular_font)
    
    ws_aditivos_distrato = workbook.create_sheet(title='Aditivos x Distrato')
    ws_aditivos_distrato.sheet_properties.tabColor = "00B050"
    write_aditivos_distrato_sheet(ws_aditivos_distrato, summary_normal, summary_distratado, price_data, bold_font, regular_font, header_fill, currency_format, accounting_format, None, None, client_normal, client_distratado)

    ws_cliente = workbook.create_sheet(title='Cliente')
    write_client_sheet(ws_cliente, client_normal, client_distratado, price_data, bold_font, header_fill, currency_format, regular_font)
    
    try:
        workbook.save(filename)
        print(f"Arquivo '{filename}' gerado com sucesso.")
    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}")

def _write_optional_items_section(sheet, price_data, start_row, bold_font, header_fill, currency_format, regular_font, parede_metragem_cells, lix_j_formula_parts, aspg_formula_parts):
    COLS = {'TIPO': 1, 'DESC': 2, 'METRAGEM': 3, 'UN': 4, 'VALOR_UNIT': 5, 'VALOR_TOTAL': 6, 'CUSTO_MO_UNIT': 7, 'CUSTO_MO_TOTAL': 8}
    current_row = start_row
    aspg_metragem_coord = None

    # Header Row
    header_row_data = [('Tipo R. Bassani', COLS['TIPO']), ('Itens Complementares Opcionais', COLS['DESC']), ('Metragem', COLS['METRAGEM']), ('Un', COLS['UN']),
                       ('Valor do Material + MO', COLS['VALOR_UNIT']), ('Valor Total', COLS['VALOR_TOTAL']), ('Custo MO', COLS['CUSTO_MO_UNIT']), ('Valor Total MO', COLS['CUSTO_MO_TOTAL'])]
    for val, col_idx in header_row_data:
        cell = sheet.cell(row=current_row, column=col_idx, value=val)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if val in ['Custo MO', 'Valor Total MO']:
            cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8)
        else:
            cell.fill = header_fill
    apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['CUSTO_MO_TOTAL'])
    current_row += 1
    section_start_row = current_row

    # --- LIX-J ---
    lix_j_row_idx = current_row
    price_info_j = price_data.get('LIX-J', {'Valor': 0, 'Un': '', 'Custo MO': 0, 'Descricao': ''})
    
    lix_j_tipo_cell = sheet.cell(row=lix_j_row_idx, column=COLS['TIPO'], value='LIX-J')
    lix_j_tipo_cell.alignment = Alignment(horizontal='center', vertical='center')
    lix_j_tipo_cell.font = regular_font

    desc_cell_j = sheet.cell(row=lix_j_row_idx, column=COLS['DESC'], value=price_info_j.get('Descricao', ''))
    desc_cell_j.alignment = Alignment(wrap_text=True, vertical='center')
    desc_cell_j.font = regular_font

    metragem_cell_j = sheet.cell(row=lix_j_row_idx, column=COLS['METRAGEM'])
    if parede_metragem_cells:
        metragem_cell_j.value = f"={'+'.join(parede_metragem_cells)}"
    else:
        metragem_cell_j.value = 0
    metragem_cell_j.number_format = currency_format
    metragem_cell_j.font = regular_font
    metragem_cell_j.alignment = Alignment(vertical='center')

    un_cell_j = sheet.cell(row=lix_j_row_idx, column=COLS['UN'], value=price_info_j.get('Un'))
    un_cell_j.font = regular_font
    un_cell_j.alignment = Alignment(horizontal='center', vertical='center')

    valor_unit_cell_j = sheet.cell(row=lix_j_row_idx, column=COLS['VALOR_UNIT'], value=price_info_j.get('Valor'))
    valor_unit_cell_j.number_format = currency_format
    valor_unit_cell_j.font = regular_font
    valor_unit_cell_j.alignment = Alignment(vertical='center')

    custo_mo_unit_cell_j = sheet.cell(row=lix_j_row_idx, column=COLS['CUSTO_MO_UNIT'], value=price_info_j.get('Custo MO'))
    custo_mo_unit_cell_j.number_format = currency_format
    custo_mo_unit_cell_j.font = regular_font
    custo_mo_unit_cell_j.alignment = Alignment(vertical='center')
    
    m_cell_j, v_cell_j, mo_cell_j = f"{get_column_letter(COLS['METRAGEM'])}{lix_j_row_idx}", f"{get_column_letter(COLS['VALOR_UNIT'])}{lix_j_row_idx}", f"{get_column_letter(COLS['CUSTO_MO_UNIT'])}{lix_j_row_idx}"
    sheet.cell(row=lix_j_row_idx, column=COLS['VALOR_TOTAL'], value=f"={m_cell_j}*{v_cell_j}").number_format = currency_format
    sheet.cell(row=lix_j_row_idx, column=COLS['CUSTO_MO_TOTAL'], value=f"={m_cell_j}*{mo_cell_j}").number_format = currency_format
    
    fill_c0c0c0 = PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")
    sheet.cell(row=lix_j_row_idx, column=COLS['CUSTO_MO_UNIT']).fill = fill_c0c0c0
    sheet.cell(row=lix_j_row_idx, column=COLS['CUSTO_MO_TOTAL']).fill = fill_c0c0c0
    apply_borders_to_range(sheet, lix_j_row_idx, COLS['TIPO'], lix_j_row_idx, COLS['CUSTO_MO_TOTAL'])
    current_row += 1

    # --- LIX-J’ ---
    lix_j_prime_row_idx = current_row
    price_info_j_prime = price_data.get('LIX-J’', {'Valor': 0, 'Un': '', 'Custo MO': 0, 'Descricao': ''})
    
    sheet.cell(row=lix_j_prime_row_idx, column=COLS['TIPO'], value=None) # Empty because it will be merged

    desc_cell_jp = sheet.cell(row=lix_j_prime_row_idx, column=COLS['DESC'], value=price_info_j_prime.get('Descricao', ''))
    desc_cell_jp.alignment = Alignment(wrap_text=True, vertical='center')
    desc_cell_jp.font = regular_font

    metragem_cell_jp = sheet.cell(row=lix_j_prime_row_idx, column=COLS['METRAGEM'])
    if lix_j_formula_parts:
        metragem_cell_jp.value = f"={'+'.join(lix_j_formula_parts)}"
    else:
        metragem_cell_jp.value = 0
    metragem_cell_jp.number_format = currency_format
    metragem_cell_jp.font = regular_font
    metragem_cell_jp.alignment = Alignment(vertical='center')

    un_cell_jp = sheet.cell(row=lix_j_prime_row_idx, column=COLS['UN'], value=price_info_j_prime.get('Un'))
    un_cell_jp.font = regular_font
    un_cell_jp.alignment = Alignment(horizontal='center', vertical='center')

    valor_unit_cell_jp = sheet.cell(row=lix_j_prime_row_idx, column=COLS['VALOR_UNIT'], value=price_info_j_prime.get('Valor'))
    valor_unit_cell_jp.number_format = currency_format
    valor_unit_cell_jp.font = regular_font
    valor_unit_cell_jp.alignment = Alignment(vertical='center')

    custo_mo_unit_cell_jp = sheet.cell(row=lix_j_prime_row_idx, column=COLS['CUSTO_MO_UNIT'], value=price_info_j_prime.get('Custo MO'))
    custo_mo_unit_cell_jp.number_format = currency_format
    custo_mo_unit_cell_jp.font = regular_font
    custo_mo_unit_cell_jp.alignment = Alignment(vertical='center')

    m_cell_jp, v_cell_jp, mo_cell_jp = f"{get_column_letter(COLS['METRAGEM'])}{lix_j_prime_row_idx}", f"{get_column_letter(COLS['VALOR_UNIT'])}{lix_j_prime_row_idx}", f"{get_column_letter(COLS['CUSTO_MO_UNIT'])}{lix_j_prime_row_idx}"
    sheet.cell(row=lix_j_prime_row_idx, column=COLS['VALOR_TOTAL'], value=f"={m_cell_jp}*{v_cell_jp}").number_format = currency_format
    sheet.cell(row=lix_j_prime_row_idx, column=COLS['CUSTO_MO_TOTAL'], value=f"={m_cell_jp}*{mo_cell_jp}").number_format = currency_format

    sheet.cell(row=lix_j_prime_row_idx, column=COLS['CUSTO_MO_UNIT']).fill = fill_c0c0c0
    sheet.cell(row=lix_j_prime_row_idx, column=COLS['CUSTO_MO_TOTAL']).fill = fill_c0c0c0
    apply_borders_to_range(sheet, lix_j_prime_row_idx, COLS['TIPO'], lix_j_prime_row_idx, COLS['CUSTO_MO_TOTAL'])
    current_row += 1

    # Merge LIX-J cells
    sheet.merge_cells(start_row=lix_j_row_idx, start_column=COLS['TIPO'], end_row=lix_j_prime_row_idx, end_column=COLS['TIPO'])
    sheet.cell(row=lix_j_row_idx, column=COLS['TIPO']).alignment = Alignment(horizontal='center', vertical='center')

    # --- Other items ---
    for tipo_code in ['ASP-G', 'LP-TUB', 'VIT']:
        price_info = price_data.get(tipo_code, {'Valor': 0, 'Un': '', 'Custo MO': 0, 'Descricao': ''})
        
        tipo_cell = sheet.cell(row=current_row, column=COLS['TIPO'], value=tipo_code)
        tipo_cell.alignment = Alignment(horizontal='center', vertical='center')
        tipo_cell.font = regular_font

        desc_cell = sheet.cell(row=current_row, column=COLS['DESC'], value=price_info.get('Descricao', ''))
        desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
        desc_cell.font = regular_font

        metragem_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'])
        if tipo_code == 'ASP-G':
            if aspg_formula_parts:
                metragem_cell.value = f"={'+'.join(aspg_formula_parts)}"
            else:
                metragem_cell.value = 0
            aspg_metragem_coord = metragem_cell.coordinate
        elif tipo_code == 'LP-TUB':
            metragem_cell.value = build_lptub_formula(aspg_metragem_coord)
        else:
            metragem_cell.value = 0
        
        metragem_cell.number_format = currency_format
        metragem_cell.font = regular_font
        metragem_cell.alignment = Alignment(vertical='center')

        un_cell = sheet.cell(row=current_row, column=COLS['UN'], value=price_info.get('Un'))
        un_cell.font = regular_font
        un_cell.alignment = Alignment(horizontal='center', vertical='center')

        valor_unit_cell = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=price_info.get('Valor'))
        valor_unit_cell.number_format = currency_format
        valor_unit_cell.font = regular_font
        valor_unit_cell.alignment = Alignment(vertical='center')

        custo_mo_unit_cell = sheet.cell(row=current_row, column=COLS['CUSTO_MO_UNIT'], value=price_info.get('Custo MO'))
        custo_mo_unit_cell.number_format = currency_format
        custo_mo_unit_cell.font = regular_font
        custo_mo_unit_cell.alignment = Alignment(vertical='center')

        m_cell, v_cell, mo_cell = f"{get_column_letter(COLS['METRAGEM'])}{current_row}", f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}", f"{get_column_letter(COLS['CUSTO_MO_UNIT'])}{current_row}"
        sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"={m_cell}*{v_cell}").number_format = currency_format
        sheet.cell(row=current_row, column=COLS['CUSTO_MO_TOTAL'], value=f"={m_cell}*{mo_cell}").number_format = currency_format

        sheet.cell(row=current_row, column=COLS['CUSTO_MO_UNIT']).fill = fill_c0c0c0
        sheet.cell(row=current_row, column=COLS['CUSTO_MO_TOTAL']).fill = fill_c0c0c0
        apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['CUSTO_MO_TOTAL'])
        current_row += 1

    # Subtotal Row
    subtotal_row = current_row
    subtotal_cell = sheet.cell(row=subtotal_row, column=COLS['TIPO'], value='SUB-TOTAL')
    subtotal_cell.font = bold_font
    subtotal_fill = PatternFill(start_color="ffff99", end_color="ffff99", fill_type="solid")
    subtotal_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=4)

    for col_idx in range(1, 5):
        sheet.cell(row=subtotal_row, column=col_idx).fill = subtotal_fill

    sheet.merge_cells(start_row=subtotal_row, start_column=COLS['VALOR_UNIT'], end_row=subtotal_row, end_column=COLS['VALOR_TOTAL'])
    total_col = get_column_letter(COLS['VALOR_TOTAL'])
    formula_total = f"=SUM({total_col}{section_start_row}:{total_col}{current_row - 1})"
    merged_value_cell = sheet.cell(row=subtotal_row, column=COLS['VALOR_UNIT'], value=formula_total)
    merged_value_cell.number_format = currency_format
    merged_value_cell.alignment = Alignment(horizontal='right', vertical='center')
    merged_value_cell.fill = subtotal_fill
    merged_value_cell.font = bold_font

    mo_col = get_column_letter(COLS['CUSTO_MO_TOTAL'])
    formula_mo = f"=SUM({mo_col}{section_start_row}:{mo_col}{current_row - 1})"
    sheet.merge_cells(start_row=subtotal_row, start_column=COLS['CUSTO_MO_UNIT'], end_row=subtotal_row, end_column=COLS['CUSTO_MO_TOTAL'])
    merged_mo_cell = sheet.cell(row=subtotal_row, column=COLS['CUSTO_MO_UNIT'], value=formula_mo)
    merged_mo_cell.number_format = currency_format
    merged_mo_cell.alignment = Alignment(horizontal='right', vertical='center')
    merged_mo_cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
    merged_mo_cell.font = Font(name='Verdana', color="FFFFFF", size=8)

    apply_borders_to_range(sheet, subtotal_row, COLS['TIPO'], subtotal_row, COLS['CUSTO_MO_TOTAL'])
    current_row += 1
    
    return current_row, merged_value_cell.coordinate, merged_mo_cell.coordinate

def _write_summary_section(sheet, summary_data, price_data, title, start_row, bold_font, header_fill, currency_format, regular_font):
    COLS = {'TIPO': 1, 'DESC': 2, 'METRAGEM': 3, 'UN': 4, 'VALOR_UNIT': 5, 'VALOR_TOTAL': 6, 'CUSTO_MO_UNIT': 7, 'CUSTO_MO_TOTAL': 8}
    
    title_cell = sheet.cell(row=start_row, column=1, value=title)
    title_cell.font = bold_font
    
    if title == 'RESUMO':
        title_cell.fill = PatternFill(start_color="b3a2c7", end_color="b3a2c7", fill_type="solid")
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        subtitle_row = start_row + 1
        subtitle_cell = sheet.cell(row=subtitle_row, column=1, value="SERVIÇOS ADICIONAIS (NOVO LAYOUT)")
        subtitle_cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8) # White font, keep bold
        subtitle_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        sheet.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=6) # Merge A:F
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Ensure the entire merged area has the fill color
        for col_idx in range(1, 7): # Columns A to F
            sheet.cell(row=subtitle_row, column=col_idx).fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")

        current_row = start_row + 2
    elif title == 'SERVIÇOS DISTRATADOS':
        title_cell.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6) # Merge A:F
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8) # White font, keep bold, and name Verdana
        # Ensure the entire merged area has the fill color
        for col_idx in range(1, 7): # Columns A to F
            sheet.cell(row=start_row, column=col_idx).fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        current_row = start_row + 1
    else:
        current_row = start_row + 1

    subtotal_valor_cells, subtotal_mo_cells = [], []
    parede_metragem_cells = []
    lix_j_formula_parts = []
    parede_cells_map = {}

    def summary_sort_key_with_tp_priority(item):
        tipo_code = item[0]
        is_car = tipo_code.endswith('-CAR')
        base_code = tipo_code.replace('-CAR', '')
        tp_priority = 0 if base_code.startswith('TP') else 1
        natural_base_sort = (tp_priority, natural_sort_key(base_code))
        car_suffix_priority = 1 if is_car else 0
        return (natural_base_sort, car_suffix_priority)

    for categoria, nome_categoria in [('Forros', 'Forros'), ('Paredes', 'Paredes e Revestimentos'), ('Guias e Montantes', 'Guias e Montantes'), ('Isolamento', 'Isolamento Acústico')]:
        items_cat = {k: v for k, v in summary_data.items() if v.get('Categoria') == categoria}
        if not items_cat: continue

        header_row_start = current_row

        header_row = [ ('Tipo R. Bassani', COLS['TIPO']), (nome_categoria, COLS['DESC']), ('Metragem', COLS['METRAGEM']), ('Un', COLS['UN']),
                       ('Valor do Material + MO', COLS['VALOR_UNIT']), ('Valor Total', COLS['VALOR_TOTAL']), ('Custo MO', COLS['CUSTO_MO_UNIT']), ('Valor Total MO', COLS['CUSTO_MO_TOTAL']) ]
        for val, col in header_row:
            cell = sheet.cell(row=current_row, column=col, value=val)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) # Add wrap_text

            if val in ['Custo MO', 'Valor Total MO']:
                cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8) # White font, keep bold
            else:
                cell.fill = header_fill # Existing header fill
        
        # Apply borders to the header row
        apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['CUSTO_MO_TOTAL'])

        current_row += 1
        section_start_row = current_row

        if categoria == 'Paredes':
            sorted_items = sorted(items_cat.items(), key=summary_sort_key_with_tp_priority)
        elif categoria == 'Isolamento':
            sorted_items = sorted(items_cat.items(), key=lambda item: (0 if item[0].startswith('TP') else 1, natural_sort_key(item[0])))
        else:
            sorted_items = sorted(items_cat.items(), key=lambda item: natural_sort_key(item[0]))

        for tipo_code, data in sorted_items:
            price_info = price_data.get(tipo_code, {'Valor': 0, 'Un': '', 'Custo MO': 0})
            tipo_cell = sheet.cell(row=current_row, column=COLS['TIPO'], value=tipo_code)
            tipo_cell.alignment = Alignment(horizontal='center', vertical='center')
            tipo_cell.font = regular_font
            
            desc_cell = sheet.cell(row=current_row, column=COLS['DESC'], value=data['Descricao'])
            desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
            desc_cell.font = regular_font
            
            metragem_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'], value=data['Quantidade'])
            metragem_cell.number_format = currency_format
            metragem_cell.font = regular_font
            metragem_cell.alignment = Alignment(vertical='center')

            if _lix_j_should_include(data['Descricao']):
                factor = _lix_j_extract_factor_cm(data['Descricao'])
                if factor is None:
                    factor = _lix_j_extract_factor_m(data['Descricao'])
                if factor is None and "fechamento vertical" in data['Descricao'].lower():
                    factor = 1
                
                if factor:
                    lix_j_formula_parts.append(f"({metragem_cell.coordinate}*{factor})")
                else:
                    lix_j_formula_parts.append(metragem_cell.coordinate)


            if categoria == 'Paredes' and data['Descricao'].startswith('Parede'):
                parede_metragem_cells.append(metragem_cell.coordinate)
                parede_cells_map[tipo_code] = metragem_cell.coordinate

            
            un_cell = sheet.cell(row=current_row, column=COLS['UN'], value=price_info['Un'])
            un_cell.font = regular_font
            un_cell.alignment = Alignment(horizontal='center', vertical='center')
            valor_unit_cell = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=price_info['Valor'])
            valor_unit_cell.number_format = currency_format
            valor_unit_cell.font = regular_font
            valor_unit_cell.alignment = Alignment(vertical='center')
            custo_mo_unit_cell = sheet.cell(row=current_row, column=COLS['CUSTO_MO_UNIT'], value=price_info['Custo MO'])
            custo_mo_unit_cell.number_format = currency_format
            custo_mo_unit_cell.font = regular_font
            custo_mo_unit_cell.alignment = Alignment(vertical='center')
            
            m_cell, v_cell, mo_cell = f"{get_column_letter(COLS['METRAGEM'])}{current_row}", f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}", f"{get_column_letter(COLS['CUSTO_MO_UNIT'])}{current_row}"
            valor_total_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"={m_cell}*{v_cell}")
            valor_total_cell.number_format = currency_format
            valor_total_cell.font = regular_font
            valor_total_cell.alignment = Alignment(vertical='center')
            custo_mo_total_cell = sheet.cell(row=current_row, column=COLS['CUSTO_MO_TOTAL'], value=f"={m_cell}*{mo_cell}")
            custo_mo_total_cell.number_format = currency_format
            custo_mo_total_cell.font = regular_font
            custo_mo_total_cell.alignment = Alignment(vertical='center')
            
            # Apply fill to G and H columns for data rows
            fill_c0c0c0 = PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")
            sheet.cell(row=current_row, column=COLS['CUSTO_MO_UNIT']).fill = fill_c0c0c0
            sheet.cell(row=current_row, column=COLS['CUSTO_MO_TOTAL']).fill = fill_c0c0c0
            
            # Apply borders to the data row
            apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['CUSTO_MO_TOTAL'])
            
            current_row += 1
            
        subtotal_row = current_row
        subtotal_cell = sheet.cell(row=subtotal_row, column=COLS['TIPO'], value='SUB-TOTAL')
        subtotal_cell.font = bold_font
        
        subtotal_fill = PatternFill(start_color="ffff99", end_color="ffff99", fill_type="solid")
        subtotal_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=4)

        for col_idx in range(1, 5):
            sheet.cell(row=subtotal_row, column=col_idx).fill = subtotal_fill

        # Apply new formatting for E:F
        sheet.merge_cells(start_row=subtotal_row, start_column=COLS['VALOR_UNIT'], end_row=subtotal_row, end_column=COLS['VALOR_TOTAL'])
        
        # Define total_col and mo_col here
        total_col = get_column_letter(COLS['VALOR_TOTAL'])
        mo_col = get_column_letter(COLS['CUSTO_MO_TOTAL'])

        # Apply formula to the merged cell (column E)
        formula_total = f"=SUM({total_col}{section_start_row}:{total_col}{current_row - 1})"
        merged_value_cell = sheet.cell(row=subtotal_row, column=COLS['VALOR_UNIT'], value=formula_total)
        merged_value_cell.number_format = currency_format
        merged_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_value_cell.fill = subtotal_fill # Also apply fill to the merged value cell
        merged_value_cell.font = bold_font

        # Apply formula to CUSTO_MO_TOTAL (column H)
        formula_mo = f"=SUM({mo_col}{section_start_row}:{mo_col}{current_row - 1})"
        
        # Merge G and H
        sheet.merge_cells(start_row=subtotal_row, start_column=COLS['CUSTO_MO_UNIT'], end_row=subtotal_row, end_column=COLS['CUSTO_MO_TOTAL'])
        merged_mo_cell = sheet.cell(row=subtotal_row, column=COLS['CUSTO_MO_UNIT'], value=formula_mo)
        merged_mo_cell.number_format = currency_format
        merged_mo_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_mo_cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        merged_mo_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # Added font color # Corrected color

        # Apply borders to the subtotal row
        apply_borders_to_range(sheet, subtotal_row, COLS['TIPO'], subtotal_row, COLS['CUSTO_MO_TOTAL'])
        
        subtotal_valor_cells.append(f"{get_column_letter(COLS['VALOR_UNIT'])}{subtotal_row}")
        subtotal_mo_cells.append(f"{get_column_letter(COLS['CUSTO_MO_UNIT'])}{subtotal_row}") # Update to reference column G
        current_row += 1
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1

    if title == 'RESUMO': # This is where the optional items table will be added
        aspg_formula_parts = build_aspg_formula_parts(summary_data, parede_cells_map)
        current_row, optional_subtotal_val, optional_subtotal_mo = _write_optional_items_section(
            sheet, price_data, current_row, bold_font, header_fill, currency_format, regular_font, parede_metragem_cells, lix_j_formula_parts, aspg_formula_parts
        )
        subtotal_valor_cells.append(optional_subtotal_val)
        subtotal_mo_cells.append(optional_subtotal_mo)
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1

    total_row = current_row
    total_label_cell = sheet.cell(row=total_row, column=COLS['TIPO'], value=f'TOTAL {title}')
    total_label_cell.font = bold_font
    total_formula = f"=+{'+'.join(subtotal_valor_cells)}" if subtotal_valor_cells else "0"
    mo_formula = f"=+{'+'.join(subtotal_mo_cells)}" if subtotal_mo_cells else "0"
    
    # Initialize total_val_cell and total_mo_cell to ensure they are always defined
    total_val_cell_coord = ""
    total_mo_cell_coord = ""

    if title == 'RESUMO':
        total_label_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Apply the fill to the merged cells as well
        for col_idx in range(1, 5): # Columns A to D
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")

        # Apply new formatting for E:F
        sheet.merge_cells(start_row=total_row, start_column=COLS['VALOR_UNIT'], end_row=total_row, end_column=COLS['VALOR_TOTAL'])
        merged_total_val_cell = sheet.cell(row=total_row, column=COLS['VALOR_UNIT'], value=total_formula) # Place formula in E
        merged_total_val_cell.number_format = currency_format
        merged_total_val_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_total_val_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        merged_total_val_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # Set font color to white
        for col_idx in range(COLS['VALOR_UNIT'], COLS['VALOR_TOTAL'] + 1):
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        
        # Assign coordinate
        total_val_cell_coord = merged_total_val_cell.coordinate

        # Apply new formatting for G:H for the TOTAL RESUMO line
        sheet.merge_cells(start_row=total_row, start_column=COLS['CUSTO_MO_UNIT'], end_row=total_row, end_column=COLS['CUSTO_MO_TOTAL'])
        merged_total_mo_cell = sheet.cell(row=total_row, column=COLS['CUSTO_MO_UNIT'], value=mo_formula) # Place formula in G
        merged_total_mo_cell.number_format = currency_format
        merged_total_mo_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_total_mo_cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        merged_total_mo_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # White font
        for col_idx in range(COLS['CUSTO_MO_UNIT'], COLS['CUSTO_MO_TOTAL'] + 1):
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        
        # Assign coordinate
        total_mo_cell_coord = merged_total_mo_cell.coordinate
    elif title == 'SERVIÇOS DISTRATADOS':
        # Apply new formatting for E:F
        sheet.merge_cells(start_row=total_row, start_column=COLS['VALOR_UNIT'], end_row=total_row, end_column=COLS['VALOR_TOTAL'])
        merged_total_val_cell_distratado = sheet.cell(row=total_row, column=COLS['VALOR_UNIT'], value=total_formula) # Place formula in E
        merged_total_val_cell_distratado.number_format = currency_format
        merged_total_val_cell_distratado.alignment = Alignment(horizontal='right', vertical='center')
        merged_total_val_cell_distratado.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        merged_total_val_cell_distratado.font = Font(name='Verdana', color="FFFFFF", size=8) # Set font color to white
        for col_idx in range(COLS['VALOR_UNIT'], COLS['VALOR_TOTAL'] + 1):
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        
        # Assign coordinate
        total_val_cell_coord = merged_total_val_cell_distratado.coordinate

        # Apply new formatting for G:H for the TOTAL SERVIÇOS DISTRATADOS line
        sheet.merge_cells(start_row=total_row, start_column=COLS['CUSTO_MO_UNIT'], end_row=total_row, end_column=COLS['CUSTO_MO_TOTAL'])
        merged_total_mo_cell_distratado = sheet.cell(row=total_row, column=COLS['CUSTO_MO_UNIT'], value=mo_formula) # Place formula in G
        merged_total_mo_cell_distratado.number_format = currency_format
        merged_total_mo_cell_distratado.alignment = Alignment(horizontal='right', vertical='center')
        merged_total_mo_cell_distratado.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        merged_total_mo_cell_distratado.font = Font(name='Verdana', color="FFFFFF", size=8) # White font
        for col_idx in range(COLS['CUSTO_MO_UNIT'], COLS['CUSTO_MO_TOTAL'] + 1):
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        
        # Assign coordinate
        total_mo_cell_coord = merged_total_mo_cell_distratado.coordinate

        # New changes for A:D for the TOTAL line
        total_label_cell.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4) # Merge A:D
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_label_cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8) # White font, keep bold
        # Ensure the entire merged area has the fill color
        for col_idx in range(1, 5): # Columns A to D
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")




    elif title == 'SERVIÇOS DISTRATADOS':
        total_label_cell.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Apply the fill to the merged cells as well
        for col_idx in range(1, 5): # Columns A to D
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")

    # Use the initialized coordinates for return
    return current_row + 1, total_val_cell_coord, total_mo_cell_coord


def write_summary_sheet(sheet, summary_normal, summary_distratado, price_data, bold_font, header_fill, currency_format, regular_font):
    sheet.column_dimensions['A'].width = 12.29
    sheet.column_dimensions['B'].width = 43.37
    sheet.column_dimensions['C'].width = 10.14
    sheet.column_dimensions['D'].width = 5.00
    sheet.column_dimensions['E'].width = 13.91
    sheet.column_dimensions['F'].width = 13.20
    sheet.column_dimensions['G'].width = 13.91
    sheet.column_dimensions['H'].width = 13.20

    next_row, normal_total_coord, normal_mo_coord = _write_summary_section(sheet, summary_normal, price_data, 'RESUMO', 1, bold_font, header_fill, currency_format, regular_font)
    next_row, distratado_total_coord, distratado_mo_coord = _write_summary_section(sheet, summary_distratado, price_data, 'SERVIÇOS DISTRATADOS', next_row, bold_font, header_fill, currency_format, regular_font)
    
    diff_row = next_row
    diff_label_cell = sheet.cell(row=diff_row, column=1, value='DIFERENÇA (NORMAL - DISTRATADO)')
    diff_label_cell.font = bold_font
    
    # Merge E and F for diff_total_cell
    sheet.merge_cells(start_row=diff_row, start_column=5, end_row=diff_row, end_column=6) # Merge E:F
    diff_total_cell = sheet.cell(row=diff_row, column=5, value=f"={normal_total_coord}-{distratado_total_coord}") # Moved to E (column 5)
    diff_total_cell.number_format = currency_format
    diff_total_cell.alignment = Alignment(horizontal='right', vertical='center')
    diff_total_cell.fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")
    diff_total_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # White font
    for col_idx in range(5, 7): # Columns E to F
        sheet.cell(row=diff_row, column=col_idx).fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")

    diff_mo_cell = sheet.cell(row=diff_row, column=8, value=f"={normal_mo_coord}-{distratado_mo_coord}") # This is column H
    diff_mo_cell.number_format = currency_format

    # Merge G and H for diff_mo_cell
    sheet.merge_cells(start_row=diff_row, start_column=7, end_row=diff_row, end_column=8) # Merge G:H
    diff_mo_cell = sheet.cell(row=diff_row, column=7, value=f"={normal_mo_coord}-{distratado_mo_coord}") # Moved to G (column 7)
    diff_mo_cell.number_format = currency_format
    diff_mo_cell.alignment = Alignment(horizontal='right', vertical='center')
    diff_mo_cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
    diff_mo_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # White font
    for col_idx in range(7, 9): # Columns G to H
        sheet.cell(row=diff_row, column=col_idx).fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")

    diff_label_cell.fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")
    sheet.merge_cells(start_row=diff_row, start_column=1, end_row=diff_row, end_column=4)
    diff_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    # Apply the fill to the merged cells as well
    for col_idx in range(1, 5): # Columns A to D
        sheet.cell(row=diff_row, column=col_idx).fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")

    apply_borders_to_range(sheet, diff_row, 1, diff_row, 8)
    

def write_aditivos_distrato_sheet(sheet, summary_normal, summary_distratado, price_data, bold_font, regular_font, header_fill, currency_format, accounting_format, bold_white_font=None, regular_white_font=None, client_normal=None, client_distratado=None):
    sheet.column_dimensions['A'].width = 12.29
    sheet.column_dimensions['B'].width = 43.37
    sheet.column_dimensions['C'].width = 10.14
    sheet.column_dimensions['D'].width = 5.00
    sheet.column_dimensions['E'].width = 13.91
    sheet.column_dimensions['F'].width = 13.20
    sheet.column_dimensions['G'].width = 2.00

    sheet.column_dimensions['H'].width = 12.29
    sheet.column_dimensions['I'].width = 43.37
    sheet.column_dimensions['J'].width = 10.14
    sheet.column_dimensions['K'].width = 5.00
    sheet.column_dimensions['L'].width = 13.91
    sheet.column_dimensions['M'].width = 13.20

    sheet.merge_cells('A1:F1')
    sheet['A1'] = 'ADITIVOS'
    sheet['A1'].font = bold_font
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A1'].fill = header_fill

    sheet.merge_cells('H1:M1')
    sheet['H1'] = 'DISTRATO'
    sheet['H1'].font = bold_font
    sheet['H1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['H1'].fill = PatternFill(start_color='31859c', end_color='31859c', fill_type='solid')

    headers = [('ID. Bloco/Torre', 1), ('Tipo R. Bassani', 2), ('Descrição', 3), ('Metragem', 4), ('Un', 5), ('Valor Total', 6)]
    for text, col in headers:
        cell = sheet.cell(row=2, column=col, value=text)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill

    for text, col in headers:
        right_col = col + 7
        cell = sheet.cell(row=2, column=right_col, value=text)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color='31859c', end_color='31859c', fill_type='solid')

    def normalize_category(category):
        if category == 'Forro':
            return 'Forro'
        if category in ['Parede', 'Revestimento', 'Paredes_e_Revestimentos', 'Paredes e Revestimentos']:
            return 'Paredes e Revestimentos'
        return None

    def collect_items_by_category(block_id, dataset, category):
        results = {}
        if not dataset or block_id not in dataset:
            return results
        for item in dataset[block_id].values():
            cat = normalize_category(item.get('Categoria'))
            if cat == category:
                results[item.get('Tipo Code', '')] = item
        return results

    def write_category_section(category_name, title, row_start):
        # Section title row
        sheet.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=6)
        sheet.merge_cells(start_row=row_start, start_column=8, end_row=row_start, end_column=13)
        left_title = sheet.cell(row=row_start, column=1, value=title)
        right_title = sheet.cell(row=row_start, column=8, value=title)
        for cell, color in [(left_title, '6aa84f'), (right_title, '31859c')]:
            cell.font = bold_white_font or bold_font
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current = row_start + 1
        section_start = current

        # Detalhar por bloco
        blocks = sorted(set((client_normal or {}).keys()) | set((client_distratado or {}).keys()), key=natural_sort_key)
        for block_id in blocks:
            # Cabeçalho por bloco
            sheet.merge_cells(start_row=current, start_column=1, end_row=current, end_column=6)
            blk = sheet.cell(row=current, column=1, value=f'ID. Bloco/Torre: {block_id}')
            blk.font = bold_font
            blk.fill = header_fill
            blk.alignment = Alignment(horizontal='left', vertical='center')

            sheet.merge_cells(start_row=current, start_column=8, end_row=current, end_column=13)
            blk2 = sheet.cell(row=current, column=8, value=f'ID. Bloco/Torre: {block_id}')
            blk2.font = bold_font
            blk2.fill = PatternFill(start_color='31859c', end_color='31859c', fill_type='solid')
            blk2.alignment = Alignment(horizontal='left', vertical='center')

            current += 1

            normal_items = collect_items_by_category(block_id, client_normal or {}, category_name)
            distr_items = collect_items_by_category(block_id, client_distratado or {}, category_name)
            tipos = sorted(set(normal_items.keys()) | set(distr_items.keys()), key=natural_sort_key)

            for tipo in tipos:
                # Lado ADITIVOS
                if tipo in normal_items:
                    item = normal_items[tipo]
                    metragem = round(item.get('BaseQuantity', 0), 2)
                    unit = price_data.get(tipo, {}).get('Un', '')
                    valor_unit = round(price_data.get(tipo, {}).get('Valor', 0), 2)
                    descricao = item.get('Descricao', '')

                    sheet.cell(row=current, column=1, value=block_id)
                    sheet.cell(row=current, column=2, value=tipo)
                    sheet.cell(row=current, column=3, value=descricao)
                    sheet.cell(row=current, column=4, value=metragem).number_format = currency_format
                    sheet.cell(row=current, column=5, value=unit)
                    sheet.cell(row=current, column=6, value=f'=ROUND({metragem}*{valor_unit},2)').number_format = accounting_format

                # Lado DISTRATO
                if tipo in distr_items:
                    item = distr_items[tipo]
                    metragem = round(item.get('BaseQuantity', 0), 2)
                    unit = price_data.get(tipo, {}).get('Un', '')
                    valor_unit = round(price_data.get(tipo, {}).get('Valor', 0), 2)
                    descricao = item.get('Descricao', '')

                    sheet.cell(row=current, column=8, value=block_id)
                    sheet.cell(row=current, column=9, value=tipo)
                    sheet.cell(row=current, column=10, value=descricao)
                    sheet.cell(row=current, column=11, value=metragem).number_format = currency_format
                    sheet.cell(row=current, column=12, value=unit)
                    sheet.cell(row=current, column=13, value=f'=ROUND({metragem}*{valor_unit},2)').number_format = accounting_format

                current += 1

        section_end = current - 1

        # Totais da seção
        sheet.cell(row=current, column=1, value=f'VALOR TOTAL DOS {title}').font = bold_font
        sheet.merge_cells(start_row=current, start_column=1, end_row=current, end_column=4)
        sheet.cell(row=current, column=6, value=f'=ROUND(SUM(F{section_start}:F{section_end}),2)').number_format = accounting_format

        sheet.cell(row=current, column=8, value=f'VALOR TOTAL DOS {title}').font = bold_font
        sheet.merge_cells(start_row=current, start_column=8, end_row=current, end_column=11)
        sheet.cell(row=current, column=13, value=f'=ROUND(SUM(M{section_start}:M{section_end}),2)').number_format = accounting_format

        total_row = current
        return section_start, section_end, total_row, current + 2

    current_row = 3
    forro_start, forro_end, forro_total_row, next_row = write_category_section('Forro', 'FORROS', current_row)
    paredes_start, paredes_end, paredes_total_row, next_row = write_category_section('Paredes e Revestimentos', 'PAREDES E REVESTIMENTOS', next_row)

    subtotal_row = next_row
    sheet.cell(row=subtotal_row, column=1, value='SUBTOTAL FORROS + PAREDES E REVESTIMENTOS').font = bold_font
    sheet.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=4)
    sheet.cell(row=subtotal_row, column=6, value=f'=ROUND(F{forro_total_row}+F{paredes_total_row},2)').number_format = accounting_format
    sheet.cell(row=subtotal_row, column=13, value=f'=ROUND(M{forro_total_row}+M{paredes_total_row},2)').number_format = accounting_format

    final_row = subtotal_row + 1
    sheet.cell(row=final_row, column=1, value='VALOR TOTAL DA PROPOSTA: (ADITIVOS - DISTRATOS)').font = bold_font
    sheet.merge_cells(start_row=final_row, start_column=1, end_row=final_row, end_column=4)
    sheet.cell(row=final_row, column=6, value=f'=ROUND(F{subtotal_row}-M{subtotal_row},2)').number_format = accounting_format
    sheet.cell(row=final_row, column=13, value=f'=ROUND(F{subtotal_row}-M{subtotal_row},2)').number_format = accounting_format

    sheet.row_dimensions[1].hidden = True
    sheet.page_setup.paperSize = 9
    sheet.page_setup.orientation = 'portrait'
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.printArea = f'A1:M{final_row}'
    sheet.sheet_view.view = 'pageBreakPreview'



def write_client_sheet(sheet, client_normal, client_distratado, price_data, bold_font, header_fill, currency_format, regular_font):
    sheet.column_dimensions['A'].width = 12.29
    sheet.column_dimensions['B'].width = 43.37
    sheet.column_dimensions['C'].width = 10.14
    sheet.column_dimensions['D'].width = 5.00
    sheet.column_dimensions['E'].width = 13.91
    sheet.column_dimensions['F'].width = 13.20
    
    current_row = 2
    normal_client_total_coords = {} # Dicionário para armazenar os totais normais por cliente
    
    # --- Bloco 1: Serviços Normais ---
    header_cell = sheet.cell(row=current_row, column=1, value="BLOCO 1: SERVIÇOS NORMAIS")
    header_cell.font = bold_font
    header_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    obra_total_cells = []
    
    sorted_clients_normal = sorted(client_normal.keys(), key=natural_sort_key)
    for client_id in sorted_clients_normal:
        client_header_cell = sheet.cell(row=current_row, column=1, value=client_id)
        client_header_cell.font = bold_font
        client_header_cell.fill = PatternFill(start_color="e46c0a", end_color="e46c0a", fill_type="solid")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        client_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        current_row, subtotal_cells, _ = _write_client_section(sheet, client_normal[client_id], price_data, current_row, bold_font, header_fill, currency_format, regular_font)
        total_formula = f"=+{'+'.join(subtotal_cells)}" if subtotal_cells else "0"
        
        total_row = current_row
        total_label_cell = sheet.cell(row=total_row, column=1, value=f'TOTAL {client_id}')
        total_label_cell.font = bold_font
        
        total_fill = PatternFill(start_color="e46c0a", end_color="e46c0a", fill_type="solid")

        # Format A:D
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, 5):
            sheet.cell(row=total_row, column=col_idx).fill = total_fill

        # Format E:F
        sheet.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=6)
        total_value_cell = sheet.cell(row=total_row, column=5, value=total_formula)
        total_value_cell.number_format = currency_format
        total_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_value_cell.fill = total_fill
        total_value_cell.font = regular_font
        
        total_coord = total_value_cell.coordinate
        obra_total_cells.append(total_coord)
        normal_client_total_coords[client_id] = total_coord # Armazena a coordenada do total
        current_row += 1
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1

    total_obra_row = current_row
    total_obra_label_cell = sheet.cell(row=total_obra_row, column=1, value='VALOR TOTAL DA OBRA')
    total_obra_label_cell.font = bold_font
    
    total_obra_fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")

    # Format A:D
    sheet.merge_cells(start_row=total_obra_row, start_column=1, end_row=total_obra_row, end_column=4)
    total_obra_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    for col_idx in range(1, 5):
        sheet.cell(row=total_obra_row, column=col_idx).fill = total_obra_fill

    # Format E:F
    sheet.merge_cells(start_row=total_obra_row, start_column=5, end_row=total_obra_row, end_column=6)
    total_obra_value_cell = sheet.cell(row=total_obra_row, column=5, value=f"=+{'+'.join(obra_total_cells)}")
    total_obra_value_cell.number_format = currency_format
    total_obra_value_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_obra_value_cell.fill = total_obra_fill
    total_obra_value_cell.font = bold_font
    
    total_obra_coord = total_obra_value_cell.coordinate
    current_row += 1
    format_empty_row(sheet, current_row, regular_font)
    current_row += 1

    # --- Bloco 2: Serviços Distratados ---
    header_cell_distratado = sheet.cell(row=current_row, column=1, value="BLOCO 2: SERVIÇOS DISTRATADOS")
    header_cell_distratado.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)
    header_cell_distratado.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    header_cell_distratado.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    obra_total_distratado_cells = []

    sorted_clients_distratado = sorted(client_distratado.keys(), key=natural_sort_key)
    for client_id in sorted_clients_distratado:
        client_header_cell = sheet.cell(row=current_row, column=1, value=client_id)
        client_header_cell.font = bold_font
        client_header_cell.fill = PatternFill(start_color="e46c0a", end_color="e46c0a", fill_type="solid")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        client_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        current_row, subtotal_cells, _ = _write_client_section(sheet, client_distratado[client_id], price_data, current_row, bold_font, header_fill, currency_format, regular_font)
        total_formula = f"=+{'+'.join(subtotal_cells)}" if subtotal_cells else "0"
        
        total_distratado_row = current_row
        total_distratado_label_cell = sheet.cell(row=total_distratado_row, column=1, value=f'TOTAL DISTRATADO {client_id}')
        total_distratado_label_cell.font = bold_font
        
        total_distratado_fill = PatternFill(start_color="93cddd", end_color="93cddd", fill_type="solid")

        # Format A:D
        sheet.merge_cells(start_row=total_distratado_row, start_column=1, end_row=total_distratado_row, end_column=4)
        total_distratado_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, 5):
            sheet.cell(row=total_distratado_row, column=col_idx).fill = total_distratado_fill

        # Format E:F
        sheet.merge_cells(start_row=total_distratado_row, start_column=5, end_row=total_distratado_row, end_column=6)
        total_distratado_value_cell = sheet.cell(row=total_distratado_row, column=5, value=total_formula)
        total_distratado_value_cell.number_format = currency_format
        total_distratado_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_distratado_value_cell.fill = total_distratado_fill
        total_distratado_value_cell.font = regular_font
        
        distratado_coord = total_distratado_value_cell.coordinate
        obra_total_distratado_cells.append(distratado_coord)
        current_row += 1 # Advance row after total distracted

        format_empty_row(sheet, current_row, regular_font)
        current_row += 1 # Add an extra row increment for the blank line

        # Adiciona a linha de SALDO para o cliente
        saldo_row = current_row
        normal_coord = normal_client_total_coords.get(client_id, "0")
        
        saldo_label_cell = sheet.cell(row=saldo_row, column=1, value=f"SALDO {client_id}")
        saldo_label_cell.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)
        
        saldo_fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")

        # Format A:D
        sheet.merge_cells(start_row=saldo_row, start_column=1, end_row=saldo_row, end_column=4)
        saldo_label_cell.alignment = Alignment(horizontal='left', vertical='center')
        for col_idx in range(1, 5):
            sheet.cell(row=saldo_row, column=col_idx).fill = saldo_fill

        # Format E:F
        sheet.merge_cells(start_row=saldo_row, start_column=5, end_row=saldo_row, end_column=6)
        saldo_value_cell = sheet.cell(row=saldo_row, column=5, value=f"={normal_coord}-{distratado_coord}")
        saldo_value_cell.number_format = currency_format
        saldo_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        saldo_value_cell.fill = saldo_fill
        saldo_value_cell.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)
        
        current_row += 1 # Advances the row after writing SALDO

    format_empty_row(sheet, current_row, regular_font)
    current_row += 1
    # --- Bloco 3: Saldo Final ---
    saldo_final_row = current_row
    saldo_final_label_cell = sheet.cell(row=saldo_final_row, column=1, value='VALOR TOTAL FINAL (OBRA - DISTRATADO)')
    saldo_final_label_cell.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)
    
    saldo_final_fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")

    # Format A:D
    sheet.merge_cells(start_row=saldo_final_row, start_column=1, end_row=saldo_final_row, end_column=4)
    saldo_final_label_cell.alignment = Alignment(horizontal='left', vertical='center')
    for col_idx in range(1, 5):
        sheet.cell(row=saldo_final_row, column=col_idx).fill = saldo_final_fill

    # Format E:F
    sheet.merge_cells(start_row=saldo_final_row, start_column=5, end_row=saldo_final_row, end_column=6)
    
    # Calculate the distratado total directly in the formula
    distratado_total_formula = f"({'+'.join(obra_total_distratado_cells)})" if obra_total_distratado_cells else "0"
    saldo_final_value_cell = sheet.cell(row=saldo_final_row, column=5, value=f"={total_obra_coord}-{distratado_total_formula}")
    saldo_final_value_cell.number_format = currency_format
    saldo_final_value_cell.alignment = Alignment(horizontal='right', vertical='center')
    saldo_final_value_cell.fill = saldo_final_fill
    saldo_final_value_cell.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)

def _write_client_section(sheet, items_by_key, price_data, start_row, bold_font, header_fill, currency_format, regular_font):
    COLS = {'TIPO': 1, 'DESC': 2, 'METRAGEM': 3, 'UN': 4, 'VALOR_UNIT': 5, 'VALOR_TOTAL': 6}
    current_row = start_row
    client_subtotal_cells, item_key_to_cell_map = [], {}

    cats = {'Forro': [], 'Paredes_e_Revestimentos': [], 'Guias e Montantes': []}
    for item_key, item_data in items_by_key.items():
        category = 'Paredes_e_Revestimentos' if item_data.get('Categoria') in ['Parede', 'Revestimento'] else item_data.get('Categoria')
        if category in cats: cats[category].append((item_key, item_data))

    for cat_name, cat_label in [('Forro', 'Forros'), ('Paredes_e_Revestimentos', 'Paredes e Revestimentos'), ('Guias e Montantes', 'Guias e Montantes')]:
        if not cats.get(cat_name): continue
        header_row_data = [('Tipo R. Bassani', COLS['TIPO']), (cat_label, COLS['DESC']), ('Metragem', COLS['METRAGEM']), ('Un', COLS['UN']), ('Valor do Material + MO', COLS['VALOR_UNIT']), ('Valor Total', COLS['VALOR_TOTAL'])]
        for val, col_idx in header_row_data:
            cell = sheet.cell(row=current_row, column=col_idx, value=val)
            cell.font, cell.fill = bold_font, header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['VALOR_TOTAL']) # Add borders
        current_row += 1
        section_start_row, ft16_placeholder_cell, ft16_base_quantity, subtraction_cell_list = current_row, None, 0, []

        if cat_name == 'Paredes_e_Revestimentos':
            sorted_items = sorted(cats[cat_name], key=lambda x: (0 if x[1]['Tipo Code'].startswith('TP') else 1, natural_sort_key(x[1]['Tipo Code'])))
        else: # Forros, Guias e Montantes
            sorted_items = sorted(cats[cat_name], key=lambda x: natural_sort_key(x[1]['Tipo Code']))
        
        for item_key, item in sorted_items:
            price_info = price_data.get(item['Tipo Code'], {'Valor': 0, 'Un': ''})
            main_item_row, is_subclass_item = current_row, item.get('is_subclass', False)

            # Calculate number of sub-items for merging
            num_sub_items = len(item.get('insulation_items', {})) + len(item.get('carenagem_items', {}))

            tipo_cell = sheet.cell(row=main_item_row, column=COLS['TIPO'], value=item['Tipo Code'])
            tipo_cell.alignment = Alignment(horizontal='center', vertical='center')
            tipo_cell.font = regular_font

            # Merge cell if there are sub-items
            if num_sub_items > 0:
                sheet.merge_cells(
                    start_row=main_item_row,
                    start_column=COLS['TIPO'],
                    end_row=main_item_row + num_sub_items,
                    end_column=COLS['TIPO']
                )

            desc_cell = sheet.cell(row=main_item_row, column=COLS['DESC'], value=item['Descricao'])
            desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
            desc_cell.font = regular_font
            
            m_cell_ref = f"{get_column_letter(COLS['METRAGEM'])}{main_item_row}"
            item_key_to_cell_map[item_key] = m_cell_ref
            
            metragem_cell = sheet.cell(row=main_item_row, column=COLS['METRAGEM'])
            metragem_cell.number_format = currency_format
            metragem_cell.font = regular_font
            metragem_cell.alignment = Alignment(vertical='center')
            
            # GARANTIR que a BaseQuantity seja sempre escrita como valor numérico inicial
            metragem_cell.value = item.get('BaseQuantity', 0) # Adicionar esta linha

            if 'formula_contributors' in item:
                parts = [f"({item_key_to_cell_map[c['item_key']]}*{c['count']})" if c['count'] > 1 else item_key_to_cell_map[c['item_key']] for c in item['formula_contributors'] if c['item_key'] in item_key_to_cell_map]
                final_formula = f"=CEILING(({'+'.join(parts)})*{item['formula_multiplier']},1)" if parts else item['BaseQuantity']
                metragem_cell.value = final_formula
            
            un_cell = sheet.cell(row=main_item_row, column=COLS['UN'], value=price_info['Un'])
            un_cell.font = regular_font
            un_cell.alignment = Alignment(horizontal='center', vertical='center')
            valor_unit_cell = sheet.cell(row=main_item_row, column=COLS['VALOR_UNIT'], value=price_info['Valor'])
            valor_unit_cell.number_format = currency_format
            valor_unit_cell.font = regular_font
            valor_unit_cell.alignment = Alignment(vertical='center')
            v_unit_cell = f"{get_column_letter(COLS['VALOR_UNIT'])}{main_item_row}"
            valor_total_cell = sheet.cell(row=main_item_row, column=COLS['VALOR_TOTAL'], value=f"={m_cell_ref}*{v_unit_cell}")
            valor_total_cell.number_format = currency_format
            valor_total_cell.font = regular_font
            valor_total_cell.alignment = Alignment(vertical='center')
            if is_subclass_item: subtraction_cell_list.append(m_cell_ref)
            if item['Tipo Code'] == 'FT16': ft16_placeholder_cell, ft16_base_quantity = m_cell_ref, item.get('FormulaBase', item['BaseQuantity'])
            apply_borders_to_range(sheet, main_item_row, COLS['TIPO'], main_item_row, COLS['VALOR_TOTAL']) # Add borders
            current_row += 1

            # --- SUB-ITEM LOOPS ---
            insulation_cell_references = [] # Initialized here for each main item
            
            # ISOLAMENTO
            for sub_item_key, sub_item in sorted(item.get('insulation_items', {}).items()):
                price_info_sub = price_data.get(sub_item['Tipo Code'], {'Valor': 0, 'Un': ''})
                sub_desc_cell = sheet.cell(row=current_row, column=COLS['DESC'], value=sub_item['Descricao'])
                sub_desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
                sub_desc_cell.font = regular_font
                
                sub_m_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'])
                sub_m_cell.number_format = currency_format
                sub_m_cell.font = regular_font
                sub_m_cell.alignment = Alignment(vertical='center')

                if sub_item.get('is_la_dupla'):
                    sub_m_cell.value = f"={m_cell_ref}*2"
                else:
                    sub_m_cell.value = sub_item['Quantidade']
                    insulation_cell_references.append(sub_m_cell.coordinate) # Only append if not 'lã dupla'
                
                # Add lines to write 'Un' and 'Valor do Material + MO'
                sub_un_cell = sheet.cell(row=current_row, column=COLS['UN'], value=price_info_sub['Un'])
                sub_un_cell.font = regular_font
                sub_un_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                sub_v_unit_cell_obj = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=price_info_sub['Valor'])
                sub_v_unit_cell_obj.number_format = currency_format
                sub_v_unit_cell_obj.font = regular_font
                sub_v_unit_cell_obj.alignment = Alignment(vertical='center')
                
                sub_v_unit_cell = f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}"
                sub_v_total_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"={sub_m_cell.coordinate}*{sub_v_unit_cell}")
                sub_v_total_cell.number_format = currency_format
                sub_v_total_cell.font = regular_font
                sub_v_total_cell.alignment = Alignment(vertical='center')
                apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['VALOR_TOTAL']) # Add borders
                current_row += 1

            # CARNAGEM
            for sub_item_key, sub_item in sorted(item.get('carenagem_items', {}).items()):
                price_info_sub = price_data.get(sub_item['Tipo Code'], {'Valor': 0, 'Un': ''})
                sub_desc_cell = sheet.cell(row=current_row, column=COLS['DESC'], value=sub_item['Descricao'])
                sub_desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
                sub_desc_cell.font = regular_font
                sub_m_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'], value=sub_item['Quantidade'])
                sub_m_cell.number_format = currency_format
                sub_m_cell.font = regular_font
                sub_m_cell.alignment = Alignment(vertical='center')
                sub_un_cell = sheet.cell(row=current_row, column=COLS['UN'], value=price_info_sub['Un'])
                sub_un_cell.font = regular_font
                sub_un_cell.alignment = Alignment(horizontal='center', vertical='center')
                sub_v_unit_cell_obj = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=price_info_sub['Valor'])
                sub_v_unit_cell_obj.number_format = currency_format
                sub_v_unit_cell_obj.font = regular_font
                sub_v_unit_cell_obj.alignment = Alignment(vertical='center')
                sub_v_unit_cell = f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}"
                sub_v_total_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"={sub_m_cell.coordinate}*{sub_v_unit_cell}")
                sub_v_total_cell.number_format = currency_format
                sub_v_total_cell.font = regular_font
                sub_v_total_cell.alignment = Alignment(vertical='center')
                apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['VALOR_TOTAL']) # Add borders
                current_row += 1

            # --- FORMULA LOGIC FOR MAIN ITEM (AFTER SUB-ITEMS ARE PROCESSED) ---
            if cat_name != 'Guias e Montantes' and not is_subclass_item and item['Tipo Code'] != 'FT16':
                final_metragem_value = item.get('BaseQuantity') # Default to BaseQuantity

                if item.get('logic_applied'): # Merged items (BASE + LÃ)
                    formula_parts = []
                    base_val = item.get('FormulaBase', item['BaseQuantity'])
                    if base_val > 0:
                        formula_parts.append(str(base_val))
                    if insulation_cell_references: # Should have one reference
                        formula_parts.extend(insulation_cell_references)
                    
                    if len(formula_parts) > 1:
                        final_metragem_value = f"={'+'.join(formula_parts)}"
                    elif len(formula_parts) == 1:
                        # Check if it's a cell reference (starts with a letter)
                        if re.match(r'^[A-Z]', formula_parts[0]):
                            final_metragem_value = f"={formula_parts[0]}"
                        else:
                            final_metragem_value = formula_parts[0] # Just the base_val
                elif item.get('insulation_items'): # Non-merged item with insulation (just LÃ)
                    if insulation_cell_references: # Should have one reference
                        final_metragem_value = f"={insulation_cell_references[0]}" # Only the insulation cell ref
                    else:
                        final_metragem_value = item.get('BaseQuantity') # Fallback if no insulation ref found (shouldn't happen)
                else: # Item without insulation, not merged (just BASE)
                    final_metragem_value = item.get('BaseQuantity') # Use the numeric value directly (no '=')

                metragem_cell.value = final_metragem_value
        
        # --- FT16 FORMULA (AFTER ALL ITEMS IN CATEGORY ARE PROCESSED) ---
        if ft16_placeholder_cell and subtraction_cell_list:
            base_val_str = str(ft16_base_quantity).replace(',', '.')
            sheet[ft16_placeholder_cell] = f"={base_val_str}-{'-'.join(subtraction_cell_list)}"
            sheet[ft16_placeholder_cell].number_format = currency_format
            sheet[ft16_placeholder_cell].font = regular_font

        # --- SUBTOTALS ---
        subtotal_row_idx = current_row
        subtotal_cell = sheet.cell(row=subtotal_row_idx, column=COLS['TIPO'], value='SUB-TOTAL')
        subtotal_cell.font = bold_font

        # Apply formatting for A:D
        subtotal_fill = PatternFill(start_color="ffff99", end_color="ffff99", fill_type="solid")
        subtotal_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=subtotal_row_idx, start_column=1, end_row=subtotal_row_idx, end_column=4)

        for col_idx in range(1, 5):
            sheet.cell(row=subtotal_row_idx, column=col_idx).fill = subtotal_fill

        # Merge E:F and apply formula
        sheet.merge_cells(start_row=subtotal_row_idx, start_column=COLS['VALOR_UNIT'], end_row=subtotal_row_idx, end_column=COLS['VALOR_TOTAL'])
        
        total_col_letter = get_column_letter(COLS['VALOR_TOTAL'])
        formula = f"=SUM({total_col_letter}{section_start_row}:{total_col_letter}{current_row - 1})"
        
        merged_value_cell = sheet.cell(row=subtotal_row_idx, column=COLS['VALOR_UNIT'], value=formula)
        merged_value_cell.number_format = currency_format
        merged_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_value_cell.fill = subtotal_fill
        merged_value_cell.font = bold_font
        
        apply_borders_to_range(sheet, subtotal_row_idx, COLS['TIPO'], subtotal_row_idx, COLS['VALOR_TOTAL']) # Add borders
        client_subtotal_cells.append(merged_value_cell.coordinate)
        current_row += 1
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1
    
    return current_row, client_subtotal_cells, item_key_to_cell_map