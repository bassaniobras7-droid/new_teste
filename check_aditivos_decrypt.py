import msoffcrypto
import openpyxl
import os

# Descriptografar o arquivo
encrypted_file = 'Relatorios_Com_Formulas.xlsx'
decrypted_file = 'Relatorios_Com_Formulas_temp.xlsx'

with open(encrypted_file, 'rb') as fin:
    office_file = msoffcrypto.OfficeFile(fin)
    office_file.load_key(password='tecnicob')
    with open(decrypted_file, 'wb') as fout:
        office_file.decrypt(fout)

# Agora abrir e verificar
wb = openpyxl.load_workbook(decrypted_file)
ws = wb['Aditivos x Distrato']

print("=" * 100)
print("ESTRUTURA: Aditivos x Distrato - Casamento por Tipo R. Bassani com Agrupamento por Bloco")
print("=" * 100)

# Verificar headers
print("\nLinha 2 (Headers):")
print(f"  Left (A2:F2): {[ws.cell(2, col).value for col in range(1, 7)]}")
print(f"  Right (H2:M2): {[ws.cell(2, col).value for col in range(8, 14)]}")

# Ver as primeiras linhas de dados
print("\nPrimeiras 20 linhas de dados (linhas 3-22):")
for row in range(3, min(23, ws.max_row + 1)):
    left_bloco = ws.cell(row, 1).value
    left_tipo = ws.cell(row, 2).value
    left_desc = ws.cell(row, 3).value
    left_total = ws.cell(row, 6).value
    
    right_bloco = ws.cell(row, 8).value
    right_tipo = ws.cell(row, 9).value
    right_desc = ws.cell(row, 10).value
    right_total = ws.cell(row, 13).value
    
    left_str = f"Bloco:{left_bloco} | Tipo:{left_tipo}" if left_tipo else ""
    right_str = f"Bloco:{right_bloco} | Tipo:{right_tipo}" if right_tipo else ""
    
    if left_str or right_str:
        print(f"  Row {row:2}: LEFT[{left_str:30}] | RIGHT[{right_str:30}]")

# Encontrar as linhas com "TOTAL"
print("\nLinhas com TOTAL GERAL:")
for row in range(3, ws.max_row + 1):
    val = ws.cell(row, 1).value
    if val and 'TOTAL' in str(val).upper():
        total_left = ws.cell(row, 6).value
        total_right = ws.cell(row, 13).value
        print(f"  Row {row}: {val}")
        print(f"    - LEFT (F{row}): {total_left}")
        print(f"    - RIGHT (M{row}): {total_right}")

wb.close()
os.remove(decrypted_file)
print("\n✓ Verificação concluída")
