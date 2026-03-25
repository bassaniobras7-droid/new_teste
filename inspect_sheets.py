import openpyxl
from openpyxl.utils import get_column_letter

ANTIGO = r"C:/Users/Wesllen.Santana/Downloads/bassani/Relatorios_Com_Formulas_dec2.xlsx"
NOVO   = r"C:/Users/Wesllen.Santana/Downloads/bassani/Relatorios_Com_Formulas_novo_dec2.xlsx"

wb_a = openpyxl.load_workbook(ANTIGO)
wb_n = openpyxl.load_workbook(NOVO)

def rgb(color_obj):
    if color_obj is None:
        return None
    try:
        t = color_obj.type
        if t == 'rgb':
            val = color_obj.rgb
            return val if val and val != '00000000' else None
        if t == 'theme':
            return f"THEME:{color_obj.theme}"
        if t == 'indexed':
            return f"INDEXED:{color_obj.indexed}"
    except:
        pass
    return None

def dump_sheet(ws, label):
    print(f"\n{'='*60}")
    print(f"{label} - {ws.title}")
    print(f"{'='*60}")
    print(f"Max: {ws.max_row} linhas x {ws.max_column} cols")

    print("\nLARGURAS DE COLUNAS:")
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.width:
            print(f"  {col_letter}: {col_dim.width:.2f}")

    print("\nALTURAS DE LINHAS:")
    for row_idx, row_dim in ws.row_dimensions.items():
        if row_dim.height:
            print(f"  Linha {row_idx}: {row_dim.height:.2f}")

    print("\nMESCLAGENS:")
    for m in sorted([str(x) for x in ws.merged_cells.ranges]):
        print(f"  {m}")

    ps = ws.page_setup
    print(f"\nPAGE SETUP: orientation={ps.orientation}, paperSize={ps.paperSize}, scale={ps.scale}")
    print(f"PRINT AREA: {ws.print_area}")
    print(f"FREEZE: {ws.freeze_panes}")

    print("\nCELULAS (nao vazias ou com estilo):")
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            fi = cell.fill
            has_fill = fi and fi.fill_type and fi.fill_type != 'none'
            fg = rgb(fi.fgColor) if has_fill and fi.fgColor else None
            has_border = False
            b = cell.border
            if b:
                for side in [b.left, b.right, b.top, b.bottom]:
                    if side and side.border_style:
                        has_border = True
                        break
            fo = cell.font
            bold = fo.bold if fo else None
            font_color = rgb(fo.color) if fo and fo.color else None
            font_size = fo.size if fo else None
            font_name = fo.name if fo else None

            if val is not None or fg or has_border:
                info_parts = []
                if val is not None:
                    info_parts.append(f"val={repr(str(val)[:50])}")
                if fg:
                    info_parts.append(f"fill=#{fg}")
                if bold:
                    info_parts.append(f"bold")
                if font_color:
                    info_parts.append(f"fcol=#{font_color}")
                if font_size:
                    info_parts.append(f"fsize={font_size}")
                if font_name:
                    info_parts.append(f"fname={font_name}")
                al = cell.alignment
                if al and al.horizontal:
                    info_parts.append(f"halign={al.horizontal}")
                if has_border:
                    info_parts.append(f"border=yes")
                print(f"  {cell.coordinate}: {', '.join(info_parts)}")

# Inspecionar abas Cliente e Resumo que pareciam sem diferencas
for sheet_name in ['Cliente', 'Resumo', 'Relação Média Material']:
    try:
        ws_a = wb_a[sheet_name]
        ws_n = wb_n[sheet_name]
    except KeyError:
        # Tentar nome com caracteres especiais
        for name in wb_a.sheetnames:
            if 'Rela' in name:
                sheet_name = name
                ws_a = wb_a[name]
                ws_n = wb_n[name]
                break

    dump_sheet(ws_a, "ANTIGO")
    dump_sheet(ws_n, "NOVO")
