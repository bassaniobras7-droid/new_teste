import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import math
import re

# ==============================================================================
# FUNÇÕES AUXILIARES DE FORMATAÇÃO
# ==============================================================================

def apply_borders_to_range(sheet, min_row, min_col, max_row, max_col):
    thin_border = Border(left=Side(border_style="thin", color="000000"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
    for row in sheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = thin_border

def format_empty_row(sheet, row_idx, regular_font):
    sheet.row_dimensions[row_idx].height = 5
    sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    # Add a fill for visibility
    gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    sheet.cell(row=row_idx, column=1).fill = gray_fill

# ==============================================================================
# FUNÇÕES AUXILIARES E DE CARREGAMENTO DE DADOS
# ==============================================================================

def clean_numeric_column(series):
    return pd.to_numeric(series.astype(str).str.replace(',', '.'), errors='coerce').fillna(0)

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', str(s))]

def load_price_data(filename='Valores Ctba.csv'):
    try:
        df_prices = pd.read_csv(filename, sep=';', header=0)
        df_prices.dropna(subset=['Tipo R. Bassani'], inplace=True)
        price_data = {}
        for _, row in df_prices.iterrows():
            tipo_code = row['Tipo R. Bassani']
            price_data[tipo_code] = {
                'Un': row.get('Un'),
                'Valor': clean_numeric_column(pd.Series([row.get('Valor do Material + MO')]))[0],
                'Custo MO': clean_numeric_column(pd.Series([row.get('Custo MO à Pagar')]))[0],
                'Descricao': row.get('Forros')
            }
        return price_data
    except FileNotFoundError:
        print(f"AVISO: Arquivo de preços '{filename}' não encontrado.")
        return {}
    except Exception as e:
        print(f"ERRO ao carregar o arquivo de preços: {e}")
        return {}

def load_subclass_data(filename='Subclasse.csv'):
    try:
        df_subclass = pd.read_csv(filename, sep=';', header=0)
        df_subclass.dropna(subset=['Tipo'], inplace=True)
        return set(df_subclass['Tipo'].astype(str).str.strip())
    except FileNotFoundError:
        print(f"AVISO: Arquivo de subclasses '{filename}' não encontrado.")
        return set()
    except Exception as e:
        print(f"ERRO ao carregar o arquivo de subclasses: {e}")
        return set()

# ==============================================================================
# PROCESSAMENTO DE DADOS
# ==============================================================================

def process_summary_data(client_data):
    summary_agg = {}
    for client_id, client_items in client_data.items():
        subclass_total_qty = sum(item_data.get('BaseQuantity', 0) for item_data in client_items.values() if item_data.get('is_subclass', False))

        for item_data in client_items.values():
            cat = item_data['Categoria']
            summary_cat = 'Paredes' if cat in ['Parede', 'Revestimento'] else ('Forros' if cat == 'Forro' else cat)

            for carenagem_item in item_data.get('carenagem_items', {}).values():
                carenagem_tipo_code = carenagem_item['Tipo Code']
                if carenagem_tipo_code not in summary_agg:
                    summary_agg[carenagem_tipo_code] = {'Quantidade': 0, 'Descricao': carenagem_item['Descricao'], 'Categoria': summary_cat}
                summary_agg[carenagem_tipo_code]['Quantidade'] += carenagem_item['Quantidade']

            for insulation_item in item_data.get('insulation_items', {}).values():
                insul_tipo_code = insulation_item['Tipo Code']
                if insul_tipo_code not in summary_agg:
                    summary_agg[insul_tipo_code] = {'Quantidade': 0, 'Descricao': insulation_item['Descricao'], 'Categoria': 'Isolamento'}
                summary_agg[insul_tipo_code]['Quantidade'] += insulation_item['Quantidade']

            main_tipo_code = item_data['Tipo Code']
            qty = item_data['BaseQuantity']
            if main_tipo_code == 'FT16' and subclass_total_qty > 0:
                qty -= subclass_total_qty
            
            if main_tipo_code not in summary_agg:
                summary_agg[main_tipo_code] = {'Quantidade': 0, 'Descricao': item_data['Descricao'], 'Categoria': summary_cat}
            summary_agg[main_tipo_code]['Quantidade'] += qty
    return summary_agg

def process_client_data(price_data, subclass_types, forro_file, generico_file, paredes_file):
    data_by_client = {}

    def add_or_aggregate_item(client_id, item_data, has_insulation=False, insulation_data=None, carenagem_data=None):
        if client_id not in data_by_client: data_by_client[client_id] = {}
        key = (item_data['Tipo Code'], item_data['Categoria'], has_insulation)
        if key not in data_by_client[client_id]:
            data_by_client[client_id][key] = {
                'Tipo Code': item_data['Tipo Code'], 'Descricao': item_data.get('Descricao'), 'BaseQuantity': 0,
                'Categoria': item_data['Categoria'], 'has_insulation': has_insulation, 'insulation_items': {},
                'carenagem_items': {}, 'is_subclass': item_data.get('is_subclass', False)
            }
        if item_data.get('Descricao') and not data_by_client[client_id][key].get('Descricao'):
            data_by_client[client_id][key]['Descricao'] = item_data.get('Descricao')
        data_by_client[client_id][key]['BaseQuantity'] += item_data.get('Quantidade', 0)
        if item_data.get('is_subclass', False): data_by_client[client_id][key]['is_subclass'] = True
        if insulation_data:
            insul_key = insulation_data['Tipo Code']
            if insul_key not in data_by_client[client_id][key]['insulation_items']:
                data_by_client[client_id][key]['insulation_items'][insul_key] = {'Tipo Code': insul_key, 'Descricao': insulation_data['Descricao'], 'Quantidade': 0, 'is_la_dupla': insulation_data.get('is_la_dupla', False)}
            data_by_client[client_id][key]['insulation_items'][insul_key]['Quantidade'] += insulation_data['Quantidade']
        if carenagem_data:
            carenagem_key = carenagem_data['Descricao']
            if carenagem_key not in data_by_client[client_id][key]['carenagem_items']:
                data_by_client[client_id][key]['carenagem_items'][carenagem_key] = {'Tipo Code': carenagem_data['Tipo Code'], 'Descricao': carenagem_data['Descricao'], 'Quantidade': 0}
            data_by_client[client_id][key]['carenagem_items'][carenagem_key]['Quantidade'] += carenagem_data['Quantidade']

    for file_path, process_func in [(forro_file, 'forro'), (generico_file, 'generico'), (paredes_file, 'paredes')]:
        try:
            df = pd.read_csv(file_path, sep=';', header=1)
            df.dropna(subset=['ID. Bloco/Torre'], inplace=True)
            if process_func == 'forro':
                df['Área'] = clean_numeric_column(df['Área'])
                df['Perímetro'] = clean_numeric_column(df['Perímetro'])
                for _, row in df.iterrows():
                    tipo_code = str(row['Sistema Construtivo R. Bassani']).strip()
                    unit = price_data.get(tipo_code, {}).get('Un', '')
                    qty = row['Área'] if unit == 'm²' else (row['Perímetro'] if unit == 'm' else row['Área'])
                    add_or_aggregate_item(row['ID. Bloco/Torre'], {'Tipo Code': tipo_code, 'Descricao': price_data.get(tipo_code, {}).get('Descricao', row['Tipo']), 'Quantidade': qty, 'Categoria': 'Forro', 'is_subclass': tipo_code in subclass_types})
            elif process_func == 'generico':
                df['Contador'] = clean_numeric_column(df['Contador'])
                for _, row in df.iterrows():
                    tipo_code = str(row['Sistema Construtivo R. Bassani']).strip()
                    add_or_aggregate_item(row['ID. Bloco/Torre'], {'Tipo Code': tipo_code, 'Descricao': price_data.get(tipo_code, {}).get('Descricao', row['Tipo']), 'Quantidade': row['Contador'], 'Categoria': 'Parede' if row['Classe'] == 'Parede' else 'Forro', 'is_subclass': tipo_code in subclass_types})
            elif process_func == 'paredes':
                df['Área'] = clean_numeric_column(df['Área'])
                df['Altura desconectada'] = clean_numeric_column(df['Altura desconectada'])
                df['Comprimento'] = clean_numeric_column(df['Comprimento'])
                for _, row in df.iterrows():
                    client_id, tipo_code, categoria = row['ID. Bloco/Torre'], str(row['Sistema Construtivo R. Bassani']).strip(), row['Classe']
                    is_subclass, osb_perfil = tipo_code in subclass_types, str(row.get('OSB/Perfil')).strip()
                    is_carenagem, is_la_dupla = osb_perfil == 'Carenagem', osb_perfil == 'Lã dupla'
                    unit, row_quantity = price_data.get(tipo_code, {}).get('Un', ''), 0
                    if categoria in ['Parede', 'Revestimento']:
                        if unit == 'm²':
                            row_quantity = row['Área']
                        elif unit == 'm':
                            row_quantity = row['Altura desconectada']
                        else:
                            row_quantity = 0
                    elif categoria == 'Forro':
                        if unit == 'm²':
                            row_quantity = row['Área']
                        elif unit == 'm':
                            row_quantity = row['Comprimento']
                        else:
                            row_quantity = 0
                    item_data, carenagem_data, insulation_data = None, None, None
                    has_insulation = pd.notna(row['Sistema de Isolamento']) and str(row['Sistema de Isolamento']).strip() != ''
                    if is_carenagem:
                        tipo_code_carenagem = f"{tipo_code}-CAR"
                        desc = price_data.get(tipo_code_carenagem, {}).get('Descricao', row['Tipo'])
                        # Check if 'desc' already ends with '(Carenagem)'
                        if not desc.strip().endswith('(Carenagem)'):
                            desc = f"{desc} (Carenagem)"
                        carenagem_data = {'Tipo Code': tipo_code_carenagem, 'Descricao': desc, 'Quantidade': row_quantity}
                        item_data = {'Tipo Code': tipo_code, 'Categoria': categoria, 'Quantidade': 0, 'is_subclass': is_subclass}
                    else:
                        item_data = {'Tipo Code': tipo_code, 'Descricao': price_data.get(tipo_code, {}).get('Descricao', row['Tipo']), 'Quantidade': row_quantity, 'Categoria': categoria, 'is_subclass': is_subclass}
                        if has_insulation:
                            tipo_code_insul = row['Sistema de Isolamento']
                            insulation_data = {'Tipo Code': tipo_code_insul, 'Descricao': price_data.get(tipo_code_insul, {}).get('Descricao', f"Isolamento {tipo_code_insul}"), 'Quantidade': row['Área'] * (2 if is_la_dupla else 1), 'is_la_dupla': is_la_dupla}
                    add_or_aggregate_item(client_id, item_data, has_insulation=has_insulation, insulation_data=insulation_data, carenagem_data=carenagem_data)
        except FileNotFoundError:
            print(f"AVISO: Arquivo '{file_path}' não encontrado.")
    return data_by_client

# ==============================================================================
# GERAÇÃO DO ARQUIVO EXCEL
# ==============================================================================

def write_excel_with_formulas(summary_normal, summary_distratado, client_normal, client_distratado, price_data, filename):
    workbook = openpyxl.Workbook()
    bold_font = Font(name='Verdana', bold=True, size=8)
    regular_font = Font(name='Verdana', size=8)
    header_fill = PatternFill(start_color="99ccff", end_color="99ccff", fill_type="solid")
    currency_format = '#,##0.00'

    ws_resumo = workbook.active
    ws_resumo.title = 'Resumo'
    write_summary_sheet(ws_resumo, summary_normal, summary_distratado, price_data, bold_font, header_fill, currency_format, regular_font)
    
    ws_cliente = workbook.create_sheet(title='Cliente')
    write_client_sheet(ws_cliente, client_normal, client_distratado, price_data, bold_font, header_fill, currency_format, regular_font)
    
    try:
        workbook.save(filename)
        print(f"Arquivo '{filename}' gerado com sucesso.")
    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}")

def _write_summary_section(sheet, summary_data, price_data, title, start_row, bold_font, header_fill, currency_format, regular_font):
    COLS = {'TIPO': 1, 'DESC': 2, 'METRAGEM': 3, 'UN': 4, 'VALOR_UNIT': 5, 'VALOR_TOTAL': 6, 'CUSTO_MO_UNIT': 7, 'CUSTO_MO_TOTAL': 8}
    
    title_cell = sheet.cell(row=start_row, column=1, value=title)
    title_cell.font = bold_font
    
    if title == 'RESUMO':
        title_cell.fill = PatternFill(start_color="b3a2c7", end_color="b3a2c7", fill_type="solid")
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        subtitle_row = start_row + 1
        subtitle_cell = sheet.cell(row=subtitle_row, column=1, value="SERVIÇOS ADICIONAIS (NOVO LAYOUT)")
        subtitle_cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8) # White font, keep bold
        subtitle_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        sheet.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=6) # Merge A:F
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Ensure the entire merged area has the fill color
        for col_idx in range(1, 7): # Columns A to F
            sheet.cell(row=subtitle_row, column=col_idx).fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")

        current_row = start_row + 2
    elif title == 'SERVIÇOS DISTRATADOS':
        title_cell.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6) # Merge A:F
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8) # White font, keep bold, and name Verdana
        # Ensure the entire merged area has the fill color
        for col_idx in range(1, 7): # Columns A to F
            sheet.cell(row=start_row, column=col_idx).fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        current_row = start_row + 1
    else:
        current_row = start_row + 1

    subtotal_valor_cells, subtotal_mo_cells = [], []

    def summary_sort_key_with_tp_priority(item):
        tipo_code = item[0]
        is_car = tipo_code.endswith('-CAR')
        base_code = tipo_code.replace('-CAR', '')
        tp_priority = 0 if base_code.startswith('TP') else 1
        natural_base_sort = (tp_priority, natural_sort_key(base_code))
        car_suffix_priority = 1 if is_car else 0
        return (natural_base_sort, car_suffix_priority)

    for categoria, nome_categoria in [('Forros', 'Forros'), ('Paredes', 'Paredes e Revestimentos'), ('Guias e Montantes', 'Guias e Montantes'), ('Isolamento', 'Isolamento Acústico')]:
        items_cat = {k: v for k, v in summary_data.items() if v.get('Categoria') == categoria}
        if not items_cat: continue

        header_row_start = current_row

        header_row = [ ('Tipo R. Bassani', COLS['TIPO']), (nome_categoria, COLS['DESC']), ('Metragem', COLS['METRAGEM']), ('Un', COLS['UN']),
                       ('Valor do Material + MO', COLS['VALOR_UNIT']), ('Valor Total', COLS['VALOR_TOTAL']), ('Custo MO', COLS['CUSTO_MO_UNIT']), ('Valor Total MO', COLS['CUSTO_MO_TOTAL']) ]
        for val, col in header_row:
            cell = sheet.cell(row=current_row, column=col, value=val)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) # Add wrap_text

            if val in ['Custo MO', 'Valor Total MO']:
                cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8) # White font, keep bold
            else:
                cell.fill = header_fill # Existing header fill
        
        # Apply borders to the header row
        apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['CUSTO_MO_TOTAL'])

        current_row += 1
        section_start_row = current_row

        if categoria == 'Paredes':
            sorted_items = sorted(items_cat.items(), key=summary_sort_key_with_tp_priority)
        elif categoria == 'Isolamento':
            sorted_items = sorted(items_cat.items(), key=lambda item: (0 if item[0].startswith('TP') else 1, natural_sort_key(item[0])))
        else:
            sorted_items = sorted(items_cat.items(), key=lambda item: natural_sort_key(item[0]))

        for tipo_code, data in sorted_items:
            price_info = price_data.get(tipo_code, {'Valor': 0, 'Un': '', 'Custo MO': 0})
            tipo_cell = sheet.cell(row=current_row, column=COLS['TIPO'], value=tipo_code)
            tipo_cell.alignment = Alignment(horizontal='center', vertical='center')
            tipo_cell.font = regular_font
            tipo_cell.font = regular_font
            desc_cell = sheet.cell(row=current_row, column=COLS['DESC'], value=data['Descricao'])
            desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
            desc_cell.font = regular_font
            desc_cell.font = regular_font
            metragem_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'], value=data['Quantidade'])
            metragem_cell.number_format = currency_format
            metragem_cell.font = regular_font
            metragem_cell.alignment = Alignment(vertical='center')
            un_cell = sheet.cell(row=current_row, column=COLS['UN'], value=price_info['Un'])
            un_cell.font = regular_font
            un_cell.alignment = Alignment(horizontal='center', vertical='center')
            valor_unit_cell = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=price_info['Valor'])
            valor_unit_cell.number_format = currency_format
            valor_unit_cell.font = regular_font
            valor_unit_cell.alignment = Alignment(vertical='center')
            custo_mo_unit_cell = sheet.cell(row=current_row, column=COLS['CUSTO_MO_UNIT'], value=price_info['Custo MO'])
            custo_mo_unit_cell.number_format = currency_format
            custo_mo_unit_cell.font = regular_font
            custo_mo_unit_cell.alignment = Alignment(vertical='center')
            
            m_cell, v_cell, mo_cell = f"{get_column_letter(COLS['METRAGEM'])}{current_row}", f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}", f"{get_column_letter(COLS['CUSTO_MO_UNIT'])}{current_row}"
            valor_total_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"={m_cell}*{v_cell}")
            valor_total_cell.number_format = currency_format
            valor_total_cell.font = regular_font
            valor_total_cell.alignment = Alignment(vertical='center')
            custo_mo_total_cell = sheet.cell(row=current_row, column=COLS['CUSTO_MO_TOTAL'], value=f"={m_cell}*{mo_cell}")
            custo_mo_total_cell.number_format = currency_format
            custo_mo_total_cell.font = regular_font
            custo_mo_total_cell.alignment = Alignment(vertical='center')
            
            # Apply fill to G and H columns for data rows
            fill_c0c0c0 = PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")
            sheet.cell(row=current_row, column=COLS['CUSTO_MO_UNIT']).fill = fill_c0c0c0
            sheet.cell(row=current_row, column=COLS['CUSTO_MO_TOTAL']).fill = fill_c0c0c0
            
            # Apply borders to the data row
            apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['CUSTO_MO_TOTAL'])
            
            current_row += 1
            
        subtotal_row = current_row
        subtotal_cell = sheet.cell(row=subtotal_row, column=COLS['TIPO'], value='SUB-TOTAL')
        subtotal_cell.font = bold_font
        
        subtotal_fill = PatternFill(start_color="ffff99", end_color="ffff99", fill_type="solid")
        subtotal_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=4)

        for col_idx in range(1, 5):
            sheet.cell(row=subtotal_row, column=col_idx).fill = subtotal_fill

        # Apply new formatting for E:F
        sheet.merge_cells(start_row=subtotal_row, start_column=COLS['VALOR_UNIT'], end_row=subtotal_row, end_column=COLS['VALOR_TOTAL'])
        
        # Define total_col and mo_col here
        total_col = get_column_letter(COLS['VALOR_TOTAL'])
        mo_col = get_column_letter(COLS['CUSTO_MO_TOTAL'])

        # Apply formula to the merged cell (column E)
        formula_total = f"=SUM({total_col}{section_start_row}:{total_col}{current_row - 1})"
        merged_value_cell = sheet.cell(row=subtotal_row, column=COLS['VALOR_UNIT'], value=formula_total)
        merged_value_cell.number_format = currency_format
        merged_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_value_cell.fill = subtotal_fill # Also apply fill to the merged value cell
        merged_value_cell.font = bold_font

        # Apply formula to CUSTO_MO_TOTAL (column H)
        formula_mo = f"=SUM({mo_col}{section_start_row}:{mo_col}{current_row - 1})"
        
        # Merge G and H
        sheet.merge_cells(start_row=subtotal_row, start_column=COLS['CUSTO_MO_UNIT'], end_row=subtotal_row, end_column=COLS['CUSTO_MO_TOTAL'])
        merged_mo_cell = sheet.cell(row=subtotal_row, column=COLS['CUSTO_MO_UNIT'], value=formula_mo)
        merged_mo_cell.number_format = currency_format
        merged_mo_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_mo_cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        merged_mo_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # Added font color # Corrected color

        # Apply borders to the subtotal row
        apply_borders_to_range(sheet, subtotal_row, COLS['TIPO'], subtotal_row, COLS['CUSTO_MO_TOTAL'])
        
        subtotal_valor_cells.append(f"{get_column_letter(COLS['VALOR_UNIT'])}{subtotal_row}")
        subtotal_mo_cells.append(f"{get_column_letter(COLS['CUSTO_MO_UNIT'])}{subtotal_row}") # Update to reference column G
        current_row += 1
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1

    total_row = current_row
    total_label_cell = sheet.cell(row=total_row, column=COLS['TIPO'], value=f'TOTAL {title}')
    total_label_cell.font = bold_font
    total_formula = f"=+{'+'.join(subtotal_valor_cells)}" if subtotal_valor_cells else "0"
    mo_formula = f"=+{'+'.join(subtotal_mo_cells)}" if subtotal_mo_cells else "0"
    
    # Logic is now handled inside the if/elif blocks to correctly deal with merged cells.
    total_val_cell = None 
    total_mo_cell = None

    if title == 'RESUMO':
        total_label_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Apply the fill to the merged cells as well
        for col_idx in range(1, 5): # Columns A to D
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")

        # Apply new formatting for E:F
        sheet.merge_cells(start_row=total_row, start_column=COLS['VALOR_UNIT'], end_row=total_row, end_column=COLS['VALOR_TOTAL'])
        merged_total_val_cell = sheet.cell(row=total_row, column=COLS['VALOR_UNIT'], value=total_formula) # Place formula in E
        merged_total_val_cell.number_format = currency_format
        merged_total_val_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_total_val_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        merged_total_val_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # Set font color to white
        for col_idx in range(COLS['VALOR_UNIT'], COLS['VALOR_TOTAL'] + 1):
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        
        # Reassign total_val_cell to the new merged cell
        total_val_cell = merged_total_val_cell

        # Apply new formatting for G:H for the TOTAL RESUMO line
        sheet.merge_cells(start_row=total_row, start_column=COLS['CUSTO_MO_UNIT'], end_row=total_row, end_column=COLS['CUSTO_MO_TOTAL'])
        merged_total_mo_cell = sheet.cell(row=total_row, column=COLS['CUSTO_MO_UNIT'], value=mo_formula) # Place formula in G
        merged_total_mo_cell.number_format = currency_format
        merged_total_mo_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_total_mo_cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        merged_total_mo_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # White font
        for col_idx in range(COLS['CUSTO_MO_UNIT'], COLS['CUSTO_MO_TOTAL'] + 1):
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        
        # Reassign total_mo_cell to the new merged cell
        total_mo_cell = merged_total_mo_cell
    elif title == 'SERVIÇOS DISTRATADOS':
        # Apply new formatting for E:F
        sheet.merge_cells(start_row=total_row, start_column=COLS['VALOR_UNIT'], end_row=total_row, end_column=COLS['VALOR_TOTAL'])
        merged_total_val_cell_distratado = sheet.cell(row=total_row, column=COLS['VALOR_UNIT'], value=total_formula) # Place formula in E
        merged_total_val_cell_distratado.number_format = currency_format
        merged_total_val_cell_distratado.alignment = Alignment(horizontal='right', vertical='center')
        merged_total_val_cell_distratado.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        merged_total_val_cell_distratado.font = Font(name='Verdana', color="FFFFFF", size=8) # Set font color to white
        for col_idx in range(COLS['VALOR_UNIT'], COLS['VALOR_TOTAL'] + 1):
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        
        # Reassign total_val_cell to the new merged cell (for the distratado total)
        total_val_cell = merged_total_val_cell_distratado

        # Apply new formatting for G:H for the TOTAL SERVIÇOS DISTRATADOS line
        sheet.merge_cells(start_row=total_row, start_column=COLS['CUSTO_MO_UNIT'], end_row=total_row, end_column=COLS['CUSTO_MO_TOTAL'])
        merged_total_mo_cell_distratado = sheet.cell(row=total_row, column=COLS['CUSTO_MO_UNIT'], value=mo_formula) # Place formula in G
        merged_total_mo_cell_distratado.number_format = currency_format
        merged_total_mo_cell_distratado.alignment = Alignment(horizontal='right', vertical='center')
        merged_total_mo_cell_distratado.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        merged_total_mo_cell_distratado.font = Font(name='Verdana', color="FFFFFF", size=8) # White font
        for col_idx in range(COLS['CUSTO_MO_UNIT'], COLS['CUSTO_MO_TOTAL'] + 1):
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        
        # Reassign total_mo_cell to the new merged cell
        total_mo_cell = merged_total_mo_cell_distratado

        # New changes for A:D for the TOTAL line
        total_label_cell.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4) # Merge A:D
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_label_cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8) # White font, keep bold
        # Ensure the entire merged area has the fill color
        for col_idx in range(1, 5): # Columns A to D
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")




    elif title == 'SERVIÇOS DISTRATADOS':
        total_label_cell.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        # Apply the fill to the merged cells as well
        for col_idx in range(1, 5): # Columns A to D
            sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")

    total_val_cell.number_format = currency_format
    total_mo_cell.number_format = currency_format
    
    apply_borders_to_range(sheet, total_row, COLS['TIPO'], total_row, COLS['CUSTO_MO_TOTAL'])
    
    return current_row + 1, total_val_cell.coordinate, total_mo_cell.coordinate

def write_summary_sheet(sheet, summary_normal, summary_distratado, price_data, bold_font, header_fill, currency_format, regular_font):
    sheet.column_dimensions['A'].width = 12.29
    sheet.column_dimensions['B'].width = 43.37
    sheet.column_dimensions['C'].width = 10.14
    sheet.column_dimensions['D'].width = 5.00
    sheet.column_dimensions['E'].width = 13.91
    sheet.column_dimensions['F'].width = 13.20
    sheet.column_dimensions['G'].width = 13.91
    sheet.column_dimensions['H'].width = 13.20

    next_row, normal_total_coord, normal_mo_coord = _write_summary_section(sheet, summary_normal, price_data, 'RESUMO', 1, bold_font, header_fill, currency_format, regular_font)
    next_row, distratado_total_coord, distratado_mo_coord = _write_summary_section(sheet, summary_distratado, price_data, 'SERVIÇOS DISTRATADOS', next_row, bold_font, header_fill, currency_format, regular_font)
    
    diff_row = next_row
    diff_label_cell = sheet.cell(row=diff_row, column=1, value='DIFERENÇA (NORMAL - DISTRATADO)')
    diff_label_cell.font = bold_font
    
    # Merge E and F for diff_total_cell
    sheet.merge_cells(start_row=diff_row, start_column=5, end_row=diff_row, end_column=6) # Merge E:F
    diff_total_cell = sheet.cell(row=diff_row, column=5, value=f"={normal_total_coord}-{distratado_total_coord}") # Moved to E (column 5)
    diff_total_cell.number_format = currency_format
    diff_total_cell.alignment = Alignment(horizontal='right', vertical='center')
    diff_total_cell.fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")
    diff_total_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # White font
    for col_idx in range(5, 7): # Columns E to F
        sheet.cell(row=diff_row, column=col_idx).fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")

    diff_mo_cell = sheet.cell(row=diff_row, column=8, value=f"={normal_mo_coord}-{distratado_mo_coord}") # This is column H
    diff_mo_cell.number_format = currency_format

    # Merge G and H for diff_mo_cell
    sheet.merge_cells(start_row=diff_row, start_column=7, end_row=diff_row, end_column=8) # Merge G:H
    diff_mo_cell = sheet.cell(row=diff_row, column=7, value=f"={normal_mo_coord}-{distratado_mo_coord}") # Moved to G (column 7)
    diff_mo_cell.number_format = currency_format
    diff_mo_cell.alignment = Alignment(horizontal='right', vertical='center')
    diff_mo_cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
    diff_mo_cell.font = Font(name='Verdana', color="FFFFFF", size=8) # White font
    for col_idx in range(7, 9): # Columns G to H
        sheet.cell(row=diff_row, column=col_idx).fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")

    diff_label_cell.fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")
    sheet.merge_cells(start_row=diff_row, start_column=1, end_row=diff_row, end_column=4)
    diff_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    # Apply the fill to the merged cells as well
    for col_idx in range(1, 5): # Columns A to D
        sheet.cell(row=diff_row, column=col_idx).fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")

    apply_borders_to_range(sheet, diff_row, 1, diff_row, 8)

def write_client_sheet(sheet, client_normal, client_distratado, price_data, bold_font, header_fill, currency_format, regular_font):
    sheet.column_dimensions['A'].width = 12.29
    sheet.column_dimensions['B'].width = 43.37
    sheet.column_dimensions['C'].width = 10.14
    sheet.column_dimensions['D'].width = 5.00
    sheet.column_dimensions['E'].width = 13.91
    sheet.column_dimensions['F'].width = 13.20
    
    current_row = 2
    normal_client_total_coords = {} # Dicionário para armazenar os totais normais por cliente
    
    # --- Bloco 1: Serviços Normais ---
    header_cell = sheet.cell(row=current_row, column=1, value="BLOCO 1: SERVIÇOS NORMAIS")
    header_cell.font = bold_font
    header_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    obra_total_cells = []
    
    sorted_clients_normal = sorted(client_normal.keys(), key=natural_sort_key)
    for client_id in sorted_clients_normal:
        client_header_cell = sheet.cell(row=current_row, column=1, value=client_id)
        client_header_cell.font = bold_font
        client_header_cell.fill = PatternFill(start_color="e46c0a", end_color="e46c0a", fill_type="solid")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        client_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        current_row, subtotal_cells, _ = _write_client_section(sheet, client_normal[client_id], price_data, current_row, bold_font, header_fill, currency_format, regular_font)
        total_formula = f"=+{'+'.join(subtotal_cells)}" if subtotal_cells else "0"
        
        total_row = current_row
        total_label_cell = sheet.cell(row=total_row, column=1, value=f'TOTAL {client_id}')
        total_label_cell.font = bold_font
        
        total_fill = PatternFill(start_color="e46c0a", end_color="e46c0a", fill_type="solid")

        # Format A:D
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, 5):
            sheet.cell(row=total_row, column=col_idx).fill = total_fill

        # Format E:F
        sheet.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=6)
        total_value_cell = sheet.cell(row=total_row, column=5, value=total_formula)
        total_value_cell.number_format = currency_format
        total_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_value_cell.fill = total_fill
        total_value_cell.font = regular_font
        
        total_coord = total_value_cell.coordinate
        obra_total_cells.append(total_coord)
        normal_client_total_coords[client_id] = total_coord # Armazena a coordenada do total
        current_row += 1
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1

    total_obra_row = current_row
    total_obra_label_cell = sheet.cell(row=total_obra_row, column=1, value='VALOR TOTAL DA OBRA')
    total_obra_label_cell.font = bold_font
    
    total_obra_fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")

    # Format A:D
    sheet.merge_cells(start_row=total_obra_row, start_column=1, end_row=total_obra_row, end_column=4)
    total_obra_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    for col_idx in range(1, 5):
        sheet.cell(row=total_obra_row, column=col_idx).fill = total_obra_fill

    # Format E:F
    sheet.merge_cells(start_row=total_obra_row, start_column=5, end_row=total_obra_row, end_column=6)
    total_obra_value_cell = sheet.cell(row=total_obra_row, column=5, value=f"=+{'+'.join(obra_total_cells)}")
    total_obra_value_cell.number_format = currency_format
    total_obra_value_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_obra_value_cell.fill = total_obra_fill
    total_obra_value_cell.font = bold_font
    
    total_obra_coord = total_obra_value_cell.coordinate
    current_row += 1
    format_empty_row(sheet, current_row, regular_font)
    current_row += 1

    # --- Bloco 2: Serviços Distratados ---
    header_cell_distratado = sheet.cell(row=current_row, column=1, value="BLOCO 2: SERVIÇOS DISTRATADOS")
    header_cell_distratado.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)
    header_cell_distratado.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    header_cell_distratado.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    obra_total_distratado_cells = []

    sorted_clients_distratado = sorted(client_distratado.keys(), key=natural_sort_key)
    for client_id in sorted_clients_distratado:
        client_header_cell = sheet.cell(row=current_row, column=1, value=client_id)
        client_header_cell.font = bold_font
        client_header_cell.fill = PatternFill(start_color="e46c0a", end_color="e46c0a", fill_type="solid")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        client_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        current_row, subtotal_cells, _ = _write_client_section(sheet, client_distratado[client_id], price_data, current_row, bold_font, header_fill, currency_format, regular_font)
        total_formula = f"=+{'+'.join(subtotal_cells)}" if subtotal_cells else "0"
        
        total_distratado_row = current_row
        total_distratado_label_cell = sheet.cell(row=total_distratado_row, column=1, value=f'TOTAL DISTRATADO {client_id}')
        total_distratado_label_cell.font = bold_font
        
        total_distratado_fill = PatternFill(start_color="93cddd", end_color="93cddd", fill_type="solid")

        # Format A:D
        sheet.merge_cells(start_row=total_distratado_row, start_column=1, end_row=total_distratado_row, end_column=4)
        total_distratado_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, 5):
            sheet.cell(row=total_distratado_row, column=col_idx).fill = total_distratado_fill

        # Format E:F
        sheet.merge_cells(start_row=total_distratado_row, start_column=5, end_row=total_distratado_row, end_column=6)
        total_distratado_value_cell = sheet.cell(row=total_distratado_row, column=5, value=total_formula)
        total_distratado_value_cell.number_format = currency_format
        total_distratado_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_distratado_value_cell.fill = total_distratado_fill
        total_distratado_value_cell.font = regular_font
        
        distratado_coord = total_distratado_value_cell.coordinate
        obra_total_distratado_cells.append(distratado_coord)
        current_row += 1 # Advance row after total distracted

        format_empty_row(sheet, current_row, regular_font)
        current_row += 1 # Add an extra row increment for the blank line

        # Adiciona a linha de SALDO para o cliente
        saldo_row = current_row
        normal_coord = normal_client_total_coords.get(client_id, "0")
        
        saldo_label_cell = sheet.cell(row=saldo_row, column=1, value=f"SALDO {client_id}")
        saldo_label_cell.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)
        
        saldo_fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")

        # Format A:D
        sheet.merge_cells(start_row=saldo_row, start_column=1, end_row=saldo_row, end_column=4)
        saldo_label_cell.alignment = Alignment(horizontal='left', vertical='center')
        for col_idx in range(1, 5):
            sheet.cell(row=saldo_row, column=col_idx).fill = saldo_fill

        # Format E:F
        sheet.merge_cells(start_row=saldo_row, start_column=5, end_row=saldo_row, end_column=6)
        saldo_value_cell = sheet.cell(row=saldo_row, column=5, value=f"={normal_coord}-{distratado_coord}")
        saldo_value_cell.number_format = currency_format
        saldo_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        saldo_value_cell.fill = saldo_fill
        saldo_value_cell.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)
        
        current_row += 1 # Advances the row after writing SALDO

    format_empty_row(sheet, current_row, regular_font)
    current_row += 1
    # --- Bloco 3: Saldo Final ---
    saldo_final_row = current_row
    saldo_final_label_cell = sheet.cell(row=saldo_final_row, column=1, value='VALOR TOTAL FINAL (OBRA - DISTRATADO)')
    saldo_final_label_cell.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)
    
    saldo_final_fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")

    # Format A:D
    sheet.merge_cells(start_row=saldo_final_row, start_column=1, end_row=saldo_final_row, end_column=4)
    saldo_final_label_cell.alignment = Alignment(horizontal='left', vertical='center')
    for col_idx in range(1, 5):
        sheet.cell(row=saldo_final_row, column=col_idx).fill = saldo_final_fill

    # Format E:F
    sheet.merge_cells(start_row=saldo_final_row, start_column=5, end_row=saldo_final_row, end_column=6)
    
    # Calculate the distratado total directly in the formula
    distratado_total_formula = f"({'+'.join(obra_total_distratado_cells)})" if obra_total_distratado_cells else "0"
    saldo_final_value_cell = sheet.cell(row=saldo_final_row, column=5, value=f"={total_obra_coord}-{distratado_total_formula}")
    saldo_final_value_cell.number_format = currency_format
    saldo_final_value_cell.alignment = Alignment(horizontal='right', vertical='center')
    saldo_final_value_cell.fill = saldo_final_fill
    saldo_final_value_cell.font = Font(name='Verdana', color="FFFFFF", bold=True, size=8)

def _write_client_section(sheet, items_by_key, price_data, start_row, bold_font, header_fill, currency_format, regular_font):
    COLS = {'TIPO': 1, 'DESC': 2, 'METRAGEM': 3, 'UN': 4, 'VALOR_UNIT': 5, 'VALOR_TOTAL': 6}
    current_row = start_row
    client_subtotal_cells, item_key_to_cell_map = [], {}

    cats = {'Forro': [], 'Paredes_e_Revestimentos': [], 'Guias e Montantes': []}
    for item_key, item_data in items_by_key.items():
        category = 'Paredes_e_Revestimentos' if item_data.get('Categoria') in ['Parede', 'Revestimento'] else item_data.get('Categoria')
        if category in cats: cats[category].append((item_key, item_data))

    for cat_name, cat_label in [('Forro', 'Forros'), ('Paredes_e_Revestimentos', 'Paredes e Revestimentos'), ('Guias e Montantes', 'Guias e Montantes')]:
        if not cats.get(cat_name): continue
        header_row_data = [('Tipo R. Bassani', COLS['TIPO']), (cat_label, COLS['DESC']), ('Metragem', COLS['METRAGEM']), ('Un', COLS['UN']), ('Valor do Material + MO', COLS['VALOR_UNIT']), ('Valor Total', COLS['VALOR_TOTAL'])]
        for val, col_idx in header_row_data:
            cell = sheet.cell(row=current_row, column=col_idx, value=val)
            cell.font, cell.fill = bold_font, header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['VALOR_TOTAL']) # Add borders
        current_row += 1
        section_start_row, ft16_placeholder_cell, ft16_base_quantity, subtraction_cell_list = current_row, None, 0, []

        if cat_name == 'Paredes_e_Revestimentos':
            sorted_items = sorted(cats[cat_name], key=lambda x: (0 if x[1]['Tipo Code'].startswith('TP') else 1, natural_sort_key(x[1]['Tipo Code'])))
        else: # Forros, Guias e Montantes
            sorted_items = sorted(cats[cat_name], key=lambda x: natural_sort_key(x[1]['Tipo Code']))
        
        for item_key, item in sorted_items:
            price_info = price_data.get(item['Tipo Code'], {'Valor': 0, 'Un': ''})
            main_item_row, is_subclass_item = current_row, item.get('is_subclass', False)

            # Calculate number of sub-items for merging
            num_sub_items = len(item.get('insulation_items', {})) + len(item.get('carenagem_items', {}))

            tipo_cell = sheet.cell(row=main_item_row, column=COLS['TIPO'], value=item['Tipo Code'])
            tipo_cell.alignment = Alignment(horizontal='center', vertical='center')
            tipo_cell.font = regular_font

            # Merge cell if there are sub-items
            if num_sub_items > 0:
                sheet.merge_cells(
                    start_row=main_item_row,
                    start_column=COLS['TIPO'],
                    end_row=main_item_row + num_sub_items,
                    end_column=COLS['TIPO']
                )

            desc_cell = sheet.cell(row=main_item_row, column=COLS['DESC'], value=item['Descricao'])
            desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
            desc_cell.font = regular_font
            
            m_cell_ref = f"{get_column_letter(COLS['METRAGEM'])}{main_item_row}"
            item_key_to_cell_map[item_key] = m_cell_ref
            
            metragem_cell = sheet.cell(row=main_item_row, column=COLS['METRAGEM'])
            metragem_cell.number_format = currency_format
            metragem_cell.font = regular_font
            metragem_cell.alignment = Alignment(vertical='center')
            
            # GARANTIR que a BaseQuantity seja sempre escrita como valor numérico inicial
            metragem_cell.value = item.get('BaseQuantity', 0) # Adicionar esta linha

            if 'formula_contributors' in item:
                parts = [f"({item_key_to_cell_map[c['item_key']]}*{c['count']})" if c['count'] > 1 else item_key_to_cell_map[c['item_key']] for c in item['formula_contributors'] if c['item_key'] in item_key_to_cell_map]
                final_formula = f"=CEILING(({'+'.join(parts)})*{item['formula_multiplier']},1)" if parts else item['BaseQuantity']
                metragem_cell.value = final_formula
            
            un_cell = sheet.cell(row=main_item_row, column=COLS['UN'], value=price_info['Un'])
            un_cell.font = regular_font
            un_cell.alignment = Alignment(horizontal='center', vertical='center')
            valor_unit_cell = sheet.cell(row=main_item_row, column=COLS['VALOR_UNIT'], value=price_info['Valor'])
            valor_unit_cell.number_format = currency_format
            valor_unit_cell.font = regular_font
            valor_unit_cell.alignment = Alignment(vertical='center')
            v_unit_cell = f"{get_column_letter(COLS['VALOR_UNIT'])}{main_item_row}"
            valor_total_cell = sheet.cell(row=main_item_row, column=COLS['VALOR_TOTAL'], value=f"={m_cell_ref}*{v_unit_cell}")
            valor_total_cell.number_format = currency_format
            valor_total_cell.font = regular_font
            valor_total_cell.alignment = Alignment(vertical='center')
            if is_subclass_item: subtraction_cell_list.append(m_cell_ref)
            if item['Tipo Code'] == 'FT16': ft16_placeholder_cell, ft16_base_quantity = m_cell_ref, item.get('FormulaBase', item['BaseQuantity'])
            apply_borders_to_range(sheet, main_item_row, COLS['TIPO'], main_item_row, COLS['VALOR_TOTAL']) # Add borders
            current_row += 1

            # --- SUB-ITEM LOOPS ---
            insulation_cell_references = [] # Initialized here for each main item
            
            # ISOLAMENTO
            for sub_item_key, sub_item in sorted(item.get('insulation_items', {}).items()):
                price_info_sub = price_data.get(sub_item['Tipo Code'], {'Valor': 0, 'Un': ''})
                sub_desc_cell = sheet.cell(row=current_row, column=COLS['DESC'], value=sub_item['Descricao'])
                sub_desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
                sub_desc_cell.font = regular_font
                
                sub_m_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'])
                sub_m_cell.number_format = currency_format
                sub_m_cell.font = regular_font
                sub_m_cell.alignment = Alignment(vertical='center')

                if sub_item.get('is_la_dupla'):
                    sub_m_cell.value = f"={m_cell_ref}*2"
                else:
                    sub_m_cell.value = sub_item['Quantidade']
                    insulation_cell_references.append(sub_m_cell.coordinate) # Only append if not 'lã dupla'
                
                # Add lines to write 'Un' and 'Valor do Material + MO'
                sub_un_cell = sheet.cell(row=current_row, column=COLS['UN'], value=price_info_sub['Un'])
                sub_un_cell.font = regular_font
                sub_un_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                sub_v_unit_cell_obj = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=price_info_sub['Valor'])
                sub_v_unit_cell_obj.number_format = currency_format
                sub_v_unit_cell_obj.font = regular_font
                sub_v_unit_cell_obj.alignment = Alignment(vertical='center')
                
                sub_v_unit_cell = f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}"
                sub_v_total_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"={sub_m_cell.coordinate}*{sub_v_unit_cell}")
                sub_v_total_cell.number_format = currency_format
                sub_v_total_cell.font = regular_font
                sub_v_total_cell.alignment = Alignment(vertical='center')
                apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['VALOR_TOTAL']) # Add borders
                current_row += 1

            # CARNAGEM
            for sub_item_key, sub_item in sorted(item.get('carenagem_items', {}).items()):
                price_info_sub = price_data.get(sub_item['Tipo Code'], {'Valor': 0, 'Un': ''})
                sub_desc_cell = sheet.cell(row=current_row, column=COLS['DESC'], value=sub_item['Descricao'])
                sub_desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
                sub_desc_cell.font = regular_font
                sub_m_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'], value=sub_item['Quantidade'])
                sub_m_cell.number_format = currency_format
                sub_m_cell.font = regular_font
                sub_m_cell.alignment = Alignment(vertical='center')
                sub_un_cell = sheet.cell(row=current_row, column=COLS['UN'], value=price_info_sub['Un'])
                sub_un_cell.font = regular_font
                sub_un_cell.alignment = Alignment(horizontal='center', vertical='center')
                sub_v_unit_cell_obj = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=price_info_sub['Valor'])
                sub_v_unit_cell_obj.number_format = currency_format
                sub_v_unit_cell_obj.font = regular_font
                sub_v_unit_cell_obj.alignment = Alignment(vertical='center')
                sub_v_unit_cell = f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}"
                sub_v_total_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"={sub_m_cell.coordinate}*{sub_v_unit_cell}")
                sub_v_total_cell.number_format = currency_format
                sub_v_total_cell.font = regular_font
                sub_v_total_cell.alignment = Alignment(vertical='center')
                apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['VALOR_TOTAL']) # Add borders
                current_row += 1

            # --- FORMULA LOGIC FOR MAIN ITEM (AFTER SUB-ITEMS ARE PROCESSED) ---
            if cat_name != 'Guias e Montantes' and not is_subclass_item and item['Tipo Code'] != 'FT16':
                final_metragem_value = item.get('BaseQuantity') # Default to BaseQuantity

                if item.get('logic_applied'): # Merged items (BASE + LÃ)
                    formula_parts = []
                    base_val = item.get('FormulaBase', item['BaseQuantity'])
                    if base_val > 0:
                        formula_parts.append(str(base_val))
                    if insulation_cell_references: # Should have one reference
                        formula_parts.extend(insulation_cell_references)
                    
                    if len(formula_parts) > 1:
                        final_metragem_value = f"={'+'.join(formula_parts)}"
                    elif len(formula_parts) == 1:
                        # Check if it's a cell reference (starts with a letter)
                        if re.match(r'^[A-Z]', formula_parts[0]):
                            final_metragem_value = f"={formula_parts[0]}"
                        else:
                            final_metragem_value = formula_parts[0] # Just the base_val
                elif item.get('insulation_items'): # Non-merged item with insulation (just LÃ)
                    if insulation_cell_references: # Should have one reference
                        final_metragem_value = f"={insulation_cell_references[0]}" # Only the insulation cell ref
                    else:
                        final_metragem_value = item.get('BaseQuantity') # Fallback if no insulation ref found (shouldn't happen)
                else: # Item without insulation, not merged (just BASE)
                    final_metragem_value = item.get('BaseQuantity') # Use the numeric value directly (no '=')

                metragem_cell.value = final_metragem_value
        
        # --- FT16 FORMULA (AFTER ALL ITEMS IN CATEGORY ARE PROCESSED) ---
        if ft16_placeholder_cell and subtraction_cell_list:
            base_val_str = str(ft16_base_quantity).replace(',', '.')
            sheet[ft16_placeholder_cell] = f"={base_val_str}-{'-'.join(subtraction_cell_list)}"
            sheet[ft16_placeholder_cell].number_format = currency_format
            sheet[ft16_placeholder_cell].font = regular_font

        # --- SUBTOTALS ---
        subtotal_row_idx = current_row
        subtotal_cell = sheet.cell(row=subtotal_row_idx, column=COLS['TIPO'], value='SUB-TOTAL')
        subtotal_cell.font = bold_font

        # Apply formatting for A:D
        subtotal_fill = PatternFill(start_color="ffff99", end_color="ffff99", fill_type="solid")
        subtotal_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=subtotal_row_idx, start_column=1, end_row=subtotal_row_idx, end_column=4)

        for col_idx in range(1, 5):
            sheet.cell(row=subtotal_row_idx, column=col_idx).fill = subtotal_fill

        # Merge E:F and apply formula
        sheet.merge_cells(start_row=subtotal_row_idx, start_column=COLS['VALOR_UNIT'], end_row=subtotal_row_idx, end_column=COLS['VALOR_TOTAL'])
        
        total_col_letter = get_column_letter(COLS['VALOR_TOTAL'])
        formula = f"=SUM({total_col_letter}{section_start_row}:{total_col_letter}{current_row - 1})"
        
        merged_value_cell = sheet.cell(row=subtotal_row_idx, column=COLS['VALOR_UNIT'], value=formula)
        merged_value_cell.number_format = currency_format
        merged_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_value_cell.fill = subtotal_fill
        merged_value_cell.font = bold_font
        
        apply_borders_to_range(sheet, subtotal_row_idx, COLS['TIPO'], subtotal_row_idx, COLS['VALOR_TOTAL']) # Add borders
        client_subtotal_cells.append(merged_value_cell.coordinate)
        current_row += 1
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1
    
    return current_row, client_subtotal_cells, item_key_to_cell_map

# ==============================================================================
# AJUSTES E REGRAS DE NEGÓCIO
# ==============================================================================

def apply_wool_logic(data_by_client):
    for client_id, client_items in data_by_client.items():
        grouped_keys = {}
        for key in list(client_items.keys()):
            tipo_code, categoria, has_insulation = key
            base_key = (tipo_code, categoria)
            if base_key not in grouped_keys: grouped_keys[base_key] = []
            grouped_keys[base_key].append(key)

        for keys in grouped_keys.values():
            if len(keys) == 2:
                key_with_wool = next((k for k in keys if k[2]), None)
                key_without_wool = next((k for k in keys if not k[2]), None)
                
                if key_with_wool and key_without_wool:
                    item_com_la, item_sem_la = client_items[key_with_wool], client_items[key_without_wool]
                    original_qty_com_la, qty_sem_la = item_com_la['BaseQuantity'], item_sem_la['BaseQuantity']
                    item_com_la['FormulaBase'] = qty_sem_la
                    item_com_la['BaseQuantity'] = qty_sem_la + original_qty_com_la
                    if item_com_la['insulation_items']:
                        insul_key = next(iter(item_com_la['insulation_items']))
                        if not item_com_la['insulation_items'][insul_key].get('is_la_dupla', False):
                            item_com_la['insulation_items'][insul_key]['Quantidade'] = original_qty_com_la
                    item_com_la['logic_applied'] = True
                    item_com_la['carenagem_items'].update(item_sem_la.get('carenagem_items', {}))
                    del client_items[key_without_wool]
    return data_by_client

def calculate_and_add_derived_items(data_by_client, price_data):
    known_profiles = ['MS48', 'MS70', 'MS90', 'F47', 'MS140']
    rules = { 'SHF-40': {'m': 0.35, 'f': ['Ru', '/400'], 'c': True}, 'SH48-40': {'m': 0.45, 'f': ['MS48', '/400'], 'c': True}, 'SH48-60': {'m': 0.45, 'f': ['MS48', '/600'], 'c': True},
              'SH70-40': {'m': 0.45, 'f': ['MS70', '/400'], 'c': True}, 'SH70-60': {'m': 0.45, 'f': ['MS70', '/600'], 'c': True}, 'SH90-40': {'m': 0.45, 'f': ['MS90', '/400'], 'c': True},
              'SH90-60': {'m': 0.45, 'f': ['MS90', '/600'], 'c': True}, 'SH140-40': {'m': 0.45, 'f': ['MS140', '/400'], 'c': True}, 'SH140-60': {'m': 0.45, 'f': ['MS140', '/600'], 'c': True} }

    for client_id, client_items in data_by_client.items():
        contributions = {code: [] for code in rules}
        for item_key, item in client_items.items():
            if item.get('Categoria') in ['Parede', 'Revestimento'] and item.get('Tipo Code', '').startswith('TP') and item.get('Descricao') and item.get('BaseQuantity', 0) > 0:
                for code, rule in rules.items():
                    if all(f in item['Descricao'] for f in rule['f']):
                        count = item['Descricao'].count('Ru') if code == 'SHF-40' else (item['Descricao'].count(next((p for p in known_profiles if re.search(r'\d+', code).group(0) in p), None)) if rule['c'] and next((p for p in known_profiles if re.search(r'\d+', code).group(0) in p), None) else 1)
                        count = 1 if count == 0 else count
                        if count > 0: contributions[code].append({'item_key': item_key, 'count': count})
        
        for code, contribs in contributions.items():
            if contribs:
                total_contrib_val = sum(client_items[c['item_key']].get('BaseQuantity', 0) * c['count'] for c in contribs)
                final_qty = math.ceil(total_contrib_val * rules[code]['m'])
                if final_qty > 0:
                    data_by_client[client_id][(code, 'Guias e Montantes', False)] = {
                        'Tipo Code': code, 'Descricao': price_data.get(code, {}).get('Descricao', f'Montante {code}'), 'BaseQuantity': final_qty, 'Categoria': 'Guias e Montantes',
                        'has_insulation': False, 'insulation_items': {}, 'carenagem_items': {}, 'is_subclass': False,
                        'formula_contributors': contribs, 'formula_multiplier': rules[code]['m']
                    }
    return data_by_client

# ==============================================================================
# BLOCO DE EXECUÇÃO PRINCIPAL
# ==============================================================================

if __name__ == '__main__':
    print("Carregando dados de preços e subclasses...")
    precos = load_price_data()
    subclasses = load_subclass_data()

    for scope in ['normal', 'distratado']:
        print(f"Processando dados para: {scope.upper()}...")
        prefix = '__' if scope == 'distratado' else ''
        client_data = process_client_data(
            precos, subclasses,
            f'{prefix}Tabela de Forro.csv',
            f'{prefix}Tabela de Modelo Genérico.csv',
            f'{prefix}Tabela de Paredes.csv'
        )
        print("  Aplicando lógicas de negócio...")
        client_data = apply_wool_logic(client_data)
        client_data = calculate_and_add_derived_items(client_data, precos)
        summary_data = process_summary_data(client_data)
        
        if scope == 'normal':
            dados_cliente_normal, dados_resumo_normal = client_data, summary_data
        else:
            dados_cliente_distratado, dados_resumo_distratado = client_data, summary_data

    print("Gerando arquivo Excel...")
    write_excel_with_formulas(
        dados_resumo_normal, dados_resumo_distratado,
        dados_cliente_normal, dados_cliente_distratado,
        precos,
        filename='Relatorios_Com_Formulas.xlsx'
    )
