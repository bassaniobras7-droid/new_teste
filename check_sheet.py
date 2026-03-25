import openpyxl
wb = openpyxl.load_workbook('Relatorios_Com_Formulas.xlsx', data_only=True)
print('sheets', wb.sheetnames)
sh = wb['Aditivos x Distrato']
for r in range(1, 7):
    print([sh.cell(row=r, column=c).value for c in range(1, 14)])
