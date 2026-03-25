import openpyxl

wb = openpyxl.load_workbook('Relatorios_Com_Formulas.xlsx')
ws = wb['Aditivos x Distrato']

print("=" * 80)
print("ESTRUTURA: Aditivos x Distrato")
print("=" * 80)

# Verificar headers
print("\nLinha 2 (Headers):")
print(f"  Left (A2:F2): {[ws.cell(2, col).value for col in range(1, 7)]}")
print(f"  Right (H2:M2): {[ws.cell(2, col).value for col in range(8, 14)]}")

# Ver as primeiras linhas de dados
print("\nPrimeiras linhas de dados (linhas 3-15):")
for row in range(3, 16):
    left_data = [ws.cell(row, col).value for col in range(1, 7)]
    right_data = [ws.cell(row, col).value for col in range(8, 14)]
    left_str = f"A:{left_data[0][:15]+'...' if isinstance(left_data[0], str) and len(left_data[0])>15 else left_data[0]} | B:{left_data[1]} | F:{left_data[5]}" if left_data[0] else ""
    right_str = f"H:{right_data[0][:15]+'...' if isinstance(right_data[0], str) and len(right_data[0])>15 else right_data[0]} | I:{right_data[1]} | M:{right_data[5]}" if right_data[0] else ""
    print(f"  Row {row}: LEFT[{left_str}] | RIGHT[{right_str}]")

# Encontrar as linhas com "TOTAL"
print("\nLinhas com TOTAL:")
for row in range(3, ws.max_row + 1):
    val = ws.cell(row, 1).value
    if val and 'TOTAL' in str(val).upper():
        print(f"  Row {row}: {val}")

wb.close()
