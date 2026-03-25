import openpyxl
from openpyxl.utils import get_column_letter
import sys

ANTIGO = r"C:/Users/Wesllen.Santana/Downloads/bassani/Relatorios_Com_Formulas_dec2.xlsx"
NOVO   = r"C:/Users/Wesllen.Santana/Downloads/bassani/Relatorios_Com_Formulas_novo_dec2.xlsx"

wb_a = openpyxl.load_workbook(ANTIGO)
wb_n = openpyxl.load_workbook(NOVO)

print("=== ABAS ===")
print(f"ANTIGO: {wb_a.sheetnames}")
print(f"NOVO:   {wb_n.sheetnames}")
print()

def rgb(color_obj):
    if color_obj is None:
        return None
    try:
        t = color_obj.type
        if t == 'rgb':
            val = color_obj.rgb
            return val if val and val != '00000000' else None
        if t == 'theme':
            return f"THEME:{color_obj.theme}"
        if t == 'indexed':
            return f"INDEXED:{color_obj.indexed}"
    except:
        pass
    return None

def fill_info(cell):
    f = cell.fill
    if f is None or f.fill_type is None or f.fill_type == 'none':
        return None
    fg = rgb(f.fgColor) if f.fgColor else None
    bg = rgb(f.bgColor) if f.bgColor else None
    if not fg and not bg:
        return None
    return {
        'type': f.fill_type,
        'fg': fg,
        'bg': bg,
    }

def font_info(cell):
    f = cell.font
    if f is None:
        return None
    return {
        'name': f.name,
        'size': f.size,
        'bold': f.bold,
        'italic': f.italic,
        'color': rgb(f.color) if f.color else None,
        'underline': f.underline,
    }

def align_info(cell):
    a = cell.alignment
    if a is None:
        return None
    d = {
        'horizontal': a.horizontal,
        'vertical': a.vertical,
        'wrap': a.wrap_text,
    }
    if all(v is None or v is False for v in d.values()):
        return None
    return d

def border_side(s):
    if s is None or s.border_style is None:
        return None
    return {'style': s.border_style, 'color': rgb(s.color) if s.color else None}

def border_info(cell):
    b = cell.border
    if b is None:
        return None
    sides = {
        'left': border_side(b.left),
        'right': border_side(b.right),
        'top': border_side(b.top),
        'bottom': border_side(b.bottom),
    }
    if all(v is None for v in sides.values()):
        return None
    return sides

def sheet_summary(ws):
    info = {}
    info['max_row'] = ws.max_row
    info['max_col'] = ws.max_column

    col_widths = {}
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.width is not None:
            col_widths[col_letter] = round(col_dim.width, 2)
    info['col_widths'] = col_widths

    row_heights = {}
    for row_idx, row_dim in ws.row_dimensions.items():
        if row_dim.height is not None:
            row_heights[row_idx] = round(row_dim.height, 2)
    info['row_heights'] = row_heights

    info['merged_cells'] = sorted([str(m) for m in ws.merged_cells.ranges])

    ps = ws.page_setup
    info['page_setup'] = {
        'orientation': ps.orientation,
        'paper_size': ps.paperSize,
        'fit_to_width': ps.fitToWidth,
        'fit_to_height': ps.fitToHeight,
        'scale': ps.scale,
    }
    info['print_area'] = ws.print_area
    info['freeze_panes'] = str(ws.freeze_panes) if ws.freeze_panes else None

    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            coord = cell.coordinate
            val = cell.value
            fi = fill_info(cell)
            fo = font_info(cell)
            al = align_info(cell)
            br = border_info(cell)
            nf = cell.number_format if cell.number_format != 'General' else None

            if val is not None or fi or br:
                cells[coord] = {
                    'value': str(val) if val is not None else None,
                    'number_format': nf,
                    'fill': fi,
                    'font': fo,
                    'align': al,
                    'border': br,
                }
    info['cells'] = cells
    return info

sheets_a = {name: sheet_summary(wb_a[name]) for name in wb_a.sheetnames}
sheets_n = {name: sheet_summary(wb_n[name]) for name in wb_n.sheetnames}

all_names = list(dict.fromkeys(wb_a.sheetnames + wb_n.sheetnames))

print("=" * 80)
print("RELATORIO DE DIFERENCAS")
print("=" * 80)

for sheet_name in all_names:
    print(f"\n\n{'#' * 80}")
    print(f"# ABA: {sheet_name}")
    print(f"{'#' * 80}")

    if sheet_name not in sheets_a:
        print(f"  [INFO] Aba existe APENAS no arquivo NOVO")
        continue
    if sheet_name not in sheets_n:
        print(f"  [INFO] Aba existe APENAS no arquivo ANTIGO")
        continue

    sa = sheets_a[sheet_name]
    sn = sheets_n[sheet_name]

    # Dimensoes
    if sa['max_row'] != sn['max_row'] or sa['max_col'] != sn['max_col']:
        print(f"\n--- DIMENSOES ---")
        print(f"  ANTIGO: {sa['max_row']} linhas x {sa['max_col']} colunas")
        print(f"  NOVO:   {sn['max_row']} linhas x {sn['max_col']} colunas")

    # Larguras de colunas
    all_cols = sorted(set(list(sa['col_widths'].keys()) + list(sn['col_widths'].keys())))
    col_diffs = {}
    for col in all_cols:
        wa = sa['col_widths'].get(col)
        wn = sn['col_widths'].get(col)
        if wa != wn:
            col_diffs[col] = (wa, wn)
    if col_diffs:
        print(f"\n--- LARGURAS DE COLUNAS ---")
        for col in sorted(col_diffs.keys()):
            wa, wn = col_diffs[col]
            print(f"  Coluna {col}: ANTIGO={wa}  -->  NOVO={wn}")

    # Alturas de linhas
    all_rows = sorted(set(list(sa['row_heights'].keys()) + list(sn['row_heights'].keys())))
    row_diffs = {}
    for row in all_rows:
        ha = sa['row_heights'].get(row)
        hn = sn['row_heights'].get(row)
        if ha != hn:
            row_diffs[row] = (ha, hn)
    if row_diffs:
        print(f"\n--- ALTURAS DE LINHAS ---")
        for row in sorted(row_diffs.keys()):
            ha, hn = row_diffs[row]
            print(f"  Linha {row:3d}: ANTIGO={ha}  -->  NOVO={hn}")

    # Merged cells
    ma = set(sa['merged_cells'])
    mn = set(sn['merged_cells'])
    if ma != mn:
        print(f"\n--- MESCLAGENS DE CELULAS ---")
        only_a = sorted(ma - mn)
        only_n = sorted(mn - ma)
        if only_a:
            print(f"  REMOVER (existem apenas no ANTIGO):")
            for m in only_a:
                print(f"    {m}")
        if only_n:
            print(f"  ADICIONAR (existem apenas no NOVO):")
            for m in only_n:
                print(f"    {m}")

    # Page setup
    psa = sa['page_setup']
    psn = sn['page_setup']
    ps_diffs = {}
    for k in set(list(psa.keys()) + list(psn.keys())):
        if psa.get(k) != psn.get(k):
            ps_diffs[k] = (psa.get(k), psn.get(k))
    if ps_diffs or sa['print_area'] != sn['print_area'] or sa['freeze_panes'] != sn['freeze_panes']:
        print(f"\n--- CONFIGURACOES DE PAGINA ---")
        for k, (va, vn) in ps_diffs.items():
            print(f"  {k}: ANTIGO={va}  -->  NOVO={vn}")
        if sa['print_area'] != sn['print_area']:
            print(f"  print_area: ANTIGO={sa['print_area']}  -->  NOVO={sn['print_area']}")
        if sa['freeze_panes'] != sn['freeze_panes']:
            print(f"  freeze_panes: ANTIGO={sa['freeze_panes']}  -->  NOVO={sn['freeze_panes']}")

    # Celulas
    all_coords = set(list(sa['cells'].keys()) + list(sn['cells'].keys()))
    # Ordenar por linha depois coluna
    def sort_key(coord):
        import re
        m = re.match(r'([A-Z]+)(\d+)', coord)
        if m:
            return (int(m.group(2)), m.group(1))
        return (0, coord)

    cell_diffs = []
    for coord in sorted(all_coords, key=sort_key):
        ca = sa['cells'].get(coord, {})
        cn = sn['cells'].get(coord, {})
        diffs = []

        # Value
        va = ca.get('value')
        vn = cn.get('value')
        if va != vn:
            diffs.append(f"valor: [{repr(va)}] --> [{repr(vn)}]")

        # Number format
        nfa = ca.get('number_format')
        nfn = cn.get('number_format')
        if nfa != nfn:
            diffs.append(f"formato_numero: [{nfa}] --> [{nfn}]")

        # Fill
        fa = ca.get('fill')
        fn = cn.get('fill')
        if fa != fn:
            diffs.append(f"fill: {fa} --> {fn}")

        # Font
        fta = ca.get('font')
        ftn = cn.get('font')
        if fta != ftn:
            font_d = []
            for k in set(list((fta or {}).keys()) + list((ftn or {}).keys())):
                va2 = (fta or {}).get(k)
                vn2 = (ftn or {}).get(k)
                if va2 != vn2:
                    font_d.append(f"{k}: {va2} --> {vn2}")
            if font_d:
                diffs.append(f"font: [{'; '.join(font_d)}]")

        # Align
        ala = ca.get('align')
        aln = cn.get('align')
        if ala != aln:
            al_d = []
            for k in set(list((ala or {}).keys()) + list((aln or {}).keys())):
                va2 = (ala or {}).get(k)
                vn2 = (aln or {}).get(k)
                if va2 != vn2:
                    al_d.append(f"{k}: {va2} --> {vn2}")
            if al_d:
                diffs.append(f"align: [{'; '.join(al_d)}]")

        # Border
        bra = ca.get('border')
        brn = cn.get('border')
        if bra != brn:
            br_d = []
            for side in ['left', 'right', 'top', 'bottom']:
                sa2 = (bra or {}).get(side)
                sn2 = (brn or {}).get(side)
                if sa2 != sn2:
                    br_d.append(f"{side}: {sa2} --> {sn2}")
            if br_d:
                diffs.append(f"border: [{'; '.join(br_d)}]")

        if diffs:
            cell_diffs.append((coord, diffs))

    if cell_diffs:
        print(f"\n--- CELULAS COM DIFERENCAS ({len(cell_diffs)} celulas) ---")
        for coord, diffs in cell_diffs:
            print(f"\n  Celula {coord}:")
            for d in diffs:
                print(f"    * {d}")

print("\n\n" + "=" * 80)
print("FIM DO RELATORIO")
print("=" * 80)
