import openpyxl
wb=openpyxl.load_workbook('Relatorios_Com_Formulas_test.xlsx', data_only=True)
ws=wb['Aditivos x Distrato']
print('row2 left', [ws.cell(row=2, column=c).value for c in range(1,7)])
print('row2 right', [ws.cell(row=2, column=c).value for c in range(8,14)])
print('row3 left', [ws.cell(row=3, column=c).value for c in range(1,7)])
print('row3 right', [ws.cell(row=3, column=c).value for c in range(8,14)])
