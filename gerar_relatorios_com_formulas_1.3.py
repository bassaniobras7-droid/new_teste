import pandas as pd
import traceback
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont # Adicionar esta linha
import math
import re
import os
import glob
import msoffcrypto

# Definições globais para colunas
COLS = {'TIPO': 1, 'DESC': 2, 'METRAGEM': 3, 'UN': 4, 'VALOR_UNIT': 5, 'VALOR_TOTAL': 6, 'CUSTO_MO_UNIT': 9, 'CUSTO_MO_TOTAL': 10, 'VALOR_FECHADO': 11, 'VALOR_TOTAL_NOVO': 12}
VISIBLE_COLS_END = 12 # A última coluna visível é a L (12)

# ==============================================================================
# FUNÇÕES AUXILIARES DE FORMATAÇÃO
# ==============================================================================

def format_bold_text_for_excel(text_content, regular_font, bold_font):
    """
    Converte uma string com marcação **texto_negrito** em um objeto CellRichText do openpyxl,
    aplicando a fonte em negrito às partes marcadas.
    """
    rich_text_list = CellRichText()
    parts = re.split(r'(\*\*.*?\*\*)', text_content)

    # Criar InlineFont a partir dos atributos de Font
    inline_regular_font = InlineFont(
        rFont=regular_font.name, # CORREÇÃO: Usar rFont em vez de name
        sz=regular_font.sz,
        b=regular_font.b,
        i=regular_font.i,
        u=regular_font.u,
        strike=regular_font.strike,
        color=regular_font.color,
        # Adicionar outros atributos relevantes se necessário
    )
    inline_bold_font = InlineFont(
        rFont=bold_font.name, # CORREÇÃO: Usar rFont em vez de name
        sz=bold_font.sz,
        b=bold_font.b,
        i=bold_font.i,
        u=bold_font.u,
        strike=bold_font.strike,
        color=bold_font.color,
        # Adicionar outros atributos relevantes se necessário
    )

    for part in parts:
        if part.startswith('**') and part.endswith('**'):
            rich_text_list.append(TextBlock(font=inline_bold_font, text=part[2:-2]))
        else:
            rich_text_list.append(TextBlock(font=inline_regular_font, text=part))
    return rich_text_list

def apply_borders_to_range(sheet, min_row, min_col, max_row, max_col):
    thin_border = Border(left=Side(border_style="thin", color="000000"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
    for row in sheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = thin_border

def format_empty_row(sheet, row_idx, regular_font):
    sheet.row_dimensions[row_idx].height = 5
    sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    # Add a fill for visibility
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # A cor de preenchimento dos espaçamento na cor branca
    sheet.cell(row=row_idx, column=1).fill = white_fill

def estimate_rows_for_text(text, sheet, start_column, end_column, font_size=8):
    """
    Estima o número de linhas necessárias para um texto, usando uma heurística
    e aplicando uma correção com base no feedback do usuário.
    """
    total_excel_width = 0
    for col_idx in range(start_column, end_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in sheet.column_dimensions and sheet.column_dimensions[col_letter].width:
            total_excel_width += sheet.column_dimensions[col_letter].width
        else:
            total_excel_width += 8.43 # Fallback para largura padrão do Excel
    
    # Estimar caracteres que cabem em uma linha visual, com base na largura das colunas mescladas.
    # Usar um fator de ~1.2 para Verdana 8 é uma estimativa razoável, pois é mais compacta que o '0' padrão.
    chars_per_visual_line = int(total_excel_width * 1.5) 
    
    if chars_per_visual_line <= 0:
        return 1

    lines = text.split('\n')
    total_estimated_lines = 0
    for line in lines:
        if len(line) == 0: # Linhas vazias contam como 1 linha
            total_estimated_lines += 1
        else:
            # Para cada linha do texto, estimar quantas linhas visuais ela ocupa devido ao wrap_text.
            total_estimated_lines += math.ceil(len(line) / chars_per_visual_line)
    
    # Aplicar a correção com base no feedback do usuário:
    # A estimativa está consistentemente 12 linhas acima do necessário (6 para cima, 6 para baixo).
    correction_factor = 12
    final_lines = total_estimated_lines - correction_factor
    
    return max(1, final_lines)
    
# ==============================================================================
# FUNÇÕES AUXILIARES E DE CARREGAMENTO DE DADOS
# ==============================================================================

def clean_numeric_column(series):
    # Garantir que valores numéricos venham corretos mesmo se contiverem espaços ou formatos mistos
    # Ex: ' 1,40 ', '1.4', '1,4'
    cleaned = series.astype(str).str.strip().str.replace(',', '.', regex=False)
    return pd.to_numeric(cleaned, errors='coerce').fillna(0)

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', str(s))]


def normalize_block_torre_id(value: str) -> str:
    """Normaliza o valor de ID. Bloco/Torre para comparação independente de caixa e acentuação."""
    s = str(value or "").strip().lower()
    # Normalizar acentos mais comuns usados nos IDs (ex: Térreo)
    s = s.replace('é', 'e').replace('ê', 'e').replace('á', 'a').replace('à', 'a').replace('ã', 'a')
    return s


def sort_client_id_key(client_id):
    """Ordena IDs de bloco/torre com prioridade fixa para Subsolo, Térreo e PVTO/Pavimento.

    Para PVTO/Pavimento, tenta extrair o número do pavimento (ex: "1º PVTO", "PVTO-1")
    para garantir a ordem correta (1, 2, 3, ...).
    """
    normalized = normalize_block_torre_id(client_id)

    # Ordenação prioritária (sempre no início):
    # 1) Subsolo
    # 2) Térreo
    # 3) PVTO / Pavimento
    if 'subsolo' in normalized:
        return (0, 0, '')
    if 'terreo' in normalized:
        return (1, 0, '')

    # PVTO/Pavimento com extração numérica (1º PVTO, 2º PVTO, PVTO-1, etc.)
    if re.search(r'\b(pvto|pavto)\b', normalized):
        # Procura número antes de PVTO ou depois (PVTO-1, PVTO 1)
        num_match = re.search(r'(\d+)\s*º?\s*(?=\b(pvto|pavto)\b)', normalized) or \
                    re.search(r'\b(pvto|pavto)\b\W*(\d+)', normalized)
        num = int(num_match.group(1)) if num_match else 0
        return (2, num, natural_sort_key(client_id))

    return (3, 0, natural_sort_key(client_id))


def load_price_data(filename='Valores Ctba.csv'):
    try:
        df_prices = pd.read_csv(filename, sep=';', header=0)
        df_prices.dropna(subset=['Tipo R. Bassani'], inplace=True)
        price_data = {}
        for _, row in df_prices.iterrows():
            tipo_code = row['Tipo R. Bassani']
            price_data[tipo_code] = {
                'Un': row.get('Un'),
                'Valor': clean_numeric_column(pd.Series([row.get('Valor do Material + MO')]))[0],
                'Custo MO': clean_numeric_column(pd.Series([row.get('Custo MO à Pagar')]))[0],
                'Descricao': row.get('Forros')
            }
        return price_data
    except FileNotFoundError:
        print(f"AVISO: Arquivo de preços '{filename}' não encontrado.")
        return {}
    except Exception as e:
        print(f"ERRO ao carregar o arquivo de preços: {e}")
        return {}

def load_subclass_data(filename='Subclasse.csv'):
    try:
        df_subclass = pd.read_csv(filename, sep=';', header=0)
        df_subclass.dropna(subset=['Tipo'], inplace=True)
        return set(df_subclass['Tipo'].astype(str).str.strip())
    except FileNotFoundError:
        print(f"AVISO: Arquivo de subclasses '{filename}' não encontrado.")
        return set()
    except Exception as e:
        print(f"ERRO ao carregar o arquivo de subclasses: {e}")
        return set()

def find_latest_excel_file(prefix):
    """
    Procura pelo arquivo Excel mais recente com o prefixo dado no diretório atual.
    Retorna o caminho do arquivo ou None se nenhum arquivo for encontrado.
    """
    search_pattern = f"{prefix}*.xlsx"
    files = glob.glob(search_pattern)
    
    if not files:
        print(f"AVISO: Nenhum arquivo Excel encontrado com o prefixo '{prefix}'.")
        return None
    
    # Ordenar arquivos pela data de modificação, do mais recente para o mais antigo
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]

# ==============================================================================
# PROCESSAMENTO DE DADOS
# ==============================================================================

def process_summary_data(client_data):
    summary_agg = {}
    for client_id, client_items in client_data.items():
        subclass_total_qty = sum(item_data.get('BaseQuantity', 0) for item_data in client_items.values() if item_data.get('is_subclass', False))

        for item_data in client_items.values():
            cat = item_data['Categoria']
            summary_cat = 'Paredes' if cat in ['Parede', 'Revestimento'] else ('Forros' if cat == 'Forro' else cat)

            for carenagem_item in item_data.get('carenagem_items', {}).values():
                carenagem_tipo_code = carenagem_item['Tipo Code']
                if carenagem_tipo_code not in summary_agg:
                    summary_agg[carenagem_tipo_code] = {'Quantidade': 0, 'Descricao': carenagem_item['Descricao'], 'Categoria': summary_cat}
                summary_agg[carenagem_tipo_code]['Quantidade'] += carenagem_item['Quantidade']

            for insulation_item in item_data.get('insulation_items', {}).values():
                insul_tipo_code = insulation_item['Tipo Code']
                if insul_tipo_code not in summary_agg:
                    summary_agg[insul_tipo_code] = {'Quantidade': 0, 'Descricao': insulation_item['Descricao'], 'Categoria': 'Isolamento'}
                summary_agg[insul_tipo_code]['Quantidade'] += insulation_item['Quantidade']

            main_tipo_code = item_data['Tipo Code']
            qty = item_data['BaseQuantity']
            if main_tipo_code == 'FT16' and subclass_total_qty > 0:
                qty -= subclass_total_qty
            
            if main_tipo_code not in summary_agg:
                summary_agg[main_tipo_code] = {'Quantidade': 0, 'Descricao': item_data['Descricao'], 'Categoria': summary_cat}
            summary_agg[main_tipo_code]['Quantidade'] += qty
    return summary_agg

def process_client_data(price_data, subclass_types, excel_file_path, forro_sheet, generico_sheet, paredes_sheet):
    data_by_client = {}

    def add_or_aggregate_item(client_id, item_data, has_insulation=False, insulation_data=None, carenagem_data=None, perfil_data=None):
        if client_id not in data_by_client: data_by_client[client_id] = {}
        key = (item_data['Tipo Code'], item_data['Categoria'], has_insulation)
        if key not in data_by_client[client_id]:
            data_by_client[client_id][key] = {
                'Tipo Code': item_data['Tipo Code'], 'Descricao': item_data.get('Descricao'), 'BaseQuantity': 0,
                'Categoria': item_data['Categoria'], 'has_insulation': has_insulation, 'insulation_items': {},
                'carenagem_items': {}, 'perfil_items': {}, 'is_subclass': item_data.get('is_subclass', False)
            }
        if item_data.get('Descricao') and not data_by_client[client_id][key].get('Descricao'):
            data_by_client[client_id][key]['Descricao'] = item_data.get('Descricao')
        data_by_client[client_id][key]['BaseQuantity'] += item_data.get('Quantidade', 0)
        if item_data.get('is_subclass', False): data_by_client[client_id][key]['is_subclass'] = True
        if insulation_data:
            insul_key = insulation_data['Tipo Code']
            if insul_key not in data_by_client[client_id][key]['insulation_items']:
                data_by_client[client_id][key]['insulation_items'][insul_key] = {'Tipo Code': insul_key, 'Descricao': insulation_data['Descricao'], 'Quantidade': 0, 'is_la_dupla': insulation_data.get('is_la_dupla', False)}
            data_by_client[client_id][key]['insulation_items'][insul_key]['Quantidade'] += insulation_data['Quantidade']
        if carenagem_data:
            carenagem_key = carenagem_data['Descricao']
            if carenagem_key not in data_by_client[client_id][key]['carenagem_items']:
                data_by_client[client_id][key]['carenagem_items'][carenagem_key] = {'Tipo Code': carenagem_data['Tipo Code'], 'Descricao': carenagem_data['Descricao'], 'Quantidade': 0}
            data_by_client[client_id][key]['carenagem_items'][carenagem_key]['Quantidade'] += carenagem_data['Quantidade']
        if perfil_data:
            perfil_key = perfil_data['Tipo Code']
            if perfil_key not in data_by_client[client_id][key]['perfil_items']:
                data_by_client[client_id][key]['perfil_items'][perfil_key] = {'Tipo Code': perfil_key, 'Descricao': perfil_data['Descricao'], 'Quantidade': 0}
            data_by_client[client_id][key]['perfil_items'][perfil_key]['Quantidade'] += perfil_data['Quantidade']

    for sheet_name, process_func in [(forro_sheet, 'forro'), (generico_sheet, 'generico'), (paredes_sheet, 'paredes')]:
        try:
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=1)
            df.dropna(subset=['ID. Bloco/Torre'], inplace=True)
            if process_func == 'forro':
                df['Área'] = clean_numeric_column(df['Área'])
                df['Perímetro'] = clean_numeric_column(df['Perímetro'])
                for _, row in df.iterrows():
                    tipo_code = str(row['Sistema Construtivo R. Bassani']).strip()
                    unit = price_data.get(tipo_code, {}).get('Un', '')
                    qty = row['Área'] if unit == 'm²' else (row['Perímetro'] if unit == 'm' else row['Área'])
                    add_or_aggregate_item(row['ID. Bloco/Torre'], {'Tipo Code': tipo_code, 'Descricao': price_data.get(tipo_code, {}).get('Descricao', row['Tipo']), 'Quantidade': qty, 'Categoria': 'Forro', 'is_subclass': tipo_code in subclass_types})
            elif process_func == 'generico':
                df['Contador'] = clean_numeric_column(df['Contador'])
                for _, row in df.iterrows():
                    tipo_code = str(row['Sistema Construtivo R. Bassani']).strip()
                    add_or_aggregate_item(row['ID. Bloco/Torre'], {'Tipo Code': tipo_code, 'Descricao': price_data.get(tipo_code, {}).get('Descricao', row['Tipo']), 'Quantidade': row['Contador'], 'Categoria': 'Parede' if row['Classe'] == 'Parede' else 'Forro', 'is_subclass': tipo_code in subclass_types})
            elif process_func == 'paredes':
                df['Área'] = clean_numeric_column(df['Área'])
                df['Altura desconectada'] = clean_numeric_column(df['Altura desconectada'])
                df['Comprimento'] = clean_numeric_column(df['Comprimento'])
                for _, row in df.iterrows():
                    client_id, tipo_code, categoria = row['ID. Bloco/Torre'], str(row['Sistema Construtivo R. Bassani']).strip(), row['Classe']
                    is_subclass, osb_perfil = tipo_code in subclass_types, str(row.get('OSB/Perfil')).strip()
                    is_carenagem = osb_perfil == 'Carenagem'
                    is_la_dupla = str(row.get('OSB/Perfil')).strip() == 'Lã dupla'
                    unit, row_quantity = price_data.get(tipo_code, {}).get('Un', ''), 0
                    if categoria in ['Parede', 'Revestimento']:
                        if unit == 'm²':
                            row_quantity = row['Área']
                        elif unit == 'm':
                            row_quantity = row['Altura desconectada']
                        else:
                            row_quantity = 0
                    elif categoria == 'Forro':
                        if unit == 'm²':
                            row_quantity = row['Área']
                        elif unit == 'm':
                            row_quantity = row['Comprimento']
                        else:
                            row_quantity = 0
                    item_data, carenagem_data, insulation_data = None, None, None
                    has_insulation = pd.notna(row['Sistema de Isolamento']) and str(row['Sistema de Isolamento']).strip() != ''
                    if is_carenagem:
                        tipo_code_carenagem = f"{tipo_code}-CAR"
                        desc = price_data.get(tipo_code_carenagem, {}).get('Descricao', row['Tipo'])
                        # Check if 'desc' already ends with '(Carenagem)'
                        if not desc.strip().endswith('(Carenagem)'):
                            desc = f"{desc} (Carenagem)"
                        carenagem_data = {'Tipo Code': tipo_code_carenagem, 'Descricao': desc, 'Quantidade': row_quantity}
                        item_data = {'Tipo Code': tipo_code, 'Categoria': categoria, 'Quantidade': 0, 'is_subclass': is_subclass}
                    else:
                        item_data = {'Tipo Code': tipo_code, 'Descricao': price_data.get(tipo_code, {}).get('Descricao', row['Tipo']), 'Quantidade': row_quantity, 'Categoria': categoria, 'is_subclass': is_subclass}
                        if has_insulation:
                            tipo_code_insul = row['Sistema de Isolamento'].strip()
                            insulation_data = {'Tipo Code': tipo_code_insul, 'Descricao': price_data.get(tipo_code_insul, {}).get('Descricao', f"Isolamento {tipo_code_insul}"), 'Quantidade': row['Área'] * (2 if is_la_dupla else 1), 'is_la_dupla': is_la_dupla}
                    add_or_aggregate_item(client_id, item_data, has_insulation=has_insulation, insulation_data=insulation_data, carenagem_data=carenagem_data)
        except Exception as e:
            print(f"AVISO: Erro ao ler o arquivo Excel '{excel_file_path}', aba '{sheet_name}': {e}")
    return data_by_client

# ==============================================================================
# GERAÇÃO DO ARQUIVO EXCEL
# ==============================================================================

def write_excel_with_formulas(summary_normal, summary_distratado, client_normal, client_distratado, price_data, filename, observacoes_gerais_content=""):
    workbook = openpyxl.Workbook()
    bold_font = Font(name='Verdana', bold=True, size=8)
    regular_font = Font(name='Verdana', size=8)
    bold_white_font = Font(name='Verdana', bold=True, size=8, color="FFFFFF") # Nova
    regular_white_font = Font(name='Verdana', size=8, color="FFFFFF") # Nova
    header_fill = PatternFill(start_color="99ccff", end_color="99ccff", fill_type="solid")
    header_fill_kl = PatternFill(start_color="33cccc", end_color="33cccc", fill_type="solid")
    # Usar formato padrão Excel (inglês) para que o Excel converta para o formato local ao abrir
    currency_format = '#,##0.00'
    # Formato contábil com símbolo R$ à esquerda e valores alinhados à direita
    accounting_format = '_("R$"* #,##0.00_);_("R$"* -#,##0.00_);_("R$"* "-"??_);_(@_)'

    ws_cliente = workbook.active
    ws_cliente.title = 'Cliente'
    ws_cliente.sheet_properties.tabColor = "FF00FF" # Rosa / Magenta
    write_client_sheet(ws_cliente, client_normal, client_distratado, price_data, bold_font, header_fill, currency_format, accounting_format, regular_font, bold_white_font, regular_white_font)
    
    ws_resumo = workbook.create_sheet(title='Resumo')
    ws_resumo.sheet_properties.tabColor = "666699" # Azul arroxeado
    write_summary_sheet(ws_resumo, summary_normal, summary_distratado, price_data, bold_font, header_fill, currency_format, accounting_format, regular_font, header_fill_kl, observacoes_gerais_content, bold_white_font, regular_white_font)

    ws_aditivos_distrato = workbook.create_sheet(title='Aditivos x Distrato')
    ws_aditivos_distrato.sheet_properties.tabColor = "00B050" # Verde
    write_aditivos_distrato_sheet(ws_aditivos_distrato, summary_normal, summary_distratado, price_data, bold_font, regular_font, currency_format, accounting_format, header_fill, bold_white_font, regular_white_font, client_normal, client_distratado)

    ws_relacao_media = workbook.create_sheet(title='Relação Média Material') # Nova aba
    ws_relacao_media.sheet_properties.tabColor = "FF0000" # Vermelho
    write_relacao_media_material_sheet(ws_relacao_media, bold_font, regular_font, currency_format, bold_white_font, regular_white_font) # Nova função para preencher a aba
    
    try:
        from openpyxl.styles.fills import Fill
        _default_fill = PatternFill()
        workbook._fills[:] = [f if isinstance(f, Fill) else _default_fill for f in workbook._fills]
        workbook._number_formats[:] = [f if isinstance(f, str) else '' for f in workbook._number_formats]
        workbook.save(filename)
        print(f"Arquivo '{filename}' gerado com sucesso.")

        # Reabrir e salvar novamente para corrigir possíveis problemas de XML
        workbook.close()
        wb_temp = openpyxl.load_workbook(filename)
        wb_temp.save(filename)
        wb_temp.close()

        # Proteger o arquivo com senha usando msoffcrypto-tool
        password = "tecnicob" #007
        temp_encrypted_filename = f"{filename}.encrypted"

        with open(filename, "rb") as f_in:
            officefile = msoffcrypto.OfficeFile(f_in)
            with open(temp_encrypted_filename, "wb") as f_out:
                officefile.encrypt(password=password, outfile=f_out)
        
        # Substituir o arquivo original pelo arquivo criptografado
        os.remove(filename)
        os.rename(temp_encrypted_filename, filename)
        print(f"Arquivo '{filename}' protegido com senha com sucesso.")

    except Exception as e:
        print(f"Erro ao salvar ou proteger o arquivo Excel: {e}")
        traceback.print_exc()

def _write_summary_section(sheet, summary_data, price_data, title, start_row, bold_font, header_fill, currency_format, accounting_format, regular_font, header_fill_kl, bold_white_font=None, regular_white_font=None):
    # Solicitação: N -> I, I -> J. Cabeçalho de I: "Custo MO".

    
    title_cell = sheet.cell(row=start_row, column=1, value=title)
    title_cell.font = bold_font
    
    if title == 'RESUMO':
        title_cell.fill = PatternFill(start_color="b3a2c7", end_color="b3a2c7", fill_type="solid")
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        apply_borders_to_range(sheet, start_row, 1, start_row, 6)
        
        subtitle_row = start_row + 1
        subtitle_cell = sheet.cell(row=subtitle_row, column=1, value="SERVIÇOS ADICIONAIS (NOVO LAYOUT)")
        subtitle_cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8)
        subtitle_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        sheet.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=6)
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, 6 + 1): # Changed to 6
            sheet.cell(row=subtitle_row, column=col_idx).fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
        apply_borders_to_range(sheet, subtitle_row, 1, subtitle_row, 6)

        current_row = start_row + 2
    elif title == 'SERVIÇOS DISTRATADOS':
        title_cell.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(color="FFFFFF", bold=True, name='Verdana', size=8)
        for col_idx in range(1, 6 + 1): # Alterado para preencher apenas até a coluna F
            sheet.cell(row=start_row, column=col_idx).fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
        apply_borders_to_range(sheet, start_row, 1, start_row, 6) # Alterado para aplicar bordas apenas até a coluna F
        current_row = start_row + 1
    else:
        current_row = start_row + 1

    subtotal_valor_cells, subtotal_mo_cells, subtotal_valor_novo_cells = [], [], []

    def summary_sort_key_with_tp_priority(item):
        tipo_code = item[0]
        is_car = tipo_code.endswith('-CAR')
        base_code = tipo_code.replace('-CAR', '')
        tp_priority = 0 if base_code.startswith('TP') else 1
        natural_base_sort = (tp_priority, natural_sort_key(base_code))
        car_suffix_priority = 1 if is_car else 0
        return (natural_base_sort, car_suffix_priority)

    for categoria, nome_categoria in [('Forros', 'Forros'), ('Paredes', 'Paredes e Revestimentos'), ('Isolamento', 'Isolamento Acústico'), ('Guias e Montantes', 'Guias e Montantes')]:
        items_cat = {k: v for k, v in summary_data.items() if v.get('Categoria') == categoria}
        if not items_cat: continue

        header_row = [ ('Tipo R. Bassani', COLS['TIPO']), (nome_categoria, COLS['DESC']), ('Metragem', COLS['METRAGEM']), ('Un', COLS['UN']),
                       ('Valor do Material + MO', COLS['VALOR_UNIT']), ('Valor Total', COLS['VALOR_TOTAL']), ('', 7), ('', 8), 
                       ('Custo MO', COLS['CUSTO_MO_UNIT']), ('Valor Total MO', COLS['CUSTO_MO_TOTAL']),
                       ('Valor Fechado', COLS['VALOR_FECHADO']), ('Valor Total', COLS['VALOR_TOTAL_NOVO']) ]
        for val, col in header_row:
            cell = sheet.cell(row=current_row, column=col, value=val)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if val in ['Custo MO', 'Valor Total MO']:
                cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
                cell.font = bold_white_font
            elif col in [COLS['VALOR_FECHADO'], COLS['VALOR_TOTAL_NOVO']]:
                cell.fill = header_fill_kl
            else:
                cell.fill = header_fill
        
        apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, VISIBLE_COLS_END)

        current_row += 1
        section_start_row = current_row

        if categoria == 'Paredes':
            sorted_items = sorted(items_cat.items(), key=summary_sort_key_with_tp_priority)
        elif categoria == 'Isolamento':
            sorted_items = sorted(items_cat.items(), key=lambda item: (0 if item[0].startswith('TP') else 1, natural_sort_key(item[0])))
        else:
            sorted_items = sorted(items_cat.items(), key=lambda item: natural_sort_key(item[0]))

        for tipo_code, data in sorted_items:
            price_info = price_data.get(tipo_code, {'Valor': 0, 'Un': '', 'Custo MO': 0})
            
            # Colunas Visíveis A-F
            sheet.cell(row=current_row, column=COLS['TIPO'], value=tipo_code).font = regular_font
            sheet.cell(row=current_row, column=COLS['TIPO']).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=current_row, column=COLS['DESC'], value=data['Descricao']).font = regular_font
            sheet.cell(row=current_row, column=COLS['DESC']).alignment = Alignment(wrap_text=True)
            metragem_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'], value=round(data['Quantidade'], 2))
            metragem_cell.number_format = currency_format
            metragem_cell.alignment = Alignment(vertical='center')
            metragem_cell.font = regular_font
            sheet.cell(row=current_row, column=COLS['UN'], value=price_info['Un']).font = regular_font
            sheet.cell(row=current_row, column=COLS['UN']).alignment = Alignment(horizontal='center', vertical='center')
            valor_fechado_cell_coord = f"{get_column_letter(COLS['VALOR_FECHADO'])}{current_row}"
            valor_unit_cell = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=f"={valor_fechado_cell_coord}")
            valor_unit_cell.number_format = currency_format
            valor_unit_cell.alignment = Alignment(vertical='center')
            valor_unit_cell.font = regular_font
            
            m_cell_coord = f"{get_column_letter(COLS['METRAGEM'])}{current_row}"
            v_unit_cell_coord = f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}"

            valor_total_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"=ROUND({m_cell_coord}*{v_unit_cell_coord}, 2)")
            valor_total_cell.number_format = accounting_format
            
            # Colunas Visíveis I e J
            custo_mo_unit_cell = sheet.cell(row=current_row, column=COLS['CUSTO_MO_UNIT'], value=round(price_info['Custo MO'], 2))
            custo_mo_unit_cell.number_format = currency_format
            custo_mo_unit_cell.fill = PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")

            custo_mo_total_cell = sheet.cell(row=current_row, column=COLS['CUSTO_MO_TOTAL'], value=f"=ROUND({m_cell_coord}*{custo_mo_unit_cell.coordinate}, 2)")
            custo_mo_total_cell.number_format = currency_format
            custo_mo_total_cell.fill = PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")

            # Novas colunas K e L
            valor_fechado_cell = sheet.cell(row=current_row, column=COLS['VALOR_FECHADO'], value=round(price_info['Valor'], 2))
            valor_fechado_cell.number_format = currency_format
            valor_fechado_cell.fill = PatternFill(start_color="99ccff", end_color="99ccff", fill_type="solid")
            
            valor_total_novo_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL_NOVO'], value=f"=ROUND({m_cell_coord}*{valor_fechado_cell.coordinate}, 2)")
            valor_total_novo_cell.number_format = currency_format
            valor_total_novo_cell.fill = PatternFill(start_color="99ccff", end_color="99ccff", fill_type="solid")

            # Estilos de fonte e alinhamento
            for col in [COLS['VALOR_TOTAL'], COLS['CUSTO_MO_UNIT'], COLS['CUSTO_MO_TOTAL'], COLS['VALOR_FECHADO'], COLS['VALOR_TOTAL_NOVO']]:
                cell = sheet.cell(row=current_row, column=col)
                cell.font = regular_font
                cell.alignment = Alignment(vertical='center')

            apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, VISIBLE_COLS_END)
            
            current_row += 1
            
        subtotal_row = current_row
        subtotal_cell = sheet.cell(row=subtotal_row, column=COLS['TIPO'], value='SUB-TOTAL')
        subtotal_cell.font = bold_font
        subtotal_fill = PatternFill(start_color="ffff99", end_color="ffff99", fill_type="solid")
        subtotal_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=4)
        for col_idx in range(1, 5): sheet.cell(row=subtotal_row, column=col_idx).fill = subtotal_fill

        valor_total_col = get_column_letter(COLS['VALOR_TOTAL'])
        formula_total = f"=ROUND(SUM({valor_total_col}{section_start_row}:{valor_total_col}{current_row - 1}), 2)"
        
        mo_col = get_column_letter(COLS['CUSTO_MO_TOTAL'])
        formula_mo = f"=ROUND(SUM({mo_col}{section_start_row}:{mo_col}{current_row - 1}), 2)"

        valor_total_novo_col = get_column_letter(COLS['VALOR_TOTAL_NOVO'])
        formula_total_novo = f"=ROUND(SUM({valor_total_novo_col}{section_start_row}:{valor_total_novo_col}{current_row - 1}), 2)"

        sheet.merge_cells(start_row=subtotal_row, start_column=COLS['VALOR_UNIT'], end_row=subtotal_row, end_column=COLS['VALOR_TOTAL'])
        merged_value_cell = sheet.cell(row=subtotal_row, column=COLS['VALOR_UNIT'], value=formula_total)
        merged_value_cell.number_format = accounting_format
        merged_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_value_cell.fill = subtotal_fill
        merged_value_cell.font = bold_font
        subtotal_valor_cells.append(merged_value_cell.coordinate)

        sheet.merge_cells(start_row=subtotal_row, start_column=7, end_row=subtotal_row, end_column=COLS['CUSTO_MO_TOTAL'])
        merged_mo_cell = sheet.cell(row=subtotal_row, column=7, value=formula_mo)
        merged_mo_cell.number_format = currency_format
        merged_mo_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_mo_cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
        merged_mo_cell.font = Font(name='Verdana', color="FFFFFF", size=8)
        subtotal_mo_cells.append(merged_mo_cell.coordinate)
        
        sheet.merge_cells(start_row=subtotal_row, start_column=COLS['VALOR_FECHADO'], end_row=subtotal_row, end_column=COLS['VALOR_TOTAL_NOVO'])
        merged_value_novo_cell = sheet.cell(row=subtotal_row, column=COLS['VALOR_FECHADO'], value=formula_total_novo)
        merged_value_novo_cell.number_format = currency_format
        merged_value_novo_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_value_novo_cell.fill = PatternFill(start_color="33cccc", end_color="33cccc", fill_type="solid")
        merged_value_novo_cell.font = bold_font
        subtotal_valor_novo_cells.append(merged_value_novo_cell.coordinate)

        apply_borders_to_range(sheet, subtotal_row, COLS['TIPO'], subtotal_row, VISIBLE_COLS_END)
        current_row += 1
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1

    total_row = current_row
    total_label_cell = sheet.cell(row=total_row, column=COLS['TIPO'], value='TOTAL DOS SERVIÇOS EXECUTADOS' if title == 'RESUMO' else f'TOTAL {title}')
    total_label_cell.font = bold_font
    total_formula = f"=ROUND({'+'.join(subtotal_valor_cells)}, 2)" if subtotal_valor_cells else "0"
    mo_formula = f"=ROUND({'+'.join(subtotal_mo_cells)}, 2)" if subtotal_mo_cells else "0"
    novo_valor_formula = f"=ROUND({'+'.join(subtotal_valor_novo_cells)}, 2)" if subtotal_valor_novo_cells else "0"
    
    fill_color_label = "77933c" if title == 'RESUMO' else "31859c"
    font_color = "FFFFFF"
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_label_cell.fill = PatternFill(start_color=fill_color_label, end_color=fill_color_label, fill_type="solid")
    total_label_cell.font = bold_white_font
    for col_idx in range(1, 5): sheet.cell(row=total_row, column=col_idx).fill = PatternFill(start_color=fill_color_label, end_color=fill_color_label, fill_type="solid")

    total_val_cell_obj = sheet.cell(row=total_row, column=COLS['VALOR_UNIT'], value=total_formula)
    total_val_cell_obj.number_format = accounting_format
    total_val_cell_obj.alignment = Alignment(horizontal='right', vertical='center')
    total_val_cell_obj.fill = PatternFill(start_color=fill_color_label, end_color=fill_color_label, fill_type="solid")
    total_val_cell_obj.font = regular_white_font
    final_total_val_coord = total_val_cell_obj.coordinate

    sheet.merge_cells(start_row=total_row, start_column=7, end_row=total_row, end_column=COLS['CUSTO_MO_TOTAL'])
    total_mo_cell_obj = sheet.cell(row=total_row, column=7, value=mo_formula)
    total_mo_cell_obj.number_format = currency_format
    total_mo_cell_obj.alignment = Alignment(horizontal='right', vertical='center')
    total_mo_cell_obj.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
    total_mo_cell_obj.font = regular_white_font
    final_total_mo_coord = total_mo_cell_obj.coordinate
    
    sheet.merge_cells(start_row=total_row, start_column=COLS['VALOR_FECHADO'], end_row=total_row, end_column=COLS['VALOR_TOTAL_NOVO'])
    total_val_novo_cell_obj = sheet.cell(row=total_row, column=COLS['VALOR_FECHADO'], value=novo_valor_formula) # Usa a nova formula para K e L
    total_val_novo_cell_obj.number_format = currency_format
    total_val_novo_cell_obj.alignment = Alignment(horizontal='right', vertical='center')
    total_val_novo_cell_obj.fill = PatternFill(start_color="33cccc", end_color="33cccc", fill_type="solid")
    total_val_novo_cell_obj.font = regular_white_font
    final_total_novo_coord = total_val_novo_cell_obj.coordinate
    
    apply_borders_to_range(sheet, total_row, COLS['TIPO'], total_row, VISIBLE_COLS_END)
    
    final_current_row = current_row + 1

    return final_current_row, final_total_val_coord, final_total_mo_coord, final_total_novo_coord

def write_summary_sheet(sheet, summary_normal, summary_distratado, price_data, bold_font, header_fill, currency_format, accounting_format, regular_font, header_fill_kl, observacoes_gerais_content="", bold_white_font=None, regular_white_font=None):
    sheet.column_dimensions['A'].width = 12.29
    sheet.column_dimensions['B'].width = 43.37
    sheet.column_dimensions['C'].width = 10.14
    sheet.column_dimensions['D'].width = 5.00
    sheet.column_dimensions['E'].width = 13.91
    sheet.column_dimensions['F'].width = 13.20 # VALOR_TOTAL
    sheet.column_dimensions['G'].hidden = True
    sheet.column_dimensions['H'].hidden = True
    sheet.column_dimensions['I'].width = 13.91 # CUSTO_MO_UNIT
    sheet.column_dimensions['J'].width = 13.20 # CUSTO_MO_TOTAL
    sheet.column_dimensions['K'].width = 13.91 # VALOR_FECHADO
    sheet.column_dimensions['L'].width = 13.20 # VALOR_TOTAL_NOVO

    # Oculta as colunas G e H conforme já definido.
    sheet.column_dimensions['G'].hidden = True
    sheet.column_dimensions['H'].hidden = True
    
    # Oculta as colunas de suporte que não são mais necessárias
    sheet.column_dimensions['N'].hidden = True # Antiga CUSTO_MO_UNIT, agora livre
    sheet.column_dimensions['P'].hidden = True # Totalizadores

    next_row, normal_total_coord, normal_mo_coord, normal_total_novo_coord = _write_summary_section(sheet, summary_normal, price_data, 'RESUMO', 1, bold_font, header_fill, currency_format, accounting_format, regular_font, header_fill_kl, bold_white_font, regular_white_font)
    
    # Adicionar uma linha em branco antes da seção 'SERVIÇOS DISTRATADOS'
    format_empty_row(sheet, next_row, regular_font)
    next_row += 1

    next_row, distratado_total_coord, distratado_mo_coord, distratado_total_novo_coord = _write_summary_section(sheet, summary_distratado, price_data, 'SERVIÇOS DISTRATADOS', next_row, bold_font, header_fill, currency_format, accounting_format, regular_font, header_fill_kl, bold_white_font, regular_white_font)
    
    diff_row = next_row
    diff_label_cell = sheet.cell(row=diff_row, column=1, value='VALOR TOTAL DA OBRA:\nVALOR À SER COBRADO COMO ESCOPO ADICIONAL RESULTANTE DA DIFERENÇA ENTRE NOVO LAYOUT E SERVIÇOS DISTRATADOS')
    diff_label_cell.font = bold_font
    
    # Diferença de VALOR TOTAL (visível em E:F)
    sheet.merge_cells(start_row=diff_row, start_column=5, end_row=diff_row, end_column=6)
    diff_total_cell = sheet.cell(row=diff_row, column=5, value=f"=ROUND({normal_total_coord}-{distratado_total_coord}, 2)")
    diff_total_cell.number_format = currency_format
    diff_total_cell.alignment = Alignment(horizontal='right', vertical='center')
    diff_total_cell.fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")
    diff_total_cell.font = regular_white_font

    # Diferença de CUSTO MO TOTAL (visível em G:J)
    sheet.merge_cells(start_row=diff_row, start_column=7, end_row=diff_row, end_column=10)
    diff_mo_cell = sheet.cell(row=diff_row, column=7, value=f"=ROUND({normal_mo_coord}-{distratado_mo_coord}, 2)")
    diff_mo_cell.number_format = currency_format
    diff_mo_cell.alignment = Alignment(horizontal='right', vertical='center')
    diff_mo_cell.fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
    diff_mo_cell.font = regular_white_font

    # Diferença de VALOR TOTAL NOVO (visível em K:L)
    sheet.merge_cells(start_row=diff_row, start_column=COLS['VALOR_FECHADO'], end_row=diff_row, end_column=COLS['VALOR_TOTAL_NOVO'])
    diff_total_novo_cell = sheet.cell(row=diff_row, column=COLS['VALOR_FECHADO'], value=f"=ROUND({normal_total_novo_coord}-{distratado_total_novo_coord}, 2)")
    diff_total_novo_cell.number_format = currency_format
    diff_total_novo_cell.alignment = Alignment(horizontal='right', vertical='center')
    diff_total_novo_cell.fill = PatternFill(start_color="33cccc", end_color="33cccc", fill_type="solid")
    diff_total_novo_cell.font = regular_white_font

    # Formatação do label da diferença (A:D)
    diff_label_fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")
    diff_label_cell.fill = diff_label_fill
    diff_label_cell.font = bold_white_font
    sheet.merge_cells(start_row=diff_row, start_column=1, end_row=diff_row, end_column=4)
    diff_label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx in range(1, 5):
        sheet.cell(row=diff_row, column=col_idx).fill = diff_label_fill

    apply_borders_to_range(sheet, diff_row, 1, diff_row, VISIBLE_COLS_END) # Borda até a coluna L

    # Adicionar linha em branco de espaçamento
    format_empty_row(sheet, diff_row + 1, regular_font)
    next_row = diff_row + 2 # Atualiza next_row para pular a linha em branco

    # Adicionar o bloco de observações gerais
    if observacoes_gerais_content: # Adiciona o bloco apenas se houver conteúdo
        obs_start_row = next_row
        num_required_lines = estimate_rows_for_text(observacoes_gerais_content, sheet, 1, 6)
        obs_end_row = obs_start_row + num_required_lines - 1 

        sheet.merge_cells(start_row=obs_start_row, start_column=1, end_row=obs_end_row, end_column=6)
        # Processar o conteúdo das observações gerais para aplicar negrito
        formatted_obs_content = format_bold_text_for_excel(observacoes_gerais_content, regular_font, bold_font)
        obs_cell = sheet.cell(row=obs_start_row, column=1, value=formatted_obs_content)
        obs_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # A fonte já está definida no objeto RichText, então não é necessário definir aqui.
        obs_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Branco

        apply_borders_to_range(sheet, obs_start_row, 1, obs_end_row, 6)
        last_content_row = obs_end_row
    else:
        last_content_row = diff_row # Última linha do total da diferença se não houver observações

    # --- Configurações Finais de Página (após todo o conteúdo ser adicionado) ---
    sheet.page_setup.paperSize = 9  # A4
    sheet.page_setup.orientation = 'portrait'
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1 # Ajustar para 1 página de largura
    sheet.page_setup.fitToHeight = 0 # Ajustar altura automaticamente
    sheet.page_setup.printArea = f'A1:F{last_content_row}' # Definir área de impressão para A1:F até a última linha de conteúdo gerado
    sheet.sheet_view.view = 'pageBreakPreview' # Abrir em Visualização da Quebra de Página
    # Removida a quebra de página de coluna após F, pois I-L não devem ser impressas separadamente


def write_aditivos_distrato_sheet(sheet, summary_normal, summary_distratado, price_data, bold_font, regular_font, header_fill, currency_format, accounting_format, bold_white_font=None, regular_white_font=None, client_normal=None, client_distratado=None):
    # Larguras baseadas no RB_10841 (ADD + DISTRATO)
    sheet.column_dimensions['A'].width = 11.71   # Tipo R. Bassani (ADD)
    sheet.column_dimensions['B'].width = 39.00   # Descrição (ADD)
    sheet.column_dimensions['C'].width = 10.29   # Metragem (ADD)
    sheet.column_dimensions['D'].width = 3.43    # Un (ADD)
    sheet.column_dimensions['E'].width = 13.86   # Valor Unit (ADD)
    sheet.column_dimensions['F'].width = 16.14   # Valor Total (ADD)
    sheet.column_dimensions['G'].width = 11.71   # Tipo R. Bassani (DIST)
    sheet.column_dimensions['H'].width = 37.71   # Descrição (DIST)
    sheet.column_dimensions['I'].width = 10.29   # Metragem (DIST)
    sheet.column_dimensions['J'].width = 3.43    # Un (DIST)
    sheet.column_dimensions['K'].width = 13.86   # Valor Unit (DIST)
    sheet.column_dimensions['L'].width = 14.29   # Valor Total (DIST)
    sheet.column_dimensions['M'].width = 10.14   # Diferença (aux, vermelho)

    # Paleta de cores RB_10841
    fill_titulo      = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')  # azul claro
    fill_add_label   = PatternFill(start_color='FBD4B4', end_color='FBD4B4', fill_type='solid')  # laranja muito claro
    fill_dis_label   = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')  # lilás muito claro
    fill_add_total_l = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')  # laranja claro (label total)
    fill_add_total_v = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')  # laranja palidíssimo (valor total)
    fill_dis_total_l = PatternFill(start_color='E5DFEC', end_color='E5DFEC', fill_type='solid')  # lilás (label total)
    fill_dis_total_v = PatternFill(start_color='CCC0D9', end_color='CCC0D9', fill_type='solid')  # lilás médio (valor total)
    font_red = Font(name='Verdana', size=8, color='FF0000')

    def _spacer(row, height=3.75):
        sheet.row_dimensions[row].height = height

    def normalize_category(cat):
        if cat == 'Forro': return 'Forro'
        if cat in ['Parede', 'Revestimento', 'Paredes_e_Revestimentos', 'Paredes e Revestimentos']:
            return 'Paredes e Revestimentos'
        return None

    def collect_items(block_id, dataset, category):
        result = {}
        if not dataset or block_id not in dataset: return result
        for item in dataset[block_id].values():
            if normalize_category(item.get('Categoria')) == category:
                result[item.get('Tipo Code', '')] = item
        return result

    current_row = 1
    all_proposta_l = []  # coordenadas L de cada "VALOR TOTAL DA PROPOSTA" por bloco

    blocks = sorted(set((client_normal or {}).keys()) | set((client_distratado or {}).keys()), key=natural_sort_key)

    for block_id in blocks:
        # --- Título do bloco (azul claro, A:L mesclado) ---
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
        tc = sheet.cell(row=current_row, column=1, value=block_id)
        tc.font = bold_font; tc.fill = fill_titulo
        tc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_borders_to_range(sheet, current_row, 1, current_row, 12)
        sheet.row_dimensions[current_row].height = 19.5
        current_row += 1

        _spacer(current_row); current_row += 1  # espaçador fino

        # --- Labels ADITIVOS / DISTRATO ---
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        al = sheet.cell(row=current_row, column=1, value='ADITIVOS')
        al.font = bold_font; al.fill = fill_add_label
        al.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_borders_to_range(sheet, current_row, 1, current_row, 6)

        sheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
        dl = sheet.cell(row=current_row, column=7, value='DISTRATO')
        dl.font = bold_font; dl.fill = fill_dis_label
        dl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_borders_to_range(sheet, current_row, 7, current_row, 12)
        sheet.row_dimensions[current_row].height = 19.5
        current_row += 1

        _spacer(current_row); current_row += 1  # espaçador fino

        # --- Categorias do bloco: Forro e Paredes e Revestimentos ---
        cat_total_f = []  # coordenadas F de totais de categoria
        cat_total_l = []  # coordenadas L de totais de categoria

        for cat_name, cat_label in [('Forro', 'Forros'), ('Paredes e Revestimentos', 'Paredes e Revestimentos')]:
            n_items = collect_items(block_id, client_normal or {}, cat_name)
            d_items = collect_items(block_id, client_distratado or {}, cat_name)
            if not n_items and not d_items:
                continue

            # Cabeçalho de colunas (col B = nome da categoria)
            col_hdrs = [
                ('Tipo R. Bassani', 1), (cat_label, 2), ('Metragem', 3), ('Un', 4),
                ('Valor do Material + MO', 5), ('Valor Total', 6)
            ]
            for text, col in col_hdrs:
                cell = sheet.cell(row=current_row, column=col, value=text)
                cell.font = bold_font; cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            apply_borders_to_range(sheet, current_row, 1, current_row, 6)

            for text, col in col_hdrs:
                cell = sheet.cell(row=current_row, column=col + 6, value=text)
                cell.font = bold_font; cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            apply_borders_to_range(sheet, current_row, 7, current_row, 12)
            sheet.row_dimensions[current_row].height = 30
            current_row += 1

            section_start = current_row
            tipos = sorted(set(n_items.keys()) | set(d_items.keys()), key=natural_sort_key)

            for tipo in tipos:
                row = current_row
                # ADD
                if tipo in n_items:
                    item = n_items[tipo]
                    metr = round(item.get('BaseQuantity', 0), 2)
                    unit = price_data.get(tipo, {}).get('Un', '')
                    vunit = round(price_data.get(tipo, {}).get('Valor', 0), 2)
                    desc = item.get('Descricao', '')
                    cA = sheet.cell(row=row, column=1, value=tipo)
                    cA.font = regular_font; cA.alignment = Alignment(horizontal='center', vertical='center')
                    cB = sheet.cell(row=row, column=2, value=desc)
                    cB.font = regular_font; cB.alignment = Alignment(vertical='center', wrap_text=True)
                    cC = sheet.cell(row=row, column=3, value=metr)
                    cC.number_format = currency_format
                    cC.font = regular_font; cC.alignment = Alignment(vertical='center')
                    cD = sheet.cell(row=row, column=4, value=unit)
                    cD.font = regular_font; cD.alignment = Alignment(horizontal='center', vertical='center')
                    cE = sheet.cell(row=row, column=5, value=vunit)
                    cE.number_format = currency_format
                    cE.font = regular_font; cE.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cF = sheet.cell(row=row, column=6, value=f'=E{row}*C{row}')
                    cF.number_format = accounting_format
                    cF.font = regular_font; cF.fill = fill_add_label
                    cF.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                apply_borders_to_range(sheet, row, 1, row, 6)

                # DIST
                if tipo in d_items:
                    item = d_items[tipo]
                    metr = round(item.get('BaseQuantity', 0), 2)
                    unit = price_data.get(tipo, {}).get('Un', '')
                    vunit = round(price_data.get(tipo, {}).get('Valor', 0), 2)
                    desc = item.get('Descricao', '')
                    cG = sheet.cell(row=row, column=7, value=tipo)
                    cG.font = regular_font; cG.alignment = Alignment(horizontal='center', vertical='center')
                    cH = sheet.cell(row=row, column=8, value=desc)
                    cH.font = regular_font; cH.alignment = Alignment(vertical='center', wrap_text=True)
                    cI = sheet.cell(row=row, column=9, value=metr)
                    cI.number_format = currency_format
                    cI.font = regular_font; cI.alignment = Alignment(vertical='center')
                    cJ = sheet.cell(row=row, column=10, value=unit)
                    cJ.font = regular_font; cJ.alignment = Alignment(horizontal='center', vertical='center')
                    cK = sheet.cell(row=row, column=11, value=vunit)
                    cK.number_format = currency_format
                    cK.font = regular_font; cK.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cL = sheet.cell(row=row, column=12, value=f'=I{row}*K{row}')
                    cL.number_format = accounting_format
                    cL.font = regular_font; cL.fill = fill_dis_label
                    cL.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                apply_borders_to_range(sheet, row, 7, row, 12)

                # Coluna M: diferença de metragem (fonte vermelha, sem bordas)
                cM = sheet.cell(row=row, column=13, value=f'=C{row}-I{row}')
                cM.font = font_red; cM.alignment = Alignment(vertical='center', wrap_text=True)

                current_row += 1

                # === SUBITENS (ISOLAMENTO ACÚSTICO) ===
                n_insulation = n_items.get(tipo, {}).get('insulation_items', {}) if tipo in n_items else {}
                d_insulation = d_items.get(tipo, {}).get('insulation_items', {}) if tipo in d_items else {}

                # Coletar todas as chaves de isolamento (ADD + DIST)
                all_insul_keys = set(n_insulation.keys()) | set(d_insulation.keys())

                for insul_key in sorted(all_insul_keys):
                    sub_row = current_row

                    # ADD - Isolamento
                    if insul_key in n_insulation:
                        sub_item = n_insulation[insul_key]
                        sub_metr = round(sub_item.get('Quantidade', 0), 2)
                        sub_unit = price_data.get(sub_item['Tipo Code'], {}).get('Un', '')
                        sub_vunit = round(price_data.get(sub_item['Tipo Code'], {}).get('Valor', 0), 2)
                        sub_desc = sub_item.get('Descricao', '')

                        # Deixar vazio coluna A (tipo) para subitens
                        cB_sub = sheet.cell(row=sub_row, column=2, value=sub_desc)
                        cB_sub.font = regular_font; cB_sub.alignment = Alignment(vertical='center', wrap_text=True)
                        cC_sub = sheet.cell(row=sub_row, column=3, value=sub_metr)
                        cC_sub.number_format = currency_format
                        cC_sub.font = regular_font; cC_sub.alignment = Alignment(vertical='center')
                        cD_sub = sheet.cell(row=sub_row, column=4, value=sub_unit)
                        cD_sub.font = regular_font; cD_sub.alignment = Alignment(horizontal='center', vertical='center')
                        cE_sub = sheet.cell(row=sub_row, column=5, value=sub_vunit)
                        cE_sub.number_format = currency_format
                        cE_sub.font = regular_font; cE_sub.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cF_sub = sheet.cell(row=sub_row, column=6, value=f'=E{sub_row}*C{sub_row}')
                        cF_sub.number_format = accounting_format
                        cF_sub.font = regular_font; cF_sub.fill = fill_add_label
                        cF_sub.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    apply_borders_to_range(sheet, sub_row, 1, sub_row, 6)

                    # DIST - Isolamento
                    if insul_key in d_insulation:
                        sub_item = d_insulation[insul_key]
                        sub_metr = round(sub_item.get('Quantidade', 0), 2)
                        sub_unit = price_data.get(sub_item['Tipo Code'], {}).get('Un', '')
                        sub_vunit = round(price_data.get(sub_item['Tipo Code'], {}).get('Valor', 0), 2)
                        sub_desc = sub_item.get('Descricao', '')

                        # Deixar vazio coluna G (tipo) para subitens
                        cH_sub = sheet.cell(row=sub_row, column=8, value=sub_desc)
                        cH_sub.font = regular_font; cH_sub.alignment = Alignment(vertical='center', wrap_text=True)
                        cI_sub = sheet.cell(row=sub_row, column=9, value=sub_metr)
                        cI_sub.number_format = currency_format
                        cI_sub.font = regular_font; cI_sub.alignment = Alignment(vertical='center')
                        cJ_sub = sheet.cell(row=sub_row, column=10, value=sub_unit)
                        cJ_sub.font = regular_font; cJ_sub.alignment = Alignment(horizontal='center', vertical='center')
                        cK_sub = sheet.cell(row=sub_row, column=11, value=sub_vunit)
                        cK_sub.number_format = currency_format
                        cK_sub.font = regular_font; cK_sub.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cL_sub = sheet.cell(row=sub_row, column=12, value=f'=I{sub_row}*K{sub_row}')
                        cL_sub.number_format = accounting_format
                        cL_sub.font = regular_font; cL_sub.fill = fill_dis_label
                        cL_sub.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    apply_borders_to_range(sheet, sub_row, 7, sub_row, 12)

                    # Coluna M: diferença de metragem para subitens
                    cM_sub = sheet.cell(row=sub_row, column=13, value=f'=C{sub_row}-I{sub_row}')
                    cM_sub.font = font_red; cM_sub.alignment = Alignment(vertical='center', wrap_text=True)

                    current_row += 1

            section_end = current_row - 1

            # Total da categoria
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            tl = sheet.cell(row=current_row, column=1, value=f'VALOR TOTAL DOS {cat_label.upper()}')
            tl.font = bold_font; tl.fill = fill_add_total_l
            tl.alignment = Alignment(horizontal='center', vertical='center')
            for c in range(1, 6): sheet.cell(row=current_row, column=c).fill = fill_add_total_l
            tf = sheet.cell(row=current_row, column=6, value=f'=SUM(F{section_start}:F{section_end})')
            tf.number_format = accounting_format
            tf.font = bold_font; tf.fill = fill_add_total_v
            tf.alignment = Alignment(horizontal='right', vertical='center')
            apply_borders_to_range(sheet, current_row, 1, current_row, 6)

            sheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=11)
            tr = sheet.cell(row=current_row, column=7, value=f'VALOR TOTAL DOS {cat_label.upper()} (DISTRATO)')
            tr.font = bold_font; tr.fill = fill_dis_total_l
            tr.alignment = Alignment(horizontal='center', vertical='center')
            for c in range(7, 12): sheet.cell(row=current_row, column=c).fill = fill_dis_total_l
            tm = sheet.cell(row=current_row, column=12, value=f'=SUM(L{section_start}:L{section_end})')
            tm.number_format = accounting_format
            tm.font = bold_font; tm.fill = fill_dis_total_v
            tm.alignment = Alignment(horizontal='right', vertical='center')
            apply_borders_to_range(sheet, current_row, 7, current_row, 12)
            sheet.row_dimensions[current_row].height = 14.25

            cat_total_f.append(f'F{current_row}')
            cat_total_l.append(f'L{current_row}')
            current_row += 1

        # Espaçador antes do sub-total
        _spacer(current_row); current_row += 1

        # Sub-Total (soma de todas as categorias do bloco)
        if len(cat_total_f) > 1:
            sub_f = f'=SUM({"+".join(cat_total_f)})'
            sub_l = f'=SUM({"+".join(cat_total_l)})'
        else:
            sub_f = f'={cat_total_f[0]}' if cat_total_f else '0'
            sub_l = f'={cat_total_l[0]}' if cat_total_l else '0'

        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        sl = sheet.cell(row=current_row, column=1, value='SUB-TOTAL')
        sl.font = bold_font; sl.fill = fill_add_total_l
        sl.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(1, 6): sheet.cell(row=current_row, column=c).fill = fill_add_total_l
        sf = sheet.cell(row=current_row, column=6, value=sub_f)
        sf.number_format = accounting_format
        sf.font = bold_font; sf.fill = fill_add_total_v
        sf.alignment = Alignment(horizontal='right', vertical='center')
        apply_borders_to_range(sheet, current_row, 1, current_row, 6)

        sheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=11)
        sr = sheet.cell(row=current_row, column=7, value='SUB-TOTAL')
        sr.font = bold_font; sr.fill = fill_dis_total_l
        sr.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(7, 12): sheet.cell(row=current_row, column=c).fill = fill_dis_total_l
        sm = sheet.cell(row=current_row, column=12, value=sub_l)
        sm.number_format = accounting_format
        sm.font = bold_font; sm.fill = fill_dis_total_v
        sm.alignment = Alignment(horizontal='right', vertical='center')
        apply_borders_to_range(sheet, current_row, 7, current_row, 12)
        sheet.row_dimensions[current_row].height = 14.25
        sub_f_coord = f'F{current_row}'
        sub_l_coord = f'L{current_row}'
        current_row += 1

        # Valor Total da Proposta (azul claro, A:K mesclado, valor em L)
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        pl = sheet.cell(row=current_row, column=1, value='VALOR TOTAL DA PROPOSTA: (ADITIVOS - DISTRATOS)')
        pl.font = bold_font; pl.fill = fill_titulo
        pl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for c in range(1, 12): sheet.cell(row=current_row, column=c).fill = fill_titulo
        pm = sheet.cell(row=current_row, column=12, value=f'={sub_f_coord}-{sub_l_coord}')
        pm.number_format = accounting_format
        pm.font = bold_font; pm.fill = fill_titulo
        pm.alignment = Alignment(horizontal='right', vertical='center')
        apply_borders_to_range(sheet, current_row, 1, current_row, 12)
        sheet.row_dimensions[current_row].height = 21.0
        all_proposta_l.append(f'L{current_row}')
        current_row += 1

        # Espaçador entre blocos
        _spacer(current_row); current_row += 1

    # Valor Total da Obra (azul claro, A:K mesclado, valor em L)
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
    ol = sheet.cell(row=current_row, column=1,
        value='VALOR TOTAL DA OBRA:\nVALOR A SER COBRADO COMO ESCOPO ADICIONAL RESULTANTE DA DIFERENÇA ENTRE NOVO LAYOUT E SERVIÇOS DISTRATADOS')
    ol.font = bold_font; ol.fill = fill_titulo
    ol.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for c in range(1, 12): sheet.cell(row=current_row, column=c).fill = fill_titulo
    om = sheet.cell(row=current_row, column=12, value=f'=SUM({",".join(all_proposta_l)})' if all_proposta_l else '0')
    om.number_format = accounting_format
    om.font = bold_font; om.fill = fill_titulo
    om.alignment = Alignment(horizontal='right', vertical='center')
    apply_borders_to_range(sheet, current_row, 1, current_row, 12)
    sheet.row_dimensions[current_row].height = 24.75
    final_row = current_row

    sheet.page_setup.paperSize = 9
    sheet.page_setup.orientation = 'portrait'
    sheet.page_setup.scale = 46
    sheet.page_setup.printArea = f'$A$1:$M${final_row}'
    sheet.sheet_view.view = 'pageBreakPreview'



def write_client_sheet(sheet, client_normal, client_distratado, price_data, bold_font, header_fill, currency_format, accounting_format, regular_font, bold_white_font=None, regular_white_font=None):
    sheet.column_dimensions['A'].width = 12.29
    sheet.column_dimensions['B'].width = 43.37
    sheet.column_dimensions['C'].width = 10.14
    sheet.column_dimensions['D'].width = 5.00
    sheet.column_dimensions['E'].width = 13.91
    sheet.column_dimensions['F'].width = 13.20

    current_row = 2
    normal_client_total_coords = {} # Dicionário para armazenar os totais normais por cliente

    # --- Bloco 1: Serviços Normais ---
    header_cell = sheet.cell(row=current_row, column=1, value="BLOCO 1: SERVIÇOS ADICIONAIS (NOVO LAYOUT)")
    header_cell.font = bold_font
    header_cell.fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_borders_to_range(sheet, current_row, 1, current_row, 6)
    current_row += 1
    obra_total_cells = []

    sorted_clients_normal = sorted(client_normal.keys(), key=sort_client_id_key)
    for client_id in sorted_clients_normal:
        client_header_cell = sheet.cell(row=current_row, column=1, value=client_id)
        client_header_cell.font = bold_font
        client_header_cell.fill = PatternFill(start_color="e46c0a", end_color="e46c0a", fill_type="solid")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        client_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        apply_borders_to_range(sheet, current_row, 1, current_row, 6)
        current_row += 1

        current_row, subtotal_cells, _ = _write_client_section(sheet, client_normal[client_id], price_data, current_row, bold_font, header_fill, currency_format, accounting_format, regular_font, bold_white_font, regular_white_font)
        total_formula = f"=ROUND({'+'.join(subtotal_cells)}, 2)" if subtotal_cells else "0"

        total_row = current_row
        total_label_cell = sheet.cell(row=total_row, column=1, value=f'TOTAL - {client_id} (SERVIÇOS ADICIONAIS)')
        total_label_cell.font = bold_font

        total_fill = PatternFill(start_color="e46c0a", end_color="e46c0a", fill_type="solid")
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, 5):
            sheet.cell(row=total_row, column=col_idx).fill = total_fill

        sheet.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=6)
        total_value_cell = sheet.cell(row=total_row, column=5, value=total_formula)
        total_value_cell.number_format = accounting_format
        total_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_value_cell.fill = total_fill
        total_value_cell.font = regular_font
        apply_borders_to_range(sheet, total_row, 1, total_row, 6)

        total_coord = total_value_cell.coordinate
        obra_total_cells.append(total_coord)
        normal_client_total_coords[client_id] = total_coord

        current_row += 1
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1

    total_obra_row = current_row
    total_obra_label_cell = sheet.cell(row=total_obra_row, column=1, value='VALOR TOTAL DA OBRA - SERVIÇOS ADICIONAIS')
    total_obra_label_cell.font = bold_font
    total_obra_fill = PatternFill(start_color="77933c", end_color="77933c", fill_type="solid")
    sheet.merge_cells(start_row=total_obra_row, start_column=1, end_row=total_obra_row, end_column=4)
    total_obra_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    for col_idx in range(1, 5):
        sheet.cell(row=total_obra_row, column=col_idx).fill = total_obra_fill

    sheet.merge_cells(start_row=total_obra_row, start_column=5, end_row=total_obra_row, end_column=6)
    total_obra_value_cell = sheet.cell(row=total_obra_row, column=5, value=f"=ROUND({'+'.join(obra_total_cells)}, 2)")
    total_obra_value_cell.number_format = currency_format
    total_obra_value_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_obra_value_cell.fill = total_obra_fill
    total_obra_value_cell.font = bold_font
    apply_borders_to_range(sheet, total_obra_row, 1, total_obra_row, 6)
    total_obra_coord = total_obra_value_cell.coordinate

    current_row += 1
    format_empty_row(sheet, current_row, regular_font)
    current_row += 1

    # --- Bloco 2: Serviços Distratados ---
    header_cell_distratado = sheet.cell(row=current_row, column=1, value="BLOCO 2: SERVIÇOS DISTRATADOS")
    header_cell_distratado.font = bold_white_font
    header_cell_distratado.fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    header_cell_distratado.alignment = Alignment(horizontal='center', vertical='center')
    apply_borders_to_range(sheet, current_row, 1, current_row, 6)
    current_row += 1
    obra_total_distratado_cells = []

    sorted_clients_distratado = sorted(client_distratado.keys(), key=sort_client_id_key)
    for client_id in sorted_clients_distratado:
        client_header_cell = sheet.cell(row=current_row, column=1, value=client_id)
        client_header_cell.font = bold_font
        client_header_cell.fill = PatternFill(start_color="e46c0a", end_color="e46c0a", fill_type="solid")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        client_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        apply_borders_to_range(sheet, current_row, 1, current_row, 6)
        current_row += 1

        current_row, subtotal_cells, _ = _write_client_section(sheet, client_distratado[client_id], price_data, current_row, bold_font, header_fill, currency_format, accounting_format, regular_font, bold_white_font, regular_white_font)
        total_formula = f"=ROUND({'+'.join(subtotal_cells)}, 2)" if subtotal_cells else "0"

        total_distratado_row = current_row
        total_distratado_label_cell = sheet.cell(row=total_distratado_row, column=1, value=f'TOTAL -  {client_id} (SERVIÇOS DISTRATADOS)')
        total_distratado_label_cell.font = bold_font
        total_distratado_fill = PatternFill(start_color="93cddd", end_color="93cddd", fill_type="solid")

        sheet.merge_cells(start_row=total_distratado_row, start_column=1, end_row=total_distratado_row, end_column=4)
        total_distratado_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, 5):
            sheet.cell(row=total_distratado_row, column=col_idx).fill = total_distratado_fill

        sheet.merge_cells(start_row=total_distratado_row, start_column=5, end_row=total_distratado_row, end_column=6)
        total_distratado_value_cell = sheet.cell(row=total_distratado_row, column=5, value=total_formula)
        total_distratado_value_cell.number_format = accounting_format
        total_distratado_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_distratado_value_cell.fill = total_distratado_fill
        total_distratado_value_cell.font = regular_font
        apply_borders_to_range(sheet, total_distratado_row, 1, total_distratado_row, 6)
        distratado_coord = total_distratado_value_cell.coordinate
        obra_total_distratado_cells.append(distratado_coord)
        current_row += 1

        format_empty_row(sheet, current_row, regular_font)
        current_row += 1

        # Adiciona a linha de SALDO para o cliente
        saldo_row = current_row
        normal_coord = normal_client_total_coords.get(client_id, "0")

        saldo_label_cell = sheet.cell(row=saldo_row, column=1, value=f"TOTAL APTO {client_id} - VALOR CONSIDERANDO O DESCONTO DOS SERVIÇOS DISTRATADOS EM RELAÇÃO AOS SERVIÇOS EXECUTADOS POR NOVO LAYOUT")
        saldo_label_cell.font = bold_white_font

        saldo_fill = PatternFill(start_color="31859c", end_color="31859c", fill_type="solid")

        sheet.merge_cells(start_row=saldo_row, start_column=1, end_row=saldo_row, end_column=4)
        saldo_label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col_idx in range(1, 5):
            sheet.cell(row=saldo_row, column=col_idx).fill = saldo_fill

        sheet.merge_cells(start_row=saldo_row, start_column=5, end_row=saldo_row, end_column=6)
        saldo_value_cell = sheet.cell(row=saldo_row, column=5, value=f"=ROUND({normal_coord}-{distratado_coord}, 2)")
        saldo_value_cell.number_format = currency_format
        saldo_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        saldo_value_cell.fill = saldo_fill
        saldo_value_cell.font = bold_white_font
        apply_borders_to_range(sheet, saldo_row, 1, saldo_row, 6)
        current_row += 1

        format_empty_row(sheet, current_row, regular_font)
        current_row += 1

    saldo_final_row = current_row
    saldo_final_label_cell = sheet.cell(row=saldo_final_row, column=1, value='VALOR TOTAL DA OBRA:\n' \
        'VALOR À SER COBRADO COMO ESCOPO ADICIONAL RESULTANTE DA DIFERENÇA \n' \
        'ENTRE NOVO LAYOUT E SERVIÇOS DISTRATADOS')
    saldo_final_label_cell.font = bold_white_font

    saldo_final_fill = PatternFill(start_color="953735", end_color="953735", fill_type="solid")

    sheet.merge_cells(start_row=saldo_final_row, start_column=1, end_row=saldo_final_row, end_column=4)
    saldo_final_label_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx in range(1, 5):
        sheet.cell(row=saldo_final_row, column=col_idx).fill = saldo_final_fill

    sheet.merge_cells(start_row=saldo_final_row, start_column=5, end_row=saldo_final_row, end_column=6)

    distratado_total_formula = f"({'+'.join(obra_total_distratado_cells)})" if obra_total_distratado_cells else "0"
    saldo_final_value_cell = sheet.cell(row=saldo_final_row, column=5, value=f"=ROUND({total_obra_coord}-{distratado_total_formula}, 2)")
    saldo_final_value_cell.number_format = currency_format
    saldo_final_value_cell.alignment = Alignment(horizontal='right', vertical='center')
    saldo_final_value_cell.fill = saldo_final_fill
    saldo_final_value_cell.font = bold_white_font
    apply_borders_to_range(sheet, saldo_final_row, 1, saldo_final_row, 6)

    sheet.row_dimensions[1].hidden = True
    sheet.page_setup.paperSize = 9
    sheet.page_setup.orientation = 'portrait'
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.printArea = f'A2:F{saldo_final_row}'
    sheet.sheet_view.view = 'pageBreakPreview'


def _write_client_section(sheet, items_by_key, price_data, start_row, bold_font, header_fill, currency_format, accounting_format, regular_font, bold_white_font=None, regular_white_font=None):
    COLS = {'TIPO': 1, 'DESC': 2, 'METRAGEM': 3, 'UN': 4, 'VALOR_UNIT': 5, 'VALOR_TOTAL': 6}
    current_row = start_row
    client_subtotal_cells, item_key_to_cell_map = [], {}

    cats = {'Forro': [], 'Paredes_e_Revestimentos': [], 'Guias e Montantes': []}
    for item_key, item_data in items_by_key.items():
        category = 'Paredes_e_Revestimentos' if item_data.get('Categoria') in ['Parede', 'Revestimento'] else item_data.get('Categoria')
        if category in cats: cats[category].append((item_key, item_data))

    for cat_name, cat_label in [('Forro', 'Forros'), ('Paredes_e_Revestimentos', 'Paredes e Revestimentos'), ('Guias e Montantes', 'Guias e Montantes')]:
        if not cats.get(cat_name): continue
        header_row_data = [('Tipo R. Bassani', COLS['TIPO']), (cat_label, COLS['DESC']), ('Metragem', COLS['METRAGEM']), ('Un', COLS['UN']), ('Valor do Material + MO', COLS['VALOR_UNIT']), ('Valor Total', COLS['VALOR_TOTAL'])]
        for val, col_idx in header_row_data:
            cell = sheet.cell(row=current_row, column=col_idx, value=val)
            cell.font = bold_font
            # Forçar cor #99CCFF nos cabeçalhos
            cell.fill = PatternFill(start_color="99ccff", end_color="99ccff", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['VALOR_TOTAL']) # Add borders
        current_row += 1
        section_start_row, ft16_placeholder_cell, ft16_base_quantity, subtraction_cell_list = current_row, None, 0, []

        if cat_name == 'Paredes_e_Revestimentos':
            sorted_items = sorted(cats[cat_name], key=lambda x: (0 if x[1]['Tipo Code'].startswith('TP') else 1, natural_sort_key(x[1]['Tipo Code'])))
        else: # Forros, Guias e Montantes
            sorted_items = sorted(cats[cat_name], key=lambda x: natural_sort_key(x[1]['Tipo Code']))
        
        for item_key, item in sorted_items:
            price_info = price_data.get(item['Tipo Code'], {'Valor': 0, 'Un': ''})
            main_item_row, is_subclass_item = current_row, item.get('is_subclass', False)

            # Calculate number of sub-items for merging
            num_sub_items = len(item.get('insulation_items', {})) + len(item.get('carenagem_items', {}))

            tipo_cell = sheet.cell(row=main_item_row, column=COLS['TIPO'], value=item['Tipo Code'])
            tipo_cell.alignment = Alignment(horizontal='center', vertical='center')
            tipo_cell.font = regular_font

            # Merge cell if there are sub-items
            if num_sub_items > 0:
                sheet.merge_cells(
                    start_row=main_item_row,
                    start_column=COLS['TIPO'],
                    end_row=main_item_row + num_sub_items,
                    end_column=COLS['TIPO']
                )

            desc_cell = sheet.cell(row=main_item_row, column=COLS['DESC'], value=item['Descricao'])
            desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
            desc_cell.font = regular_font
            
            m_cell_ref = f"{get_column_letter(COLS['METRAGEM'])}{main_item_row}"
            item_key_to_cell_map[item_key] = m_cell_ref
            
            metragem_cell = sheet.cell(row=main_item_row, column=COLS['METRAGEM'])
            metragem_cell.number_format = currency_format
            metragem_cell.font = regular_font
            metragem_cell.alignment = Alignment(vertical='center')
            
            # GARANTIR que a BaseQuantity seja sempre escrita como valor numérico inicial
            base_qty = item.get('BaseQuantity', 0)
            try:
                base_qty_num = float(str(base_qty).strip().replace(',', '.'))
            except Exception:
                base_qty_num = 0.0
            metragem_cell.value = round(base_qty_num, 2)

            if 'formula_contributors' in item:
                parts = [f"({item_key_to_cell_map[c['item_key']]}*{c['count']})" if c['count'] > 1 else item_key_to_cell_map[c['item_key']] for c in item['formula_contributors'] if c['item_key'] in item_key_to_cell_map]
                if parts:
                    final_formula = f"=CEILING(({'+'.join(parts)})*{item['formula_multiplier']},1)"
                else:
                    # Caso não haja partes da fórmula, manter a metragem numérica (evita strings como '1.4')
                    final_formula = round(base_qty_num, 2)
                metragem_cell.value = final_formula
            
            un_cell = sheet.cell(row=main_item_row, column=COLS['UN'], value=price_info['Un'])
            un_cell.font = regular_font
            un_cell.alignment = Alignment(horizontal='center', vertical='center')
            valor_unit_cell = sheet.cell(row=main_item_row, column=COLS['VALOR_UNIT'], value=round(price_info['Valor'], 2))
            valor_unit_cell.number_format = accounting_format
            valor_unit_cell.font = regular_font
            valor_unit_cell.alignment = Alignment(vertical='center')
            v_unit_cell = f"{get_column_letter(COLS['VALOR_UNIT'])}{main_item_row}"
            valor_total_cell = sheet.cell(row=main_item_row, column=COLS['VALOR_TOTAL'], value=f"=ROUND({m_cell_ref}*{v_unit_cell}, 2)")
            valor_total_cell.number_format = accounting_format
            valor_total_cell.font = regular_font
            valor_total_cell.alignment = Alignment(vertical='center')
            if is_subclass_item: subtraction_cell_list.append(m_cell_ref)
            if item['Tipo Code'] == 'FT16': ft16_placeholder_cell, ft16_base_quantity = m_cell_ref, item.get('FormulaBase', item['BaseQuantity'])
            apply_borders_to_range(sheet, main_item_row, COLS['TIPO'], main_item_row, COLS['VALOR_TOTAL']) # Add borders
            current_row += 1

            # --- SUB-ITEM LOOPS ---
            insulation_cell_references = [] # Initialized here for each main item
            
            # ISOLAMENTO
            for sub_item_key, sub_item in sorted(item.get('insulation_items', {}).items()):
                price_info_sub = price_data.get(sub_item['Tipo Code'], {'Valor': 0, 'Un': ''})
                sub_desc_cell = sheet.cell(row=current_row, column=COLS['DESC'], value=sub_item['Descricao'])
                sub_desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
                sub_desc_cell.font = regular_font
                
                sub_m_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'])
                sub_m_cell.number_format = currency_format
                sub_m_cell.font = regular_font
                sub_m_cell.alignment = Alignment(vertical='center')

                if sub_item.get('is_la_dupla'):
                    sub_m_cell.value = f"={m_cell_ref}*2"
                else:
                    sub_m_cell.value = sub_item['Quantidade']
                    insulation_cell_references.append(sub_m_cell.coordinate) # Only append if not 'lã dupla'
                
                # Add lines to write 'Un' and 'Valor do Material + MO'
                sub_un_cell = sheet.cell(row=current_row, column=COLS['UN'], value=price_info_sub['Un'])
                sub_un_cell.font = regular_font
                sub_un_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                sub_v_unit_cell_obj = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=price_info_sub['Valor'])
                sub_v_unit_cell_obj.number_format = accounting_format
                sub_v_unit_cell_obj.font = regular_font
                sub_v_unit_cell_obj.alignment = Alignment(horizontal='right', vertical='center')
                
                sub_v_unit_cell = f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}"
                sub_v_total_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"={sub_m_cell.coordinate}*{sub_v_unit_cell}")
                sub_v_total_cell.number_format = accounting_format
                sub_v_total_cell.font = regular_font
                sub_v_total_cell.alignment = Alignment(horizontal='right', vertical='center')
                apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['VALOR_TOTAL']) # Add borders
                current_row += 1

            # CARNAGEM
            for sub_item_key, sub_item in sorted(item.get('carenagem_items', {}).items()):
                price_info_sub = price_data.get(sub_item['Tipo Code'], {'Valor': 0, 'Un': ''})
                sub_desc_cell = sheet.cell(row=current_row, column=COLS['DESC'], value=sub_item['Descricao'])
                sub_desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
                sub_desc_cell.font = regular_font
                sub_m_cell = sheet.cell(row=current_row, column=COLS['METRAGEM'], value=sub_item['Quantidade'])
                sub_m_cell.number_format = currency_format
                sub_m_cell.font = regular_font
                sub_m_cell.alignment = Alignment(vertical='center')
                sub_un_cell = sheet.cell(row=current_row, column=COLS['UN'], value=price_info_sub['Un'])
                sub_un_cell.font = regular_font
                sub_un_cell.alignment = Alignment(horizontal='center', vertical='center')
                sub_v_unit_cell_obj = sheet.cell(row=current_row, column=COLS['VALOR_UNIT'], value=price_info_sub['Valor'])
                sub_v_unit_cell_obj.number_format = accounting_format
                sub_v_unit_cell_obj.font = regular_font
                sub_v_unit_cell_obj.alignment = Alignment(horizontal='right', vertical='center')
                sub_v_unit_cell = f"{get_column_letter(COLS['VALOR_UNIT'])}{current_row}"
                sub_v_total_cell = sheet.cell(row=current_row, column=COLS['VALOR_TOTAL'], value=f"={sub_m_cell.coordinate}*{sub_v_unit_cell}")
                sub_v_total_cell.number_format = accounting_format
                sub_v_total_cell.font = regular_font
                sub_v_total_cell.alignment = Alignment(horizontal='right', vertical='center')
                apply_borders_to_range(sheet, current_row, COLS['TIPO'], current_row, COLS['VALOR_TOTAL']) # Add borders
                current_row += 1

            # --- FORMULA LOGIC FOR MAIN ITEM (AFTER SUB-ITEMS ARE PROCESSED) ---
            if cat_name != 'Guias e Montantes' and not is_subclass_item and item['Tipo Code'] != 'FT16':
                # Usar o valor numérico já calculado anteriormente para evitar strings como '1.4'
                final_metragem_value = round(base_qty_num, 2)

                if item.get('logic_applied'): # Merged items (BASE + LÃ)
                    formula_parts = []
                    # Garantir que base_val seja numérico
                    try:
                        base_val = float(str(item.get('FormulaBase', item['BaseQuantity'])).strip().replace(',', '.'))
                    except Exception:
                        base_val = 0.0
                    if base_val > 0:
                        formula_parts.append(str(base_val))
                    if insulation_cell_references: # Should have one reference
                        formula_parts.extend(insulation_cell_references)
                    
                    if len(formula_parts) > 1:
                        final_metragem_value = f"={'+'.join(formula_parts)}"
                    elif len(formula_parts) == 1:
                        # Check if it's a cell reference (starts with a letter)
                        if re.match(r'^[A-Z]', formula_parts[0]):
                            final_metragem_value = f"={formula_parts[0]}"
                        else:
                            final_metragem_value = float(formula_parts[0]) if formula_parts[0].replace('.','',1).isdigit() else formula_parts[0]
                elif item.get('insulation_items'): # Non-merged item with insulation (just LÃ)
                    if insulation_cell_references: # Should have one reference
                        final_metragem_value = f"={insulation_cell_references[0]}" # Only the insulation cell ref
                    else:
                        final_metragem_value = round(base_qty_num, 2) # Fallback if no insulation ref found
                else: # Item without insulation, not merged (just BASE)
                    final_metragem_value = round(base_qty_num, 2) # Use the numeric value directly (no '=')

                metragem_cell.value = final_metragem_value
        
        # --- FT16 FORMULA (AFTER ALL ITEMS IN CATEGORY ARE PROCESSED) ---
        if ft16_placeholder_cell and subtraction_cell_list:
            base_val_str = str(ft16_base_quantity).replace(',', '.')
            sheet[ft16_placeholder_cell] = f"={base_val_str}-{'-'.join(subtraction_cell_list)}"
            sheet[ft16_placeholder_cell].number_format = currency_format
            sheet[ft16_placeholder_cell].font = regular_font

        # --- SUBTOTALS ---
        subtotal_row_idx = current_row
        subtotal_cell = sheet.cell(row=subtotal_row_idx, column=COLS['TIPO'], value='SUB-TOTAL')
        subtotal_cell.font = bold_font

        # Apply formatting for A:D
        subtotal_fill = PatternFill(start_color="ffff99", end_color="ffff99", fill_type="solid")
        subtotal_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=subtotal_row_idx, start_column=1, end_row=subtotal_row_idx, end_column=4)

        for col_idx in range(1, 5):
            sheet.cell(row=subtotal_row_idx, column=col_idx).fill = subtotal_fill

        # Merge E:F and apply formula
        sheet.merge_cells(start_row=subtotal_row_idx, start_column=COLS['VALOR_UNIT'], end_row=subtotal_row_idx, end_column=COLS['VALOR_TOTAL'])
        
        total_col_letter = get_column_letter(COLS['VALOR_TOTAL'])
        formula = f"=SUM({total_col_letter}{section_start_row}:{total_col_letter}{current_row - 1})"
        
        merged_value_cell = sheet.cell(row=subtotal_row_idx, column=COLS['VALOR_UNIT'], value=formula)
        merged_value_cell.number_format = currency_format
        merged_value_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_value_cell.fill = subtotal_fill
        merged_value_cell.font = bold_font
        
        apply_borders_to_range(sheet, subtotal_row_idx, COLS['TIPO'], subtotal_row_idx, COLS['VALOR_TOTAL']) # Add borders
        client_subtotal_cells.append(merged_value_cell.coordinate)
        current_row += 1
        format_empty_row(sheet, current_row, regular_font)
        current_row += 1
    
    return current_row, client_subtotal_cells, item_key_to_cell_map

# ==============================================================================
# AJUSTES E REGRAS DE NEGÓCIO
# ==============================================================================

def apply_wool_logic(data_by_client):
    for client_id, client_items in data_by_client.items():
        grouped_keys = {}
        for key in list(client_items.keys()):
            tipo_code, categoria, has_insulation = key
            base_key = (tipo_code, categoria)
            if base_key not in grouped_keys: grouped_keys[base_key] = []
            grouped_keys[base_key].append(key)

        for keys in grouped_keys.values():
            if len(keys) == 2:
                key_with_wool = next((k for k in keys if k[2]), None)
                key_without_wool = next((k for k in keys if not k[2]), None)
                
                if key_with_wool and key_without_wool:
                    item_com_la, item_sem_la = client_items[key_with_wool], client_items[key_without_wool]
                    original_qty_com_la, qty_sem_la = item_com_la['BaseQuantity'], item_sem_la['BaseQuantity']
                    item_com_la['FormulaBase'] = qty_sem_la
                    item_com_la['BaseQuantity'] = qty_sem_la + original_qty_com_la
                    if item_com_la['insulation_items']:
                        insul_key = next(iter(item_com_la['insulation_items']))
                        if not item_com_la['insulation_items'][insul_key].get('is_la_dupla', False):
                            item_com_la['insulation_items'][insul_key]['Quantidade'] = original_qty_com_la
                    item_com_la['logic_applied'] = True
                    item_com_la['carenagem_items'].update(item_sem_la.get('carenagem_items', {}))
                    del client_items[key_without_wool]
    return data_by_client

def calculate_and_add_derived_items(data_by_client, price_data):
    known_profiles = ['MS48', 'MS70', 'MS90', 'F47', 'MS140']
    rules = { 'SHF-40': {'m': 0.35, 'f': ['Ru', '/400'], 'c': True}, 'SH48-40': {'m': 0.45, 'f': ['MS48', '/400'], 'c': True}, 'SH48-60': {'m': 0.45, 'f': ['MS48', '/600'], 'c': True},
              'SH70-40': {'m': 0.45, 'f': ['MS70', '/400'], 'c': True}, 'SH70-60': {'m': 0.45, 'f': ['MS70', '/600'], 'c': True}, 'SH90-40': {'m': 0.45, 'f': ['MS90', '/400'], 'c': True},
              'SH90-60': {'m': 0.45, 'f': ['MS90', '/600'], 'c': True}, 'SH140-40': {'m': 0.45, 'f': ['MS140', '/400'], 'c': True}, 'SH140-60': {'m': 0.45, 'f': ['MS140', '/600'], 'c': True} }

    for client_id, client_items in data_by_client.items():
        contributions = {code: [] for code in rules}
        for item_key, item in client_items.items():
            if item.get('Categoria') in ['Parede', 'Revestimento'] and item.get('Tipo Code', '').startswith('TP') and item.get('Descricao') and item.get('BaseQuantity', 0) > 0:
                for code, rule in rules.items():
                    if all(f in item['Descricao'] for f in rule['f']):
                        count = item['Descricao'].count('Ru') if code == 'SHF-40' else (item['Descricao'].count(next((p for p in known_profiles if re.search(r'\d+', code).group(0) in p), None)) if rule['c'] and next((p for p in known_profiles if re.search(r'\d+', code).group(0) in p), None) else 1)
                        count = 1 if count == 0 else count
                        if count > 0: contributions[code].append({'item_key': item_key, 'count': count})
        
        for code, contribs in contributions.items():
            if contribs:
                total_contrib_val = sum(client_items[c['item_key']].get('BaseQuantity', 0) * c['count'] for c in contribs)
                final_qty = math.ceil(total_contrib_val * rules[code]['m'])
                if final_qty > 0:
                    data_by_client[client_id][(code, 'Guias e Montantes', False)] = {
                        'Tipo Code': code, 'Descricao': price_data.get(code, {}).get('Descricao', f'Montante {code}'), 'BaseQuantity': final_qty, 'Categoria': 'Guias e Montantes',
                        'has_insulation': False, 'insulation_items': {}, 'carenagem_items': {}, 'is_subclass': False,
                        'formula_contributors': contribs, 'formula_multiplier': rules[code]['m']
                    }
    return data_by_client


def remove_guias_e_montantes(data_by_client):
    """Remove todos os itens de categoria 'Guias e Montantes' do conjunto de dados.

    Isso permite gerar orçamentos sem incluir essa categoria quando o usuário
    opta por não adicioná-la.
    """
    for client_id, client_items in data_by_client.items():
        keys_to_remove = [k for k, v in client_items.items() if v.get('Categoria') == 'Guias e Montantes']
        for k in keys_to_remove:
            del client_items[k]
    return data_by_client

# ==============================================================================
# BLOCO DE EXECUÇÃO PRINCIPAL
# ==============================================================================

def write_relacao_media_material_sheet(sheet, bold_font, regular_font, currency_format, bold_white_font, regular_white_font):
    # Definir larguras das colunas
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 10
    sheet.column_dimensions['J'].width = 15
    sheet.column_dimensions['K'].width = 15
    sheet.column_dimensions['L'].width = 15
    sheet.column_dimensions['M'].width = 15
    sheet.column_dimensions['N'].width = 18 # Ajuste para 'Fornecedor'

    # Definir as cores de preenchimento
    fill_ccccff = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
    fill_ffff00 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Definir os cabeçalhos e suas cores
    headers_part1 = [('Códigos', 1), ('Itens', 2), ('Classe', 3)] # Colunas 1, 2, 3
    headers_part2 = [('Qtds NOVO LAYOUT', 4), ('Qtds DISTRATO', 5), ('Qtd Final', 6), ('Perda', 7)] # Colunas 4, 5, 6, 7
    headers_part3 = [('Total', 8), ('Ud', 9), ('Custo Unitário', 10), ('Custo Total', 11), ('Valor Venda Direta', 12), ('VD Total', 13), ('Fornecedor', 14)] # Colunas 8, 9, 10, 11, 12, 13, 14

    current_row = 1 # Linha do cabeçalho
    
    # Escrever e formatar a Parte 1 dos cabeçalhos
    for col_name, col_idx in headers_part1:
        cell = sheet.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = bold_font
        cell.fill = fill_ccccff
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Escrever e formatar a Parte 2 dos cabeçalhos
    for col_name, col_idx in headers_part2:
        cell = sheet.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = bold_font
        cell.fill = fill_ffff00
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Escrever e formatar a Parte 3 dos cabeçalhos
    for col_name, col_idx in headers_part3:
        cell = sheet.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = bold_font
        cell.fill = fill_ccccff
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Aplicar bordas ao cabeçalho (apenas a primeira linha, colunas 1 a 14)
    apply_borders_to_range(sheet, current_row, 1, current_row, 14)

    # Aplicar bordas às 46 linhas de dados (da linha 2 à linha 47, colunas 1 a 14)
    # Note que o preenchimento será aplicado apenas no cabeçalho devido à lógica acima.
    apply_borders_to_range(sheet, current_row + 1, 1, current_row + 46, 14)

    # Adicionar preenchimento nas colunas D, E, F e G para as 46 linhas de dados
    fill_ffcc99 = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    red_font = Font(color="FF0000", name='Verdana', size=8) # Fonte vermelha
    for r_idx in range(current_row + 1, current_row + 46 + 1): # Linhas 2 a 47
        # Aplicar preenchimento
        for c_idx in range(4, 7 + 1): # Colunas D (4) a G (7)
            sheet.cell(row=r_idx, column=c_idx).fill = fill_ffcc99
        
        # Inserir fórmulas e valor fixo
        # Coluna F: Qtd Final
        sheet.cell(row=r_idx, column=6).value = f'=CEILING(IF(E{r_idx}=0,ABS(E{r_idx}-D{r_idx}),E{r_idx}-D{r_idx})*G{r_idx},1)'
        
        # Coluna G: Perda (valor 1 com fonte vermelha)
        cell_g = sheet.cell(row=r_idx, column=7, value=1)
        cell_g.font = red_font
        
        # Coluna H: Total
        sheet.cell(row=r_idx, column=8).value = f'=CEILING(IF(E{r_idx}<1,ABS(F{r_idx}),E{r_idx})*G{r_idx},1)'
        
        # Coluna K: Custo Total
        sheet.cell(row=r_idx, column=11).value = f'=IFERROR($H{r_idx}*J{r_idx},"-")'
        
        # Coluna M: VD Total
        sheet.cell(row=r_idx, column=13).value = f'=IFERROR($H{r_idx}*L{r_idx},"-")'

    # Adicionar linha de soma total para as colunas K e M
    sum_row = current_row + 47 # Linha 48 (1 para cabeçalho + 46 para dados + 1 para soma)
    fill_ffff00 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Célula de soma para a coluna K
    sum_k_cell = sheet.cell(row=sum_row, column=11, value=f'=SUM(K{current_row + 1}:K{current_row + 46})')
    sum_k_cell.fill = fill_ffff00
    sum_k_cell.font = bold_font
    sum_k_cell.number_format = currency_format
    sum_k_cell.alignment = Alignment(horizontal='right', vertical='center') # Alinhamento

    # Célula de soma para a coluna M
    sum_m_cell = sheet.cell(row=sum_row, column=13, value=f'=SUM(M{current_row + 1}:M{current_row + 46})')
    sum_m_cell.fill = fill_ffff00
    sum_m_cell.font = bold_font
    sum_m_cell.number_format = currency_format
    sum_m_cell.alignment = Alignment(horizontal='right', vertical='center') # Alinhamento
    
    # Aplicar bordas à linha de soma (apenas colunas K e M)
    apply_borders_to_range(sheet, sum_row, 11, sum_row, 11) # Borda para K
    apply_borders_to_range(sheet, sum_row, 13, sum_row, 13) # Borda para M

if __name__ == '__main__':
    print("Carregando dados de preços e subclasses...")
    precos = load_price_data()
    subclasses = load_subclass_data()

    # Ler observações gerais do arquivo
    try:
        with open('OBSERVACOES_GERAIS.txt', 'r', encoding='utf-8') as f:
            observacoes_gerais_content = f.read()
    except FileNotFoundError:
        print("AVISO: Arquivo 'OBSERVACOES_GERAIS.txt' não encontrado. As observações gerais não serão incluídas.")
        observacoes_gerais_content = "" # Conteúdo vazio se o arquivo não for encontrado
    except Exception as e:
        print(f"ERRO ao ler 'OBSERVACOES_GERAIS.txt': {e}")
        observacoes_gerais_content = ""

    # Perguntar ao usuário se deve incluir Guias e Montantes no orçamento
    resposta = input("Orçamento com Guias e Montantes? (s/n): ").strip().lower()
    include_guias_montantes = not resposta.startswith('n')
    print(f"Incluir Guias e Montantes: {'SIM' if include_guias_montantes else 'NÃO'}")

    # Definir os arquivos Excel e os nomes das abas
    excel_file_add = find_latest_excel_file('ADD_')
    excel_file_dis = find_latest_excel_file('DIS_')

    if not excel_file_add or not excel_file_dis:
        print("Erro: Não foi possível encontrar todos os arquivos Excel necessários. Abortando.")
        exit() # Terminate script if files are not found

    sheet_forro = '1.__Tabela de Forro'
    sheet_generico = '2.__Tabela de Modelo Genérico'
    sheet_paredes = '3.__Tabela de Paredes'

    print("Processando dados para SERVIÇOS ADICIONAIS (NOVO LAYOUT)...")
    dados_cliente_normal = process_client_data(
        precos, subclasses,
        excel_file_add,
        sheet_forro,
        sheet_generico,
        sheet_paredes
    )
    print("  Aplicando lógicas de negócio...")
    dados_cliente_normal = apply_wool_logic(dados_cliente_normal)
    if include_guias_montantes:
        dados_cliente_normal = calculate_and_add_derived_items(dados_cliente_normal, precos)
    else:
        dados_cliente_normal = remove_guias_e_montantes(dados_cliente_normal)
    dados_resumo_normal = process_summary_data(dados_cliente_normal)

    print("Processando dados para SERVIÇOS DISTRATADOS...")
    dados_cliente_distratado = process_client_data(
        precos, subclasses,
        excel_file_dis,
        sheet_forro,
        sheet_generico,
        sheet_paredes
    )
    print("  Aplicando lógicas de negócio...")
    dados_cliente_distratado = apply_wool_logic(dados_cliente_distratado)
    if include_guias_montantes:
        dados_cliente_distratado = calculate_and_add_derived_items(dados_cliente_distratado, precos)
    else:
        dados_cliente_distratado = remove_guias_e_montantes(dados_cliente_distratado)
    dados_resumo_distratado = process_summary_data(dados_cliente_distratado)

    print("Gerando arquivo Excel...")
    write_excel_with_formulas(
        dados_resumo_normal, dados_resumo_distratado,
        dados_cliente_normal, dados_cliente_distratado,
        precos,
        filename='Relatorios_Com_Formulas.xlsx',
            observacoes_gerais_content=observacoes_gerais_content
        )
