import openpyxl
wb=openpyxl.load_workbook('Relatorios_Com_Formulas_test.xlsx',data_only=True)
print('sheets', wb.sheetnames)
ws=wb['Aditivos x Distrato']
print('Aditivos header', ws['A1'].value, ws['H1'].value)
